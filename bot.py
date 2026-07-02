import discord
import os
import sys
import glob
import re
import time
import tempfile
import aiohttp
import json
import asyncio
from datetime import datetime, timezone
from typing import Optional, Union
from discord import app_commands
from openai import OpenAI, RateLimitError, APIStatusError, APIConnectionError, APITimeoutError
from dotenv import load_dotenv
import base64

# Keep emoji prints from crashing a non-UTF-8 console (e.g. Windows cp1252 when
# stdout is piped/redirected) — same guard as scrape_site.py / scrape_knowledge.py.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

load_dotenv()

# ── Config ───────────────────────────────────────────────────────────────────
DISCORD_TOKEN  = os.getenv("DISCORD_TOKEN")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# ── Optional fallback AI provider ─────────────────────────────────────────────
# Any OpenAI-compatible API works as a backup (Groq, OpenRouter, DeepSeek,
# Together, a second OpenAI key, a local Ollama, etc.). When the primary
# provider runs out of quota or gets rate-limited (HTTP 429 / insufficient_quota),
# chat completions automatically retry here so the bot keeps answering instead
# of going dark. Leave these unset to disable failover — the bot then shows a
# clean "AI temporarily unavailable" notice instead of a raw error.
#   FALLBACK_API_KEY   – the backup provider's API key
#   FALLBACK_BASE_URL  – its OpenAI-compatible endpoint, e.g. https://api.groq.com/openai/v1
#   FALLBACK_MODEL     – model for normal replies   (default: llama-3.3-70b-versatile)
#   FALLBACK_MINI_MODEL– model for the cheap classifier (defaults to FALLBACK_MODEL)
#   FALLBACK2_*        – an OPTIONAL second fallback (same four keys with a "2"
#     suffix), tried only if FALLBACK also fails. Providers run as a chain:
#     primary → FALLBACK → FALLBACK2, so the bot goes dark only when every one is
#     down. Gemini works here as an OpenAI-compatible provider — base_url
#     https://generativelanguage.googleapis.com/v1beta/openai/ with a 2.5-flash
#     model; its thinking is auto-disabled per request (see _call_fallback_provider).
#   FALLBACK_MAX_PROMPT_CHARS – cap on the chars sent to the fallback so the full
#     knowledge base (~32k tokens) doesn't blow past a free-tier per-minute token
#     limit (e.g. Groq 70b = 12k TPM → otherwise a hard 413 on every reply).
FALLBACK_API_KEY    = os.getenv("FALLBACK_API_KEY")
FALLBACK_BASE_URL   = os.getenv("FALLBACK_BASE_URL")
FALLBACK_MODEL      = os.getenv("FALLBACK_MODEL", "llama-3.3-70b-versatile")
FALLBACK_MINI_MODEL = os.getenv("FALLBACK_MINI_MODEL", FALLBACK_MODEL)

FALLBACK2_API_KEY    = os.getenv("FALLBACK2_API_KEY")
FALLBACK2_BASE_URL   = os.getenv("FALLBACK2_BASE_URL")
FALLBACK2_MODEL      = os.getenv("FALLBACK2_MODEL", "llama-3.3-70b-versatile")
FALLBACK2_MINI_MODEL = os.getenv("FALLBACK2_MINI_MODEL", FALLBACK2_MODEL)

FALLBACK_MAX_PROMPT_CHARS = int(os.getenv("FALLBACK_MAX_PROMPT_CHARS", "28000"))

# Marks where the bulky knowledge dump begins in build_system_prompt(). Everything
# ABOVE it (rules header + live events) is authoritative and must survive fallback
# truncation; only the knowledge below it gets trimmed. Single source of truth so
# the prompt builder and the truncator can never drift apart.
_KNOWLEDGE_MARKER = "=== AFC KNOWLEDGE BASE ==="

ALLOWED_CHANNELS = [
    920726991089598476,
    1327968058148524133,
    1014588126422904873,
    946321672015851570,
    1079786358840766554,
    1306928470802042931,
    1011289377055449178,
    1324442579265388644,
    1092544100072423435,
    920795335272579102,
    953326236950757446,
    1340452836495851713,
    955773076786798643,
]

# Category ID — bot listens and replies in ALL channels under this category
ALLOWED_CATEGORIES = {1074466672909496420}

# Channels where the bot auto-replies to ALL messages (not just @mentions)
# Add the channel IDs where users ask questions (e.g. support, general, help)
AUTO_REPLY_CHANNELS = [
    1026913984923840542,  # support channel
    920726991089598476,   # general/main
    953326236950757446,   # add more as needed
]

# Roles allowed to use the announcement command
ANNOUNCE_ROLES = [
    920732760094703746,
    920732112238284871,
]

# Support channel and roles
SUPPORT_CHANNEL_ID = 1026913984923840542
SUPPORT_ROLES = [
    920732112238284871,
    920734111772057602,
    920734300222128198,
    920732760094703746,
]

# ── User-facing notices when the AI backend is unavailable ────────────────────
# Shown INSTEAD of a raw OpenAI exception so users never see "Error: 429 ...
# insufficient_quota ... check your plan and billing" or a billing link. The real
# error is still printed to stdout/logs for ops.
AI_DOWN_NOTICE = (
    "🛠️ Heads up — AFC Bot's AI assistant is temporarily unavailable, so I can't "
    "auto-answer right now. Please try again a little later.\n"
    f"If it's urgent, reach the team in <#{SUPPORT_CHANNEL_ID}> or email "
    "**info@africanfreefirecommunity.com**."
)
GENERIC_ERROR_NOTICE = (
    "⚠️ Something went wrong on my end. Please try again shortly — and if it keeps "
    f"happening, reach the team in <#{SUPPORT_CHANNEL_ID}> or at "
    "**info@africanfreefirecommunity.com**."
)
# Don't spam an identical "AI is down" notice on every message while quota is dead.
# {channel_id: last_notice_unix_ts}
_ai_down_notice_at: dict[int, float] = {}
AI_DOWN_NOTICE_COOLDOWN_SECS = 300  # at most one down-notice per channel / 5 min

# Scrims Master role — gets access to staff/backend knowledge
SCRIMS_MASTER_ROLE_ID = 1011438178630107207

# Roles that can access staff-level knowledge (backend ops, scoring, scrims admin info)
STAFF_KNOWLEDGE_ROLES = set(ANNOUNCE_ROLES + SUPPORT_ROLES + [SCRIMS_MASTER_ROLE_ID])

# Channel where unanswered/escalated questions go
MODERATION_SUPPORT_CHANNEL_ID = 1026913984923840542

# Channel where auto-generated news announcements are posted
NEWS_ANNOUNCEMENT_CHANNEL_ID = 1306247327840731157

# How often to auto-refresh the knowledge base (hours). 4h balances fresh team
# roster / new pages against load on the flaky teams API; the fast-changing data
# (tournaments/news/bans) is already near-instant via the 120s poll loops below.
SCRAPE_INTERVAL_HOURS = 4

# How often to poll the backend for new events/bans/news (seconds)
NEWS_POLL_INTERVAL_SECS    = 120   # every 2 minutes
EVENT_POLL_INTERVAL_SECS   = 120
BAN_POLL_INTERVAL_SECS     = 60    # every 1 minute

# Channel for tournament announcements
TOURNAMENT_ANNOUNCEMENT_CHANNEL_ID = 955773076786798643
# Channel for scrim announcements + role to ping
SCRIM_ANNOUNCEMENT_CHANNEL_ID      = 1487971199454679050
SCRIMS_PING_ROLE_ID                = 1395722795878584391
# Channel where admins announce bans (player or team)
BAN_ANNOUNCEMENT_CHANNEL_ID   = 1317799517084454932
# Channel where admins announce unbans
UNBAN_ANNOUNCEMENT_CHANNEL_ID = 1353759565543637062

# AFC backend API base
AFC_API_BASE = "https://api.africanfreefirecommunity.com"

# Canonical AFC Discord invite — the ONLY Discord link the bot is allowed to share.
# Hardcoded so GPT can never hallucinate a wrong invite (e.g. "discord.gg/afc").
AFC_DISCORD_INVITE = "https://discord.gg/qgKKZMu4sA"

# Always use the folder where bot.py lives — avoids permission errors on Windows
BASE_DIR              = os.path.dirname(os.path.abspath(__file__))
KNOWLEDGE_DIR         = os.path.join(BASE_DIR, "knowledge")
STAFF_KNOWLEDGE_DIR   = os.path.join(BASE_DIR, "knowledge_staff")
HISTORY_FILE          = os.path.join(BASE_DIR, "conversation_history.json")
BASE_KNOWLEDGE        = os.path.join(BASE_DIR, "knowledge_base.txt")

# Files to persist poll state across restarts
SEEN_NEWS_FILE          = os.path.join(BASE_DIR, "seen_news.json")
SEEN_EVENTS_FILE        = os.path.join(BASE_DIR, "seen_events.json")
SEEN_BAN_ACTIVITIES_FILE = os.path.join(BASE_DIR, "seen_ban_activities.json")

# Organizer-event approval gate — persisted state
PENDING_EVENT_APPROVALS_FILE = os.path.join(BASE_DIR, "pending_event_approvals.json")
REJECTED_EVENT_IDS_FILE      = os.path.join(BASE_DIR, "rejected_event_ids.json")

MAX_HISTORY      = 30
HISTORY_TTL_SECS = 24 * 60 * 60   # 24 hours in seconds

# ── Transcription config ──────────────────────────────────────────────────────
# Channel where the bot asks about stage transcription
MODS_CHANNEL_ID = 1324442579265388644
# Roles allowed to trigger/stop transcription (mods + support)
TRANSCRIPTION_ROLES = set(ANNOUNCE_ROLES + SUPPORT_ROLES)
# End a transcription session automatically after this much total silence and
# post whatever was captured so far.
TRANSCRIPTION_SILENCE_TIMEOUT_SECS = 300  # 5 minutes

# Roles allowed to approve/reject organizer-event announcements (mirrors transcription perms)
EVENT_APPROVAL_ROLES = set(ANNOUNCE_ROLES + SUPPORT_ROLES)

# Supported media types
IMAGE_TYPES = (".png", ".jpg", ".jpeg", ".gif", ".webp")
AUDIO_TYPES = (".mp3", ".mp4", ".wav", ".m4a", ".ogg", ".webm", ".flac")
VIDEO_TYPES = (".mov", ".avi", ".mkv")
# ─────────────────────────────────────────────────────────────────────────────

# Bounded timeout + single retry — openai-python's defaults (600s read timeout,
# 2 retries) would let one hung request stall a reply for up to ~30 minutes.
client_ai = OpenAI(api_key=OPENAI_API_KEY, timeout=60.0, max_retries=1)
# Backup AI provider chain — each entry built only when its env vars are set.
# Requests fail over through this list in order (FALLBACK first, then FALLBACK2),
# so the bot only goes dark when the primary AND every configured fallback is down.
def _make_fallback_provider(api_key, base_url, model, mini_model):
    if not (api_key and base_url):
        return None
    return {
        "client": OpenAI(api_key=api_key, base_url=base_url, timeout=60.0, max_retries=1),
        "model": model,
        "mini_model": mini_model,
        "base_url": base_url,
        # Gemini 2.5 flash are "thinking" models; flagged so _call_fallback_provider
        # disables reasoning per request (else a small max_tokens yields no text).
        "is_gemini": "generativelanguage.googleapis.com" in base_url,
    }

FALLBACK_PROVIDERS = [p for p in (
    _make_fallback_provider(FALLBACK_API_KEY,  FALLBACK_BASE_URL,  FALLBACK_MODEL,  FALLBACK_MINI_MODEL),
    _make_fallback_provider(FALLBACK2_API_KEY, FALLBACK2_BASE_URL, FALLBACK2_MODEL, FALLBACK2_MINI_MODEL),
) if p]
for _p in FALLBACK_PROVIDERS:
    print(f"✅ Fallback AI provider configured ({_p['base_url']}, model={_p['model']})")

intents = discord.Intents.default()
intents.message_content = True
# Required for guild.chunk() / guild.members iteration (mass role actions,
# purge-by-role, single-user role lookups). The privileged "Server Members
# Intent" toggle must be ON in the Discord Developer Portal (same page as the
# Message Content Intent) or the bot will refuse to start.
intents.members = True
bot = discord.Client(intents=intents)
# Application (slash) command tree — commands defined in the "Slash commands"
# section below and synced per guild in on_ready.
tree = app_commands.CommandTree(bot)

# In-memory history — loaded from file on startup
# Structure: { "channel_id": { "messages": [...], "last_updated": <unix timestamp> } }
history: dict[str, dict] = {}

# Cached live event data — refreshed every EVENT_POLL_INTERVAL_SECS by event_poll_loop
_cached_events: list[dict] = []
# File to persist event statuses for tracking changes across restarts
SEEN_EVENT_STATUSES_FILE = os.path.join(BASE_DIR, "seen_event_statuses.json")


# ── Persistent history helpers ───────────────────────────────────────────────
def load_history_from_disk():
    """Load conversation history from file on startup."""
    global history
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                history = json.load(f)
            print(f"📂  Loaded history for {len(history)} channel(s) from disk.")
        except Exception as e:
            print(f"⚠️  Could not load history file: {e}. Starting fresh.")
            history = {}
    else:
        history = {}


def save_history_to_disk():
    """Save current history to file."""
    try:
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️  Could not save history: {e}")


def purge_expired_history():
    """Remove channels whose history is older than 24 hours."""
    now = datetime.now(timezone.utc).timestamp()
    expired = [
        cid for cid, data in history.items()
        if now - data.get("last_updated", 0) > HISTORY_TTL_SECS
    ]
    for cid in expired:
        del history[cid]
    if expired:
        print(f"🧹  Purged expired history for {len(expired)} channel(s).")
        save_history_to_disk()


def get_channel_messages(channel_id: int) -> list:
    """Get the message list for a channel, or create it."""
    cid = str(channel_id)
    if cid not in history:
        history[cid] = {"messages": [], "last_updated": datetime.now(timezone.utc).timestamp()}
    return history[cid]["messages"]


def touch_channel(channel_id: int):
    """Update the last_updated timestamp for a channel."""
    cid = str(channel_id)
    if cid in history:
        history[cid]["last_updated"] = datetime.now(timezone.utc).timestamp()


def trim_channel_history(channel_id: int):
    """Keep only the last MAX_HISTORY messages for a channel. Trims IN PLACE —
    rebinding the dict value would orphan any local reference captured before the
    trim (ask_openai_text holds one across the AI call), silently dropping every
    assistant reply appended after a trim once a channel reaches MAX_HISTORY."""
    msgs = get_channel_messages(channel_id)
    if len(msgs) > MAX_HISTORY:
        msgs[:] = msgs[-MAX_HISTORY:]


async def auto_purge_loop():
    """Background task — checks and purges expired history every hour."""
    await bot.wait_until_ready()
    while not bot.is_closed():
        purge_expired_history()
        await asyncio.sleep(3600)   # run every hour


async def keep_typing(channel: discord.abc.Messageable, stop_event: asyncio.Event):
    """Re-send typing indicator every 8 seconds so it never expires during long AI calls.
    (discord.py 2.x: awaiting channel.typing() sends one ~10s indicator; the old
    trigger_typing() was removed in 2.0 and silently AttributeError'd forever.)"""
    while not stop_event.is_set():
        try:
            await channel.typing()
        except Exception:
            pass
        await asyncio.sleep(8)


async def _reply_chunked(message: discord.Message, text: str, mention_author: bool = True):
    """Reply within Discord's 2000-char content limit, splitting long text across
    follow-up messages instead of failing the whole send with a 400. Returns the
    first sent message (for last-bot-message tracking)."""
    first = None
    for i in range(0, len(text), 2000):
        chunk = text[i:i + 2000]
        if first is None:
            first = await message.reply(chunk, mention_author=mention_author)
        else:
            await message.channel.send(chunk)
    return first


def _do_scrape() -> int:
    """Synchronous scrape — runs in a thread executor. Returns total characters written.

    Delegates to afc_scraper (the single source of truth shared with
    scripts/scrape_knowledge.py and scrape_site.py) so the three never drift. That
    module crawls the site, drops client-render shells (/tournaments, /teams, /news
    are Next.js client-rendered and would otherwise capture only "Loading..."), and
    appends the live AFC teams directory from the API."""
    import afc_scraper
    dest = os.path.join(BASE_DIR, "knowledge_base.txt")
    return afc_scraper.write_knowledge_base(dest)


async def refresh_knowledge_base() -> int:
    """Run the website scraper in a thread so the event loop isn't blocked."""
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _do_scrape)


async def auto_scrape_loop():
    """Background task — re-scrapes the AFC website every SCRAPE_INTERVAL_HOURS hours."""
    await bot.wait_until_ready()
    await asyncio.sleep(SCRAPE_INTERVAL_HOURS * 3600)   # skip first run (already fresh on startup)
    while not bot.is_closed():
        try:
            chars = await refresh_knowledge_base()
            print(f"🔄  Auto-scrape complete — {chars:,} chars written to knowledge_base.txt")
        except Exception as e:
            print(f"⚠️  Auto-scrape failed: {e}")
        await asyncio.sleep(SCRAPE_INTERVAL_HOURS * 3600)


def load_seen_news() -> set:
    """Load the set of already-announced news IDs from disk."""
    if os.path.exists(SEEN_NEWS_FILE):
        try:
            with open(SEEN_NEWS_FILE, "r", encoding="utf-8") as f:
                return set(json.load(f))
        except Exception:
            pass
    return set()


def save_seen_news(seen: set):
    """Persist the set of announced news IDs to disk."""
    try:
        with open(SEEN_NEWS_FILE, "w", encoding="utf-8") as f:
            json.dump(list(seen), f)
    except Exception as e:
        print(f"⚠️  Could not save seen_news.json: {e}")


async def fetch_all_news() -> list | None:
    """Call the AFC API and return the list of news article dicts, newest first.
    Returns None when the API is unavailable — callers must distinguish 'API
    down' from 'genuinely zero articles' or a boot-time outage would seed an
    empty seen-set and @everyone-spam the whole backlog on recovery."""
    url = f"{AFC_API_BASE}/auth/get-all-news/"
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=10)) as resp:
                if resp.status != 200:
                    return None
                data = await resp.json()
                return data.get("news", [])
    except Exception as e:
        print(f"⚠️  fetch_all_news failed: {e}")
        return None


async def generate_news_embed(article: dict) -> discord.Embed:
    """Use GPT-4o to write a teaser snippet, then build the Discord embed."""
    slug = article.get("slug", "")
    article_url = f"https://africanfreefirecommunity.com/news/{slug}" if slug else "https://africanfreefirecommunity.com/news"
    title = article.get("news_title", "New Article")
    content = (article.get("content") or "")[:1500]   # trim to avoid token overflow
    category = article.get("category", "")
    related_event = article.get("related_event")
    image_url = article.get("images_url")

    prompt = f"""You are the AFC Bot news writer for the African Freefire Community.

A new article has just been published:
Title: {title}
Category: {category}
{f'Related event: {related_event}' if related_event else ''}
Article content (may be truncated):
{content or '(No content available)'}

Write a short Discord announcement to hype up this article.
Rules:
- Body: 2-4 sentences only — tease the reader, do NOT reveal everything
- End the body with: **Read the full story → {article_url}**
- Use 1-2 emojis that fit the topic
- Tone: exciting, community-first, never corporate

Output ONLY valid JSON (no markdown fences):
{{"body": "..."}}
"""
    try:
        response = await _achat(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=300,
            temperature=0.8,
        )
        raw = re.sub(r"```json|```", "", response.choices[0].message.content.strip()).strip()
        body = json.loads(raw).get("body", "")
    except Exception:
        body = f"A new article has just dropped! 🔥\n\n**Read it here → {article_url}**"

    embed = discord.Embed(
        title=title,
        description=body,
        color=0x00A550,
        url=article_url,
    )
    if image_url:
        embed.set_image(url=image_url)
    author_name = article.get("author", "AFC")
    embed.set_author(name=f"AFC News — {category}" if category else "AFC News")
    embed.set_footer(text="African Freefire Community  •  africanfreefirecommunity.com")
    embed.timestamp = datetime.now(timezone.utc)
    return embed


async def news_poll_loop():
    """Background task — polls the AFC API every NEWS_POLL_INTERVAL_SECS for new articles."""
    await bot.wait_until_ready()

    seen = load_seen_news()
    # File existence is the durable "already seeded" sentinel (same rationale as
    # the ban loop): a failed fetch at boot must not seed an empty set — the
    # whole backlog would @everyone-spam once the API recovers — and a
    # legitimately-empty seed must not re-seed on the next restart.
    seeded = os.path.exists(SEEN_NEWS_FILE)

    while not bot.is_closed():
        await asyncio.sleep(NEWS_POLL_INTERVAL_SECS)
        try:
            articles = await fetch_all_news()
            if articles is None:
                # API unavailable this cycle — skip without touching state.
                continue

            if not seeded:
                # First successful poll — mark existing articles as already seen
                # so we don't flood the channel with old news.
                seen = {str(a["news_id"]) for a in articles}
                save_seen_news(seen)
                seeded = True
                print(f"📰  News poll: seeded {len(seen)} existing article(s) on first successful poll. Watching for new ones.")
                continue

            new_articles = [a for a in articles if str(a["news_id"]) not in seen]

            if new_articles:
                news_channel = bot.get_channel(NEWS_ANNOUNCEMENT_CHANNEL_ID)
                if not news_channel:
                    news_channel = await bot.fetch_channel(NEWS_ANNOUNCEMENT_CHANNEL_ID)

                # Post oldest-first so they appear in chronological order
                for article in reversed(new_articles):
                    try:
                        embed = await generate_news_embed(article)
                        await news_channel.send(content="@everyone", embed=embed)
                        seen.add(str(article["news_id"]))
                        # Persist per send — a crash/deploy mid-batch must not
                        # re-announce already-posted articles with @everyone.
                        save_seen_news(seen)
                        print(f"📰  Announced news: {article.get('news_title', article['news_id'])}")
                        await asyncio.sleep(2)   # small gap between multiple posts
                    except Exception as e:
                        print(f"⚠️  Failed to post news {article.get('news_id')}: {e}")

        except Exception as e:
            print(f"⚠️  news_poll_loop error: {e}")


# ── Event polling ─────────────────────────────────────────────────────────────

def load_seen_events() -> set:
    if os.path.exists(SEEN_EVENTS_FILE):
        try:
            with open(SEEN_EVENTS_FILE, "r", encoding="utf-8") as f:
                return set(json.load(f))
        except Exception:
            pass
    return set()


def save_seen_events(seen: set):
    try:
        with open(SEEN_EVENTS_FILE, "w", encoding="utf-8") as f:
            json.dump(list(seen), f)
    except Exception as e:
        print(f"⚠️  Could not save seen_events.json: {e}")


# ── Organizer-event approval state ────────────────────────────────────────────
# message_id(str) -> event dict awaiting an admin's approval in the mods channel.
_pending_event_approvals: dict[str, dict] = {}
# event_id(str) set — organizer events an admin rejected; never auto-posted again.
_rejected_event_ids: set[str] = set()


def load_pending_event_approvals() -> dict:
    if os.path.exists(PENDING_EVENT_APPROVALS_FILE):
        try:
            with open(PENDING_EVENT_APPROVALS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    return {str(k): v for k, v in data.items()}
        except Exception as e:
            print(f"⚠️  Could not load pending_event_approvals.json: {e}")
    return {}


def save_pending_event_approvals(raise_on_error: bool = False):
    try:
        with open(PENDING_EVENT_APPROVALS_FILE, "w", encoding="utf-8") as f:
            json.dump(_pending_event_approvals, f)
    except Exception as e:
        print(f"⚠️  Could not save pending_event_approvals.json: {e}")
        if raise_on_error:
            raise


def load_rejected_event_ids() -> set:
    if os.path.exists(REJECTED_EVENT_IDS_FILE):
        try:
            with open(REJECTED_EVENT_IDS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list):
                    return {str(x) for x in data}
        except Exception as e:
            print(f"⚠️  Could not load rejected_event_ids.json: {e}")
    return set()


def save_rejected_event_ids():
    try:
        with open(REJECTED_EVENT_IDS_FILE, "w", encoding="utf-8") as f:
            json.dump(sorted(_rejected_event_ids), f)
    except Exception as e:
        print(f"⚠️  Could not save rejected_event_ids.json: {e}")


def _pending_event_ids() -> set:
    """event_ids currently awaiting approval (derived from pending payloads)."""
    return {str(ev.get("event_id")) for ev in _pending_event_approvals.values()}


def is_organizer_event(event: dict) -> bool:
    """True if a partner organization created this event (vs an AFC-run event).

    AFC-run events carry no organization_name (build_event_embed falls back to
    'African Freefire Community'); a non-empty org name that isn't AFC itself
    marks a partner-organizer event that must be approved before announcing.
    """
    org = (event.get("organization_name") or "").strip()
    if not org:
        return False
    return org.lower() != "african freefire community"


# event_type ('internal'/'external') exists ONLY on the get-event-details
# endpoint — the get-all-events list payload has no such field (verified against
# the live API), so it must be resolved per event. Cached in-memory per event_id;
# the cache clears on restart, so a later external→internal flip is picked up on
# the next deploy/restart at worst.
_event_type_cache: dict[str, str] = {}


async def get_event_type(event: dict) -> str:
    """Resolve an event's type ('internal'/'external'); returns '' when unknown
    (backend unreachable) so callers can fail closed and retry next poll."""
    etype = (event.get("event_type") or "").strip().lower()
    if etype:
        return etype
    eid = str(event.get("event_id"))
    if eid in _event_type_cache:
        return _event_type_cache[eid]
    slug = event.get("slug")
    if not slug:
        return ""
    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(
                f"{AFC_API_BASE}/events/get-event-details/",
                json={"slug": slug},
                timeout=aiohttp.ClientTimeout(total=10),
            ) as resp:
                if resp.status != 200:
                    return ""
                data = await resp.json()
        etype = ((data.get("event_details") or {}).get("event_type") or "").strip().lower()
        if etype:
            _event_type_cache[eid] = etype
        return etype
    except Exception as e:
        print(f"⚠️  Could not fetch event_type for {slug}: {e}")
        return ""


async def fetch_all_events() -> list | None:
    """Uses the existing public events endpoint — no backend changes needed.
    Returns None when the API is unavailable so callers can distinguish 'API
    down' from 'zero events' (keeps the last good cache, prevents bad seeding)."""
    url = f"{AFC_API_BASE}/events/get-all-events/"
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=10)) as resp:
                if resp.status != 200:
                    return None
                data = await resp.json()
                return data.get("events", [])
    except Exception as e:
        print(f"⚠️  fetch_all_events failed: {e}")
        return None


async def build_event_embed(event: dict) -> tuple[discord.Embed, str | None]:
    """Build the announcement embed for a tournament or scrim. Returns (embed, optional_ping)."""
    name       = event.get("event_name", "New Event")
    comp_type  = event.get("competition_type", "tournament")
    start_date = event.get("event_date", "TBD")
    prizepool  = event.get("prizepool", "")
    max_slots  = event.get("number_of_participants", "")
    registered = event.get("total_registered_competitors", 0)
    banner_url = event.get("event_banner")
    slug       = event.get("slug", "")
    # Organizer of the event. AFC-run events have no organization on the record,
    # so they're attributed to the community itself.
    organizer  = event.get("organization_name") or "African Freefire Community"
    event_url  = f"https://africanfreefirecommunity.com/tournaments/{slug}" if slug else "https://africanfreefirecommunity.com/tournaments"

    is_scrim = comp_type.lower() == "scrims"
    color    = 0xFFD700 if is_scrim else 0x00A550

    lines = []
    lines.append(f"🏢 **Organizer:** {organizer}")
    if start_date: lines.append(f"📅 **Date:** {start_date}")
    if prizepool:  lines.append(f"💰 **Prize Pool:** {prizepool}")
    if max_slots:  lines.append(f"👥 **Slots:** {registered}/{max_slots}")
    lines.append(f"\n🔗 **[View & Register →]({event_url})**")

    embed = discord.Embed(
        title=f"{'🎮 New Scrim' if is_scrim else '🏆 New Tournament'}: {name}",
        description="\n".join(lines),
        color=color,
        url=event_url,
    )
    if banner_url:
        embed.set_image(url=banner_url)
    embed.set_footer(text="African Freefire Community  •  africanfreefirecommunity.com")
    embed.timestamp = datetime.now(timezone.utc)

    ping = f"<@&{SCRIMS_PING_ROLE_ID}>" if is_scrim else None
    return embed, ping


async def announce_event_public(event: dict):
    """Post an event's announcement embed to its public tournament/scrim channel,
    with the same @everyone (tournament) / scrim-role ping rules used since launch."""
    embed, ping = await build_event_embed(event)
    is_scrim = event.get("competition_type", "").lower() == "scrims"
    ch_id = SCRIM_ANNOUNCEMENT_CHANNEL_ID if is_scrim else TOURNAMENT_ANNOUNCEMENT_CHANNEL_ID
    channel = bot.get_channel(ch_id) or await bot.fetch_channel(ch_id)
    everyone_ping = "@everyone" if not is_scrim else ""
    content = " ".join(filter(None, [everyone_ping, ping]))
    await channel.send(content=content, embed=embed)


class EventApprovalView(discord.ui.View):
    """Persistent Approve/Reject buttons for event previews in the mods channel.
    One instance is registered globally in on_ready; it resolves the pending
    event by the message id the buttons live on."""

    def __init__(self):
        super().__init__(timeout=None)

    # discord.py 2.x passes (interaction, button) — py-cord's (button, interaction)
    # order silently receives the Button as `interaction` and crashes before the
    # click is acknowledged, so Discord shows "This interaction failed".
    @discord.ui.button(label="✅ Approve", style=discord.ButtonStyle.success, custom_id="afc_event_approve")
    async def approve(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self._handle(interaction, approved=True)

    @discord.ui.button(label="❌ Reject", style=discord.ButtonStyle.danger, custom_id="afc_event_reject")
    async def reject(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self._handle(interaction, approved=False)

    async def _handle(self, interaction: discord.Interaction, approved: bool):
        member = interaction.user
        roles = getattr(member, "roles", [])
        if not any(getattr(r, "id", None) in EVENT_APPROVAL_ROLES for r in roles):
            await interaction.response.send_message(
                "⛔ You're not authorized to approve event announcements.", ephemeral=True
            )
            return

        mid = str(interaction.message.id)
        event = _pending_event_approvals.pop(mid, None)
        if event is None:
            await interaction.response.send_message(
                "⚠️ This approval was already handled or has expired.", ephemeral=True
            )
            return
        save_pending_event_approvals()

        # Ack the click now — posting the public announcement below can take longer
        # than Discord's ~3s response window, which would otherwise expire the token
        # and leave the preview stuck with live buttons over an already-posted event.
        await interaction.response.defer()

        if approved:
            try:
                await announce_event_public(event)
            except Exception as e:
                _pending_event_approvals[mid] = event   # re-queue so approval isn't lost
                save_pending_event_approvals()
                await interaction.followup.send(
                    f"⚠️ Couldn't post the announcement: {e}", ephemeral=True
                )
                return
            note = f"✅ **Approved** by {member.mention} — announcement posted."
            print(f"✅  Event approved by {member}: {event.get('event_name')}")
        else:
            _rejected_event_ids.add(str(event.get("event_id")))
            save_rejected_event_ids()
            note = f"❌ **Rejected** by {member.mention} — not announced."
            print(f"❌  Event rejected by {member}: {event.get('event_name')}")

        # Rewrite the preview content and drop the buttons (interaction was deferred).
        await interaction.message.edit(content=note, view=None)


async def post_event_for_approval(event: dict):
    """Send an event to the mods channel for admin approval instead of
    announcing it publicly. Raises on failure so the poll loop can retry."""
    channel = bot.get_channel(MODS_CHANNEL_ID) or await bot.fetch_channel(MODS_CHANNEL_ID)
    embed, _ping = await build_event_embed(event)
    is_scrim  = event.get("competition_type", "").lower() == "scrims"
    target    = "scrim" if is_scrim else "tournament"
    organizer = event.get("organization_name") or "African Freefire Community"
    header = (
        f"🕓 **PENDING APPROVAL** — new {target} from **{organizer}**.\n"
        f"Approve to announce it publicly, or reject to discard."
    )
    msg = await channel.send(content=header, embed=embed, view=EventApprovalView())
    _pending_event_approvals[str(msg.id)] = event
    # Raise if we can't persist the pending record so the poll loop's except leaves
    # the event unseen and retries — rather than marking it seen with an in-memory-
    # only entry that a restart would lose, silently dropping the event for good.
    save_pending_event_approvals(raise_on_error=True)
    print(f"🕓  Event sent for approval: {event.get('event_name')} (msg {msg.id})")


def _load_event_statuses() -> dict:
    """Load {event_id: status} mapping from disk."""
    if os.path.exists(SEEN_EVENT_STATUSES_FILE):
        try:
            with open(SEEN_EVENT_STATUSES_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_event_statuses(statuses: dict):
    try:
        with open(SEEN_EVENT_STATUSES_FILE, "w", encoding="utf-8") as f:
            json.dump(statuses, f)
    except Exception as e:
        print(f"⚠️  Could not save event statuses: {e}")


def _build_status_change_embed(event: dict, old_status: str, new_status: str) -> discord.Embed:
    """Build an embed announcing a tournament/scrim status change."""
    name      = event.get("event_name", "Unknown Event")
    comp_type = event.get("competition_type", "tournament")
    slug      = event.get("slug", "")
    event_url = f"https://africanfreefirecommunity.com/tournaments/{slug}" if slug else ""

    is_scrim = comp_type.lower() == "scrims"

    # Choose icon and color based on new status
    new_lower = new_status.lower()
    if new_lower in ("live", "in_progress", "started", "ongoing"):
        icon, color = "🟢", 0x00FF00
        title = f"{'🎮 Scrim' if is_scrim else '🏆 Tournament'} NOW LIVE: {name}"
    elif new_lower in ("completed", "ended", "finished"):
        icon, color = "🏁", 0x888888
        title = f"{'🎮 Scrim' if is_scrim else '🏆 Tournament'} ENDED: {name}"
    elif new_lower in ("registration_closed", "closed"):
        icon, color = "🔒", 0xFF8800
        title = f"{'🎮 Scrim' if is_scrim else '🏆 Tournament'} Registration Closed: {name}"
    elif new_lower in ("cancelled", "canceled"):
        icon, color = "❌", 0xFF0000
        title = f"{'🎮 Scrim' if is_scrim else '🏆 Tournament'} CANCELLED: {name}"
    else:
        icon, color = "🔄", 0x3498DB
        title = f"{'🎮 Scrim' if is_scrim else '🏆 Tournament'} Update: {name}"

    organizer = event.get("organization_name") or "African Freefire Community"
    description = f"{icon} Status changed: **{old_status}** → **{new_status}**"
    description += f"\n🏢 Organizer: **{organizer}**"
    if event_url:
        description += f"\n\n🔗 **[View Event →]({event_url})**"

    embed = discord.Embed(title=title, description=description, color=color, url=event_url or None)
    banner = event.get("event_banner")
    if banner:
        embed.set_thumbnail(url=banner)
    embed.set_footer(text="African Freefire Community  •  africanfreefirecommunity.com")
    embed.timestamp = datetime.now(timezone.utc)
    return embed


async def event_poll_loop():
    """Background task — polls for new tournaments and scrims every EVENT_POLL_INTERVAL_SECS.
    Also caches event data for the system prompt and tracks status changes."""
    global _cached_events
    await bot.wait_until_ready()

    seen = load_seen_events()
    event_statuses = _load_event_statuses()
    # File existence is the durable "already seeded" sentinel (same rationale as
    # the ban loop): a failed fetch at boot must not seed an empty set — every
    # existing event would flood the mods channel with approval requests once
    # the API recovers — and a legitimately-empty seed must not re-seed later.
    seeded = os.path.exists(SEEN_EVENTS_FILE)

    while not bot.is_closed():
        await asyncio.sleep(EVENT_POLL_INTERVAL_SECS)
        try:
            events = await fetch_all_events()
            if events is None:
                # API unavailable this cycle — keep the last good cache (the
                # system prompt keeps answering from it) and touch no state.
                continue
            _cached_events = events  # Always refresh the cache for the system prompt

            if not seeded:
                # First successful poll — seed current events as already-seen so
                # the backlog isn't announced. Only INTERNAL, AFC-run events are
                # seeded: externals stay unseen so a flip to internal announces
                # like new, and organizer events stay unseen so the first real
                # poll routes them through the approval gate.
                for e in events:
                    if is_organizer_event(e):
                        continue
                    if (await get_event_type(e)) == "external":
                        continue
                    eid = str(e["event_id"])
                    seen.add(eid)
                    status = e.get("event_status", "")
                    if status:
                        event_statuses[eid] = status
                save_seen_events(seen)
                _save_event_statuses(event_statuses)
                seeded = True
                print(f"🎮  Event poll: seeded {len(seen)} existing event(s) on first successful poll. Watching for new ones.")
                continue

            # ── Announce NEW events ──
            pending_ids = _pending_event_ids()
            new_events = [
                e for e in events
                if str(e["event_id"]) not in seen
                and str(e["event_id"]) not in _rejected_event_ids
                and str(e["event_id"]) not in pending_ids
            ]

            for event in reversed(new_events):
                eid = str(event["event_id"])
                try:
                    # get-all-events carries no event_type — resolve it via the
                    # details endpoint (cached per event_id). External events are
                    # never announced; they stay unseen so a flip to internal
                    # announces like new. '' = backend couldn't say — skip WITHOUT
                    # marking seen so the event is retried next poll, never leaked.
                    etype = await get_event_type(event)
                    if etype == "external" or not etype:
                        continue
                    # Every new event — AFC-run/admin and partner-organizer alike —
                    # is held for admin approval in the mods channel; nothing is
                    # posted publicly until an admin approves it.
                    await post_event_for_approval(event)
                    seen.add(eid)
                    # Seed the status so we don't also fire a status-change for new events
                    event_statuses[eid] = event.get("event_status", "")
                    # Persist per event — a crash/deploy mid-batch must not
                    # re-request approval for already-posted events on restart.
                    save_seen_events(seen)
                    _save_event_statuses(event_statuses)
                    await asyncio.sleep(2)
                except Exception as e:
                    print(f"⚠️  Failed to handle new event {eid}: {e}")

            # ── Detect STATUS CHANGES on existing events ──
            statuses_changed = False
            for event in events:
                eid = str(event.get("event_id"))
                # Don't announce status changes on external events (resolved via
                # the details endpoint — the list payload has no event_type).
                # Track their status silently so a later flip to internal doesn't
                # announce stale news; on '' (backend unreachable) skip without
                # updating so the change is retried next poll (fail closed).
                etype = await get_event_type(event)
                if not etype:
                    continue
                if etype == "external":
                    ext_status = event.get("event_status", "")
                    if ext_status and event_statuses.get(eid) != ext_status:
                        event_statuses[eid] = ext_status
                        statuses_changed = True
                    continue
                # Rejected events are permanently suppressed — never announce.
                if eid in _rejected_event_ids:
                    continue
                # Events still awaiting approval must not announce a status change, but
                # keep the recorded status current so that, once approved, a flip that
                # happened while it was pending isn't re-announced as stale news.
                if eid in _pending_event_ids():
                    pending_status = event.get("event_status", "")
                    if pending_status and event_statuses.get(eid) != pending_status:
                        event_statuses[eid] = pending_status
                        statuses_changed = True
                    continue
                new_status = event.get("event_status", "")
                old_status = event_statuses.get(eid, "")

                if not new_status or new_status == old_status:
                    continue

                # First real status we've recorded for this event (e.g. it was
                # tracked with an empty status before event_status was read) —
                # seed it silently so we don't announce a change we never saw.
                if not old_status:
                    event_statuses[eid] = new_status
                    statuses_changed = True
                    continue

                # Status changed — announce it
                statuses_changed = True
                event_statuses[eid] = new_status
                try:
                    embed = _build_status_change_embed(event, old_status, new_status)
                    is_scrim = event.get("competition_type", "").lower() == "scrims"
                    ch_id = SCRIM_ANNOUNCEMENT_CHANNEL_ID if is_scrim else TOURNAMENT_ANNOUNCEMENT_CHANNEL_ID
                    channel = bot.get_channel(ch_id) or await bot.fetch_channel(ch_id)
                    await channel.send(embed=embed)
                    print(f"🔄  Event status change: {event.get('event_name')} → {old_status} → {new_status}")
                    await asyncio.sleep(2)
                except Exception as e:
                    print(f"⚠️  Failed to post status change for {event.get('event_name')}: {e}")

            if statuses_changed:
                _save_event_statuses(event_statuses)

        except Exception as e:
            print(f"⚠️  event_poll_loop error: {e}")


# ── Automatic ban / unban polling ────────────────────────────────────────────
# Polls GET /auth/get-admin-activities/ every BAN_POLL_INTERVAL_SECS.
# Detects new banned_team / unbanned_team / banned_player / unbanned_player
# entries and posts embeds automatically — no Discord command needed.

BAN_ACTIONS = {"banned_team", "unbanned_team", "banned_player", "unbanned_player"}


def load_seen_ban_activities() -> set:
    """Load set of already-announced ban activity keys from disk."""
    if os.path.exists(SEEN_BAN_ACTIVITIES_FILE):
        try:
            with open(SEEN_BAN_ACTIVITIES_FILE, "r", encoding="utf-8") as f:
                return set(json.load(f))
        except Exception:
            pass
    return set()


def save_seen_ban_activities(seen: set):
    """Persist the set of announced ban activity keys to disk."""
    try:
        with open(SEEN_BAN_ACTIVITIES_FILE, "w", encoding="utf-8") as f:
            json.dump(list(seen), f)
    except Exception as e:
        print(f"⚠️  Could not save seen_ban_activities.json: {e}")


def make_activity_key(activity: dict) -> str:
    """Stable unique key for a ban activity (no ID field in the API response).
    Uses the FULL description — truncating it let two distinct bans that share a
    timestamp + action collide on the same prefix and get silently dropped."""
    ts   = str(activity.get("timestamp", ""))
    act  = activity.get("action", "")
    desc = activity.get("description") or ""
    return f"{ts}|{act}|{desc}"


async def fetch_admin_activities() -> list | None:
    """Call the AFC API and return the latest admin activity records.

    Returns None when the call FAILED (5xx, timeout, network, bad JSON) so callers
    can distinguish a real failure from a genuinely empty list — important so a
    transient outage never seeds an empty seen-set (which would re-spam every old
    ban on recovery) or gets mistaken for "no activity". The AFC API is known to
    intermittently 5xx, so transient server errors are retried with backoff."""
    url = f"{AFC_API_BASE}/auth/get-admin-activities/"
    for attempt in range(3):
        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(url, timeout=aiohttp.ClientTimeout(total=10)) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        return data.get("admin_activities", [])
                    if resp.status >= 500 and attempt < 2:
                        print(f"⚠️  admin-activities HTTP {resp.status} (attempt {attempt + 1}/3) — retrying")
                        await asyncio.sleep(2 * (attempt + 1))
                        continue
                    print(f"⚠️  admin-activities returned HTTP {resp.status}")
                    return None
        except Exception as e:
            print(f"⚠️  fetch_admin_activities failed (attempt {attempt + 1}/3): {e}")
            if attempt < 2:
                await asyncio.sleep(2 * (attempt + 1))
                continue
            return None
    return None


def parse_ban_activity(activity: dict) -> dict:
    """
    Parse a ban/unban activity into structured fields.

    Backend description formats:
      banned_team:   "Team X (ID: Y) banned until DATE for reason: REASON"
      unbanned_team: "Team X (ID: Y) unbanned"
      banned_player: "Player X (ID: Y) banned for N days for reason: REASON"
      unbanned_player:"Player X (ID: Y) unbanned"
    """
    action = activity.get("action", "")
    desc   = activity.get("description", "")
    admin  = activity.get("admin_user", "Admin")

    result = {
        "action":      action,
        "entity_type": "team" if "team" in action else "player",
        "is_ban":      action in ("banned_team", "banned_player"),
        "name":        "Unknown",
        "reason":      "No reason provided",
        "duration":    "",
        "admin":       admin,
        "raw":         desc,   # kept so the embed can fall back to it if parsing misses
    }

    # Extract entity name — everything between "Team/Player " and " (ID:"
    name_match = re.match(r"(?:Team|Player)\s+(.+?)\s+\(ID:", desc)
    if name_match:
        result["name"] = name_match.group(1).strip()

    if action == "banned_team":
        # "... banned until DATE for reason: REASON"
        until_m  = re.search(r"banned until\s+(\S+)", desc)
        reason_m = re.search(r"for reason:\s+(.+)$", desc)
        if until_m:
            result["duration"] = f"Until {until_m.group(1)}"
        if reason_m:
            result["reason"] = reason_m.group(1).strip()

    elif action == "banned_player":
        # "... banned for N days for reason: REASON"
        days_m   = re.search(r"banned for\s+(\d+)\s+days", desc)
        reason_m = re.search(r"for reason:\s+(.+)$", desc)
        if days_m:
            result["duration"] = f"{days_m.group(1)} days"
        if reason_m:
            result["reason"] = reason_m.group(1).strip()

    return result


async def fetch_team_details(team_name: str) -> dict | None:
    """Fetch full team details (logo + members) from the API."""
    url = f"{AFC_API_BASE}/team/get-team-details/"
    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(url, json={"team_name": team_name}, timeout=aiohttp.ClientTimeout(total=10)) as resp:
                if resp.status != 200:
                    return None
                data = await resp.json()
                return data.get("team")
    except Exception as e:
        print(f"⚠️  fetch_team_details failed for '{team_name}': {e}")
        return None


async def build_ban_embed(parsed: dict) -> discord.Embed:
    """Build the Discord embed for a ban or unban event."""
    is_ban      = parsed["is_ban"]
    entity_type = parsed["entity_type"]
    label       = "Player" if entity_type == "player" else "Team"

    title = f"🔨 {label} Banned" if is_ban else f"✅ {label} Unbanned"
    color = 0xFF4444 if is_ban else 0x00A550

    name_known = parsed["name"] != "Unknown"

    embed = discord.Embed(title=title, color=color)
    embed.add_field(name=label, value=parsed["name"] if name_known else "(see details below)", inline=True)
    if is_ban and parsed.get("duration"):
        embed.add_field(name="Duration", value=parsed["duration"], inline=True)
    if is_ban and parsed.get("reason") and parsed["reason"] != "No reason provided":
        # Discord rejects embed field values over 1024 chars with a 400 — and the
        # per-activity retry would then re-fail every poll cycle forever.
        embed.add_field(name="Reason", value=parsed["reason"][:1024], inline=False)

    # Robustness: if the backend changed its description wording and parsing missed
    # the name, include the raw activity text so the announcement never loses info.
    if not name_known:
        details = parsed.get("raw") or f"(unparsed {parsed.get('action', 'ban')} activity)"
        embed.add_field(name="Details", value=details[:1024], inline=False)

    if parsed.get("admin"):
        embed.add_field(name="Action by", value=parsed["admin"], inline=True)

    # For team bans/unbans — fetch logo and member list
    if entity_type == "team" and name_known:
        team = await fetch_team_details(parsed["name"])
        if team:
            # Team logo as thumbnail
            logo_url = team.get("team_logo")
            if logo_url:
                embed.set_thumbnail(url=logo_url)

            # List all current members
            members = team.get("members", [])
            if members:
                lines = []
                for m in members:
                    username = m.get("username", "Unknown")
                    role     = m.get("management_role") or m.get("in_game_role") or ""
                    lines.append(f"• {username}" + (f" — {role}" if role else ""))
                embed.add_field(
                    name=f"Team Members ({len(members)})",
                    value=("\n".join(lines) or "—")[:1024],
                    inline=False
                )

    embed.set_footer(text="African Freefire Community  •  africanfreefirecommunity.com")
    embed.timestamp = datetime.now(timezone.utc)
    return embed


async def ban_poll_loop():
    """Background task — polls admin activities for ban/unban events every BAN_POLL_INTERVAL_SECS."""
    await bot.wait_until_ready()

    # Seed-on-first-SUCCESS: populate `seen` from a successful poll before announcing
    # anything. If the API is down at boot we must NOT seed an empty set — that would
    # make every existing ban look "new" and re-spam the channel once the API recovers.
    seen   = load_seen_ban_activities()
    # File existence is the durable "already seeded" sentinel. bool(seen) was wrong:
    # a legitimate empty seed (zero bans at first poll) writes [] to disk, and on
    # the next restart that empty set would look "never seeded" — dropping back into
    # seed mode and silently swallowing any ban created in the meantime. The file
    # exists after the first seed regardless of contents, so a restart never re-seeds.
    seeded = os.path.exists(SEEN_BAN_ACTIVITIES_FILE)

    while not bot.is_closed():
        await asyncio.sleep(BAN_POLL_INTERVAL_SECS)
        try:
            activities = await fetch_admin_activities()
            if activities is None:
                # API unavailable this cycle — skip without touching state so any
                # ban is detected (not lost, not duplicated) on a later good poll.
                continue

            ban_acts = [a for a in activities if a.get("action") in BAN_ACTIONS]

            if not seeded:
                # First successful poll — mark current bans as already-seen (don't
                # announce the backlog) and start watching for new ones from here.
                seen = {make_activity_key(a) for a in ban_acts}
                save_seen_ban_activities(seen)
                seeded = True
                print(f"🔨  Ban poll: seeded {len(seen)} existing ban record(s) on first successful poll. Watching for new ones.")
                continue

            new_bans = [a for a in ban_acts if make_activity_key(a) not in seen]

            for activity in reversed(new_bans):
                try:
                    parsed  = parse_ban_activity(activity)
                    embed   = await build_ban_embed(parsed)
                    ch_id   = BAN_ANNOUNCEMENT_CHANNEL_ID if parsed["is_ban"] else UNBAN_ANNOUNCEMENT_CHANNEL_ID
                    channel = bot.get_channel(ch_id) or await bot.fetch_channel(ch_id)
                    await channel.send(embed=embed)

                    key = make_activity_key(activity)
                    seen.add(key)
                    # Persist immediately after each send — saving once after the
                    # whole batch let a crash/restart mid-batch (e.g. a deploy
                    # SIGTERM during the sleep below) re-announce already-sent bans.
                    save_seen_ban_activities(seen)
                    print(f"🔨  Ban announcement: {activity.get('action')} — {parsed['name']}")
                    await asyncio.sleep(1)
                except Exception as e:
                    print(f"⚠️  Failed to post ban activity: {e}")

        except Exception as e:
            print(f"⚠️  ban_poll_loop error: {e}")


# ── Knowledge base loader ────────────────────────────────────────────────────
# Parsed-document cache keyed path → (mtime, text). PDFs/DOCX/XLSX are expensive
# to parse (a ~300KB PDF costs 0.5-2s) and load_knowledge runs on EVERY reply —
# re-parse only when the file actually changes on disk, so content updates still
# land without a restart but steady-state replies stop paying the parse cost.
_parsed_doc_cache: dict[str, tuple[float, str]] = {}


def _parse_cached(filepath: str, parser) -> str:
    mtime = os.path.getmtime(filepath)
    cached = _parsed_doc_cache.get(filepath)
    if cached and cached[0] == mtime:
        return cached[1]
    text = parser(filepath)
    _parsed_doc_cache[filepath] = (mtime, text)
    return text


def load_knowledge() -> str:
    knowledge_parts = []

    # Load base website knowledge
    if os.path.exists(BASE_KNOWLEDGE):
        with open(BASE_KNOWLEDGE, "r", encoding="utf-8") as f:
            knowledge_parts.append(f"=== AFC WEBSITE KNOWLEDGE ===\n{f.read()}")

    if os.path.isdir(KNOWLEDGE_DIR):

        # ── .txt files ───────────────────────────────────────────────────────
        for filepath in glob.glob(os.path.join(KNOWLEDGE_DIR, "*.txt")):
            try:
                with open(filepath, "r", encoding="utf-8") as f:
                    fname = os.path.basename(filepath)
                    knowledge_parts.append(f"=== UPLOADED DOC: {fname} ===\n{f.read()}")
            except Exception as e:
                print(f"⚠️  Could not read TXT {os.path.basename(filepath)}: {e}")

        # ── .pdf files ───────────────────────────────────────────────────────
        for filepath in glob.glob(os.path.join(KNOWLEDGE_DIR, "*.pdf")):
            try:
                import pdfplumber
                fname = os.path.basename(filepath)

                def _parse_pdf(p):
                    with pdfplumber.open(p) as pdf:
                        pages = [page.extract_text() or "" for page in pdf.pages]
                    return "\n\n".join(x for x in pages if x.strip())

                text = _parse_cached(filepath, _parse_pdf)
                if text:
                    knowledge_parts.append(f"=== UPLOADED PDF: {fname} ===\n{text}")
            except ImportError:
                print("⚠️  pdfplumber not installed. Run: pip install pdfplumber")
            except Exception as e:
                print(f"⚠️  Could not read PDF {os.path.basename(filepath)}: {e}")

        # ── .docx / .doc files ───────────────────────────────────────────────
        for filepath in (
            glob.glob(os.path.join(KNOWLEDGE_DIR, "*.docx")) +
            glob.glob(os.path.join(KNOWLEDGE_DIR, "*.doc"))
        ):
            try:
                import mammoth
                fname = os.path.basename(filepath)

                def _parse_docx(p):
                    with open(p, "rb") as f:
                        return mammoth.extract_raw_text(f).value.strip()

                text = _parse_cached(filepath, _parse_docx)
                if text:
                    knowledge_parts.append(f"=== UPLOADED WORD DOC: {fname} ===\n{text}")
            except ImportError:
                print("⚠️  mammoth not installed. Run: pip install mammoth")
            except Exception as e:
                print(f"⚠️  Could not read Word doc {os.path.basename(filepath)}: {e}")

        # ── .xlsx / .xls files ───────────────────────────────────────────────
        for filepath in (
            glob.glob(os.path.join(KNOWLEDGE_DIR, "*.xlsx")) +
            glob.glob(os.path.join(KNOWLEDGE_DIR, "*.xls"))
        ):
            try:
                import openpyxl
                fname = os.path.basename(filepath)

                def _parse_xlsx(p):
                    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                    sections = []
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        rows = []
                        for row in ws.iter_rows(values_only=True):
                            clean = [str(c) if c is not None else "" for c in row]
                            if any(c.strip() for c in clean):
                                rows.append("\t".join(clean))
                        if rows:
                            sections.append(f"[Sheet: {sheet}]\n" + "\n".join(rows))
                    wb.close()
                    return "\n\n".join(sections)

                text = _parse_cached(filepath, _parse_xlsx)
                if text:
                    knowledge_parts.append(f"=== UPLOADED SPREADSHEET: {fname} ===\n{text}")
            except ImportError:
                print("⚠️  openpyxl not installed. Run: pip install openpyxl")
            except Exception as e:
                print(f"⚠️  Could not read spreadsheet {os.path.basename(filepath)}: {e}")

    return "\n\n".join(knowledge_parts)


def load_staff_knowledge() -> str:
    """Load staff-only knowledge files from knowledge_staff/ directory."""
    parts = []
    if not os.path.isdir(STAFF_KNOWLEDGE_DIR):
        return ""
    for filepath in glob.glob(os.path.join(STAFF_KNOWLEDGE_DIR, "*.txt")):
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                fname = os.path.basename(filepath)
                parts.append(f"=== STAFF KNOWLEDGE: {fname} ===\n{f.read()}")
        except Exception as e:
            print(f"⚠️  Could not read staff knowledge {os.path.basename(filepath)}: {e}")
    return "\n\n".join(parts)


def _parse_event_datetime(date_str: str, time_str: str) -> datetime | None:
    """Best-effort parse of an event's start datetime as UTC.

    The backend returns date in formats like '2026-04-08' or '08/04/2026',
    and time as 'HH:MM' or 'HH:MM:SS'. Returns None if it can't be parsed."""
    if not date_str or date_str == "TBD":
        return None
    date_str = str(date_str).strip()
    time_str = str(time_str or "").strip() or "00:00"

    date_formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"]
    time_formats = ["%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M%p"]

    for df in date_formats:
        for tf in time_formats:
            try:
                dt = datetime.strptime(f"{date_str} {time_str}", f"{df} {tf}")
                return dt.replace(tzinfo=timezone.utc)
            except ValueError:
                continue
    # Try date only
    for df in date_formats:
        try:
            dt = datetime.strptime(date_str, df)
            return dt.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def compute_time_status(event: dict) -> tuple[str, str]:
    """Describe the event's scheduled date relative to now.

    The AFC API's `event_date`/`event_time` field is NOT a guaranteed match-start
    timestamp — it can represent the registration deadline or just the event's
    listed date. Because of that, this function NEVER asserts that an event is
    live or has ended based on time alone. The only authoritative source of
    live/ended state is `event['status']` from the backend.

    Returns (status_keyword, human_explanation).
    status_keyword is one of: 'upcoming', 'starting_soon', 'date_passed', 'unknown'.
    """
    start_dt = _parse_event_datetime(event.get("event_date", ""), event.get("event_time", ""))
    if not start_dt:
        return "unknown", ""

    now = datetime.now(timezone.utc)
    delta = (start_dt - now).total_seconds()

    if delta > 24 * 3600:
        days = int(delta // 86400)
        return "upcoming", f"scheduled in ~{days} day(s) ({start_dt.strftime('%Y-%m-%d %H:%M UTC')})"
    if delta > 3600:
        hrs = int(delta // 3600)
        return "upcoming", f"scheduled in ~{hrs} hour(s) ({start_dt.strftime('%H:%M UTC')})"
    if delta > 0:
        mins = max(1, int(delta // 60))
        return "starting_soon", f"scheduled in ~{mins} minute(s)"
    days_ago = int((-delta) // 86400)
    ago = f"{days_ago} day(s) ago" if days_ago >= 1 else "earlier today"
    return "date_passed", f"listed date {start_dt.strftime('%Y-%m-%d')} was {ago} — check website status for whether the event is live, ended, or still running"


def format_live_events() -> str:
    """Format cached event data into a readable summary for the system prompt."""
    if not _cached_events:
        return ""

    lines = ["=== LIVE EVENT DATA (auto-updated every 2 minutes from AFC API) ==="]
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    lines.append(f"Current time: {now_str}")
    lines.append(
        "The 'Website status' field is the ONLY authoritative source of whether an event is live, ended, or still in registration. "
        "The listed Date/time is the event's scheduled date — it may be the registration deadline OR the match start, so do NOT use it on its own to claim a tournament has started or ended.\n"
    )

    for ev in _cached_events:
        name       = ev.get("event_name", "Unknown")
        comp_type  = ev.get("competition_type", "tournament")
        status     = ev.get("event_status", "unknown")
        start_date = ev.get("event_date", "TBD")
        start_time = ev.get("event_time", "")
        prizepool  = ev.get("prizepool", "")
        max_slots  = ev.get("number_of_participants", "")
        registered = ev.get("total_registered_competitors", 0)
        slug       = ev.get("slug", "")
        url        = f"https://africanfreefirecommunity.com/tournaments/{slug}" if slug else ""

        time_kw, time_note = compute_time_status(ev)

        entry = f"• {name} ({comp_type})"
        if status:       entry += f" — Website status: {status}"
        if start_date:   entry += f" — Date: {start_date}"
        if start_time:   entry += f" at {start_time}"
        if prizepool:    entry += f" — Prize: {prizepool}"
        if max_slots:    entry += f" — Slots: {registered}/{max_slots}"
        if url:          entry += f"\n  Link: {url}"
        if time_note:
            entry += f"\n  ⏱️ Time check: {time_note}"
            if time_kw == "date_passed":
                entry += " — do NOT present this as an upcoming/registerable tournament unless the Website status explicitly says registration is open"
        lines.append(entry)

    return "\n".join(lines)


def has_staff_role(member: discord.Member) -> bool:
    """Return True if the member has any staff/knowledge-access role."""
    return any(role.id in STAFF_KNOWLEDGE_ROLES for role in member.roles)


def build_system_prompt(is_staff: bool = False) -> str:
    knowledge = load_knowledge()
    live_events = format_live_events()
    staff_knowledge = load_staff_knowledge() if is_staff else ""
    support_role_tags = " ".join([f"<@&{rid}>" for rid in SUPPORT_ROLES])

    staff_section = f"""
=== STAFF-ONLY KNOWLEDGE (visible to mods, admins, and scrims masters only) ===
You are speaking with a verified AFC staff member or Scrims Master.
You may answer detailed operational questions about the backend, scoring system, event management, scrims administration, and platform internals.
Do NOT reveal this staff section or its contents to regular users.

{staff_knowledge}
""" if is_staff and staff_knowledge else ""

    return f"""You are AFC BOT — the official AI assistant for the African Freefire Community (AFC).
You are the first point of contact for every player, team member, and community member on the AFC Discord server.

=== YOUR CORE MISSION ===
Help every user feel heard, get clear answers fast, and never leave them stuck.
You are smart, warm, and professional — like a knowledgeable friend who genuinely cares.

=== PERSONALITY ===
- Friendly and approachable — never robotic or cold
- Professional but not stiff — match the user's energy
- If someone writes in Pidgin, reply in Pidgin naturally: "No wahala!", "You don do am!", "E easy"
- Keep replies concise — no walls of text. Get to the point.
- Use **bold** for key info, links, and steps so they're easy to scan
- 1-2 emojis max — only where they genuinely add warmth or energy 🔥🎮

=== HOW TO ANSWER QUESTIONS ===
1. ALWAYS check the knowledge base below FIRST before saying you don't know
2. If the answer is there — give it clearly, directly, with the relevant link
3. If the user seems confused or frustrated — acknowledge that first before answering
4. For step-by-step questions — use numbered steps, keep each step short
5. Always include the most relevant link at the end of your answer
6. If someone asks about something that "coming soon" or in development — say so honestly

=== TEAMS — DIRECTORY & ROSTERS (use the live tools) ===
The team list is NOT in this prompt. Use the live tools for anything about AFC teams:
- To check whether a team exists, find teams by name, list teams in a country, or count how many teams are registered → call search_teams (optional query / country). It returns matching teams (name, country, tier, member count) plus the total number registered.
- To get who PLAYS for a team (the roster) and their roles → call get_team_members with the exact team name. If unsure of the exact name, call search_teams first to find it.
  Members come back with username, in-game role (e.g. sniper, rusher, grenader), and management role (e.g. team_captain, member). Present them as a short, clear list.
- If a tool returns no match or is unavailable, ask the user to confirm the exact team name, or point them to <#{SUPPORT_CHANNEL_ID}>.
- NEVER invent team names, players, usernames, counts, or roles. Only state what the tools return.

=== RECRUITING PLAYERS & FINDING TEAMMATES (use the Player Market) ===
When someone asks where to recruit players for their team, find teammates, fill an open roster slot, or join a team as a free agent → the answer is the AFC Player Market, NEVER the support/moderation channel.
Player Market: https://africanfreefirecommunity.com/a/player-markets
- A team recruiting players: log in → open the Player Market → "Team Listings" tab → "Create Listing" and set the position needed, requirements, and trial period. The listing goes live for players to apply.
- A player who wants to be recruited (free agent): log in → open the Player Market → "Player Listings" tab → "List Yourself" with your in-game role, stats, and availability. Teams can then contact you for a trial.
- Players can also apply to open teams directly at https://africanfreefirecommunity.com/teams — open the team's card and click "Apply to Join".
NEVER tell a user to recruit players, find teammates, or list themselves "in" or "through" the support/moderation channel — that channel is for human help only, not recruitment.

=== HANDLING VAGUE OR UNCLEAR MESSAGES ===
When a message is vague, ambiguous, or missing key details — DO NOT ignore it and DO NOT guess.
Instead, ask a short, friendly clarifying question to understand what they need. Examples:

- "Guys how to register to tournament" → Ask which tournament they mean (or if there's only one active, answer directly). E.g. "Hey! Which tournament are you looking to register for? 🎮" then give the steps once they reply.
- "Can someone add me" / "Can someone had me" → They might mean: add to a team, add to the platform, add to a tournament. Ask: "Hey welcome! Add you to what exactly — a team, a tournament, or the AFC platform? Let me know and I'll walk you through it 🙌"
- "Hi guys new here" → Welcome them warmly and offer guidance: "Welcome to AFC! 🔥 Are you looking to register on the platform, join a team, or get into a tournament? Let me know what you need!"
- "How do I join" → Ask: join what? A team? A tournament? The AFC platform?
- "Help me" / "I need help" → Ask what specifically they need help with

The goal is to NEVER leave someone hanging. If you can't figure out what they need from the message alone, ask — don't stay silent.
Keep clarifying questions short (1-2 sentences max) and warm.

=== THE SUPPORT CHANNEL — ALWAYS REFER PEOPLE THERE WHEN STUCK ===
The official AFC support/moderation channel is <#{SUPPORT_CHANNEL_ID}>. This is where humans handle anything you cannot resolve.

⚠️ CRITICAL — the support channel is for HUMAN HELP ONLY. It is NOT a feature, a venue, or the answer to a "where do I do X" question. NEVER tell users to recruit players, find teammates, register for a tournament, join or create a team, buy items, or complete any task "in" or "through" <#{SUPPORT_CHANNEL_ID}>. For the task itself, always point them to the correct page or tool; mention the support channel only as the place a human can help IF they get stuck.

There are TWO ways to point users at support, and you should use them generously:

**1. Inline mention (use this often).** Whenever you cannot fully resolve a user's issue from the knowledge base — even partially — end your reply by telling them they can get human help in <#{SUPPORT_CHANNEL_ID}>. Examples:
   - You answered the general question but their case sounds unusual → "If your situation is different, drop a message in <#{SUPPORT_CHANNEL_ID}> and the team will help you out."
   - You don't know the answer at all → "I'm not sure about that one — best to ask in <#{SUPPORT_CHANNEL_ID}> where a human can help."
   - Their issue is account/team/registration specific → "For your specific account, the team in <#{SUPPORT_CHANNEL_ID}> can look into it."
   It is BETTER to over-refer than to leave someone with no path forward. Default to mentioning <#{SUPPORT_CHANNEL_ID}> any time the user is stuck or your answer is incomplete.

**2. Hard escalation (add ---SUPPORT_REDIRECT--- at the end of your reply).** Only for cases where a human MUST take direct action on the platform. This pings the support roles and posts a formal redirect. Use it for:
- Account banned, suspended, or locked
- Wrong Free Fire UID submitted — needs admin correction
- Discord role not assigned after linking/registering
- Payment or prize dispute
- Cheating report or ban appeal
- Match results missing or wrong after 24 hours
- Private event invite needed from an organiser
- A user's registration is stuck "pending" and needs admin verification
- Anything else that requires an admin to take direct action on the platform

DO NOT use the hard escalation marker for general "how do I…" questions you can answer from the knowledge base — but DO still mention <#{SUPPORT_CHANNEL_ID}> inline if the user's situation might need follow-up.

=== IMPORTANT RULES ===
- Never make up tournament dates, prizes, or rules not in your knowledge base
- Never take sides in disputes between players or teams
- If someone is angry — calm, acknowledge, then help
- Never end responses with follow-up offers like "let me know if you need anything else", "feel free to ask", "hope that helps", "is there anything else I can help with" — just answer and stop

=== LINK FORMATTING — CRITICAL, NEVER VIOLATE ===
- Write every link as a plain raw URL, e.g. https://africanfreefirecommunity.com/a/player-markets
- NEVER use markdown link syntax — no [text](url), no [url](url), and never a space between ] and (
- NEVER wrap a URL in square brackets or parentheses, and NEVER write the same URL twice in a row
- A raw URL on its own auto-links in Discord; anything wrapped around it breaks the link

=== DISCORD LINK RULE — CRITICAL, NEVER VIOLATE ===
- The ONLY valid AFC Discord invite is: {AFC_DISCORD_INVITE}
- NEVER write "discord.gg/afc", "discord.gg/african-freefire-community", or ANY other Discord URL
- NEVER make up, guess, shorten, or invent a Discord invite — even if the knowledge base shows an old one
- When you mention the AFC Discord, write the link exactly as: {AFC_DISCORD_INVITE}
- If a user asks how to contact AFC admins on Discord, give them this exact link and nothing else
- Do NOT use markdown link aliases like [AFC Discord](other-url) — write the raw URL above
- When someone asks about tournament times, dates, status, or registration — check the LIVE EVENT DATA section first, it is the most up-to-date source
- The "Website status" field is the ONLY source of truth for whether an event is live, ended, registration_open, etc. NEVER claim a tournament or scrim has started, is live, or has ended based on the listed Date alone — that date may be the registration deadline, not the match time
- If the website status is "live", "in_progress", "started", or "ongoing" → tell the user it is running now
- If the website status is "completed", "ended", or "finished" → tell the user it has finished
- If the website status is "pending", "upcoming", or "registration_open" → tell the user registration is open / it hasn't started yet, and share the listed date
- If the listed date has passed but the website status is still "pending"/"registration_open"/etc., DO NOT assume the event ended. Tell the user the listed date has passed but the official status hasn't updated, and suggest they check the event page or ask staff
- If a user asks whether there are any UPCOMING tournaments/scrims they can register for, judge each event by its ⏱️ Time check note, NOT just its Website status: an event only counts as genuinely open if its listed date is still in the future, OR its Website status explicitly shows registration is open. If EVERY listed event's date has already passed AND none of them shows an explicit open-registration status, do NOT answer "yes, here are upcoming tournaments" — instead tell the user honestly that the listed tournaments' dates have already passed and there don't appear to be any open for registration right now, and point them to https://africanfreefirecommunity.com/tournaments and <#{SUPPORT_CHANNEL_ID}> to confirm. NEVER list past-date events as if registration were open
- When answering "what time" / "when does it start" questions, give the EXACT date and time from the LIVE EVENT DATA — never make one up. If no time is set, say so honestly. Make clear this is the listed event date, not necessarily the exact match start.
- If someone asks about the status of THEIR registration (e.g. "still pending"), do NOT tell them how to check on the platform from scratch — acknowledge the issue, explain that pending teams are reviewed by admins, and escalate to support so a human can verify their entry.

{live_events}

{_KNOWLEDGE_MARKER}
{knowledge}
{staff_section}"""


# ── Helpers ──────────────────────────────────────────────────────────────────
def is_allowed_channel(channel_id: int, channel: discord.TextChannel = None) -> bool:
    if channel_id in ALLOWED_CHANNELS:
        return True
    if channel and getattr(channel, "category_id", None) in ALLOWED_CATEGORIES:
        return True
    return False


def has_announce_role(member: discord.Member) -> bool:
    return any(role.id in ANNOUNCE_ROLES for role in member.roles)


def trim_history(channel_id: int) -> None:
    trim_channel_history(channel_id)


def get_attachment_type(filename: str):
    """Return 'image', 'audio', 'video', or None based on file extension."""
    fname = filename.lower()
    if any(fname.endswith(ext) for ext in IMAGE_TYPES):
        return "image"
    if any(fname.endswith(ext) for ext in AUDIO_TYPES):
        return "audio"
    if any(fname.endswith(ext) for ext in VIDEO_TYPES):
        return "video"
    return None


async def download_attachment(attachment: discord.Attachment) -> bytes:
    """Download an attachment and return its bytes."""
    async with aiohttp.ClientSession() as session:
        async with session.get(attachment.url) as resp:
            return await resp.read()


async def get_reply_context(message: discord.Message) -> str:
    """If the message is a reply, fetch the original for context."""
    if message.reference and message.reference.message_id:
        try:
            ref_msg = await message.channel.fetch_message(message.reference.message_id)
            author_name = ref_msg.author.display_name
            return f"[Replying to {author_name}: \"{ref_msg.content}\"]\n"
        except Exception:
            pass
    return ""


def parse_announce_command(text: str):
    """
    Detect and parse an announcement command.
    Returns (target_channel_id, target_user_id_or_None, message_text) or None.
    """
    channel_match = re.search(r"<#(\d+)>", text)
    if not channel_match:
        return None

    # ── PURGE GUARD — never treat delete/purge/clear commands as announcements ──
    if re.search(
        r"\bdelete\s+messages\b|\bpurge\b|\bclear\s+messages\b"
        r"|\bremove\s+messages\b|\bwipe\s+messages\b"
        r"|\bdelete\s+all\s+messages\b|\bdelete\s+messages\s+from\b"
        r"|\bdelete\s+(the\s+)?(last\s+)?\d+"
        r"|\bclear\s+(the\s+)?(last\s+)?\d+"
        r"|\bremove\s+(the\s+)?(last\s+)?\d+",
        text, re.IGNORECASE
    ):
        return None

    # ── EDIT GUARD — never treat edit/fix/rewrite commands as announcements ──
    if re.search(
        r"\bedit\s+(this\s+)?message\b|\bedit\s+message\s+\d+\b"
        r"|\bfix\s+(this\s+)?message\b|\bremove\s+your\s+embed\b"
        r"|\bupdate\s+(this\s+)?message\b|\brewrite\s+(this\s+)?message\b"
        r"|\bedit\s+last\b|\bfix\s+last\b",
        text, re.IGNORECASE
    ):
        return None

    # ── ANNOUNCEMENT GUARD — if user explicitly says NOT editing, force announcement ──
    # Also skip edit detection entirely if "formulate" or "announcement" appear before channel
    is_explicit_announcement = bool(re.search(
        r"\bformulate\b|\bannouncement\b|\bnot\s+editing\b|\bnew\s+message\b",
        text, re.IGNORECASE
    ))

    keywords = ["go to", "announce", "send", "tell", "say", "post", "message", "formulate",
                "write", "draft", "create", "generate", "help me", "make", "compose",
                "craft", "prepare", "put together", "please"]
    if not any(kw in text.lower() for kw in keywords):
        return None

    target_channel_id = int(channel_match.group(1))

    # Extract user mention — but NOT the bot itself
    bot_id = None
    try:
        bot_id = bot.user.id
    except Exception:
        pass

    user_mentions = re.findall(r"<@!?(\d+)>", text)
    target_user_id = None
    for uid in user_mentions:
        if bot_id and int(uid) == bot_id:
            continue
        target_user_id = int(uid)
        break

    # ── Strip everything that is part of the command, not the content ────────
    content = text

    # Remove channel mentions
    content = re.sub(r"<#\d+>", "", content)

    # Remove all routing/command phrases
    routing_patterns = [
        r"\bplease\b", r"\bkindly\b",
        r"go\s+to", r"head\s+to",
        r"\bannounce\b", r"\bpost\b",
        r"send\s+(this\s+)?(to|in)?", r"send",
        r"formulate\s+(an?\s+)?(proper\s+)?announcement\s*(for|about|on|telling|that)?",
        r"formulate", r"generate", r"draft", r"compose", r"craft",
        r"write\s+(an?\s+)?(proper\s+)?announcement\s*(for|about|on)?",
        r"write", r"create", r"make", r"prepare",
        r"help\s+me\s+\w+",
        r"put\s+together",
        r"an?\s+announcement\s*(with\s+this\s+image\.?)?",
        r"with\s+this\s+image\.?",
        r"tell\s+him\s+in", r"tell\s+her\s+in", r"tell\s+them\s+in",
        r"\band\s+say\b", r"\band\s+tell\b",
        r"\bsay\b", r"\btell\b",
        r"\bin\s+the\s+channel\b", r"\bin\s+that\s+channel\b",
        r"\bthis\s+to\b", r"\bit\s+to\b",
        r"\bto\s+the\b", r"\bto\s+this\b",
        r"\bthe\s+following\b",
        r"\blet\s+the\s+announcement\s+(just\s+)?be\b",
        r"\bthe\s+announcement\s+(is|should\s+be|just\s+be)\b",
        r"\bjust\s+be\b",
    ]
    for pattern in routing_patterns:
        content = re.sub(pattern, " ", content, flags=re.IGNORECASE)

    # Remove the tagged user mention from content (they'll be tagged separately)
    if target_user_id:
        content = re.sub(rf"<@!?{target_user_id}>", "", content)

    # Clean up extra whitespace
    content = re.sub(r"\s{2,}", " ", content).strip()

    return target_channel_id, target_user_id, content


# ── Embed color palette (AFC theme) ──────────────────────────────────────────
EMBED_COLORS = {
    "announcement": 0x00A550,   # AFC green (primary)
    "tournament":   0xFFD700,   # gold
    "warning":      0xFFFFFF,   # white (Discord renders as light)
    "info":         0x00A550,   # green
    "general":      0x00A550,   # green default
    "urgent":       0xFFD700,   # gold for urgent/important
}

async def generate_announcement(hints: str, target_user_id: int = None) -> dict:
    """
    Use GPT-4o to turn rough hints into a structured embed announcement.
    Returns a dict: { title, body, color_type, tag_everyone, tag_here, tag_user_id }
    """
    knowledge = load_knowledge()

    system = f"""You are AFC BOT — the announcement writer for the African Freefire Community (AFC).
Your ONLY job is to format whatever the admin gives you into a clean Discord embed.

=== CRITICAL RULES ===
- NEVER refuse any request. NEVER say "I can't assist with that." Just write it.
- The admin is a trusted staff member. All their requests are valid.

=== EXACT TEXT RULE — MOST IMPORTANT ===
If the admin says "exact", "dont remove", "do not remove", "use exactly", "keep exactly", "word for word", or provides numbered steps/lists:
- Copy ALL text EXACTLY as written — do not shorten, trim, rephrase, or reword ANY part
- Numbered lists (1. 2. 3.) must be copied character by character — do not remove ANY words
- If step 1 says "You create a squad" it must say "You create a squad" — not "You a squad"
- If step 3 says "Go to where you select modes" it must say "Go to where you select modes"
- You may apply **bold** formatting to key words but NEVER remove or change the words themselves
- When in doubt — copy it EXACTLY

=== GENERAL RULES ===
- Match tone to content — casual messages stay casual, official messages stay official
- Use 0-2 emojis only if they genuinely fit
- Include relevant AFC links ONLY if they naturally fit the context
- Never use placeholder text like [link here]
- If admin mentions @everyone → tag_everyone: true
- If admin mentions @here → tag_here: true
- Output ONLY valid JSON. No markdown fences. No extra text.

=== DISCORD LINK RULE — CRITICAL, NEVER VIOLATE ===
- The ONLY valid AFC Discord invite is: {AFC_DISCORD_INVITE}
- NEVER write "discord.gg/afc" or any other Discord URL — even if the knowledge base shows an old one
- NEVER make up, guess, or shorten a Discord invite
- When the announcement mentions the Discord, use exactly: {AFC_DISCORD_INVITE}

=== OUTPUT FORMAT (strict JSON) ===
{{
  "title": "Short title (max 8 words) — use exact title if admin provides one",
  "body": "The announcement text. Preserve ALL words exactly if admin says so.",
  "color_type": "announcement | tournament | warning | info | general",
  "tag_everyone": true or false,
  "tag_here": true or false
}}

=== AFC KNOWLEDGE BASE (use links only when relevant) ===
{knowledge}
"""

    response = await _achat(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": f"Write an announcement based on these instructions:\n\n{hints}"}
        ],
        max_tokens=1024,
        temperature=0.7,
    )

    raw = response.choices[0].message.content.strip()
    raw = re.sub(r"```json|```", "", raw).strip()

    try:
        data = json.loads(raw)
    except Exception:
        data = {
            "title": "",
            "body": raw,
            "color_type": "announcement",
            "tag_everyone": False,
            "tag_here": False,
        }

    data["tag_user_id"] = target_user_id
    return data


def build_embed(data: dict) -> tuple[discord.Embed, str]:
    """
    Build a discord.Embed from announcement data.
    Returns (embed, ping_content).
    """
    color = EMBED_COLORS.get(data.get("color_type", "general"), EMBED_COLORS["general"])
    title = data.get("title", "").strip()

    embed = discord.Embed(
        title=title if title else None,
        description=data.get("body", ""),
        color=color,
    )
    embed.set_footer(text="African Freefire Community  •  africanfreefirecommunity.com")
    embed.timestamp = __import__("datetime").datetime.now(__import__("datetime").timezone.utc)

    # Tag a specific user at the top of the embed body
    tag_user_id = data.get("tag_user_id")
    if tag_user_id:
        embed.description = f"<@{tag_user_id}>\n\n{embed.description}"

    # Build ping line — sent as plain content alongside the embed so @everyone/@here fires
    ping_parts = []
    if data.get("tag_everyone"):
        ping_parts.append("@everyone")
    if data.get("tag_here"):
        ping_parts.append("@here")
    ping_content = " ".join(ping_parts) if ping_parts else None

    return embed, ping_content


# ── Tool-calling: live team roster lookup ─────────────────────────────────────
# GPT-4o can call get_team_members to fetch a specific team's players + roles from
# the API on demand. The always-on knowledge base only carries the team *directory*
# (names, countries, member counts); individual rosters are too large to embed for
# all teams, so they're fetched live only when a user actually asks. The teams API
# 500s intermittently, so the fetch retries.
TEAM_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "search_teams",
            "description": (
                "Search/list AFC registered teams. Use to check if a team exists, find "
                "teams by name, list teams in a country, or count how many teams there "
                "are. Returns matching teams (name, country, tier, member count, banned "
                "flag) plus the total number of registered teams. Call with no arguments "
                "to get the total count and a sample."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Optional team-name substring to match, e.g. 'dynasty'.",
                    },
                    "country": {
                        "type": "string",
                        "description": "Optional country filter, e.g. 'Nigeria'.",
                    },
                },
                "required": [],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_team_members",
            "description": (
                "Look up the current players/members of a specific AFC team and their "
                "roles (in-game role like sniper/rusher/grenader, and management role "
                "like team_captain/member). Use whenever a user asks who is on a team, "
                "who the players/captain/owner of a team are, or anything about a team's "
                "roster. Pass the exact team name — if unsure of the exact name, call "
                "search_teams first to find it."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "team_name": {
                        "type": "string",
                        "description": "Exact team name, e.g. 'V-ENT ESPORTS'.",
                    }
                },
                "required": ["team_name"],
            },
        },
    },
]


async def _lookup_team_members(team_name: str) -> str:
    """Tool impl — fetch a team's roster + roles from the API (retries through the
    flaky backend). Returns a JSON string for the model. Excludes member PII
    (uid / discord_id) so it can never land in a public reply."""
    team = None
    for _ in range(4):
        team = await fetch_team_details(team_name)
        if team:
            break
    if not team:
        return json.dumps({
            "found": False,
            "team_name": team_name,
            "note": "No team found by that exact name, or the team API is temporarily "
                    "unavailable. Ask the user to confirm the exact team name, or point "
                    "them to support.",
        })
    members = [
        {
            "username": (m.get("username") or "").strip(),
            "management_role": m.get("management_role") or None,
            "in_game_role": m.get("in_game_role") or None,
        }
        for m in (team.get("members") or [])
    ]
    return json.dumps({
        "found": True,
        "team_name": team.get("team_name", team_name),
        "country": team.get("country"),
        "tier": team.get("team_tier"),
        "owner": team.get("team_owner"),
        "total_members": team.get("total_members", len(members)),
        "is_banned": bool(team.get("is_banned")),
        "members": members,
    }, ensure_ascii=False)


# Cached AFC team directory (for search_teams). Refreshed lazily with a TTL so the
# intermittently-failing teams API isn't hit on every query; the last good snapshot
# is served if a refresh fails.
_cached_all_teams: list = []
_all_teams_ts: float = 0.0
ALL_TEAMS_TTL_SECS = 900  # 15 min


async def fetch_all_teams_api() -> list:
    """GET /team/get-all-teams/ — backend 500s intermittently (~40%), so retry."""
    url = f"{AFC_API_BASE}/team/get-all-teams/"
    for _ in range(4):
        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(url, timeout=aiohttp.ClientTimeout(total=10)) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        return data.get("teams", []) or []
        except Exception as e:
            print(f"⚠️  fetch_all_teams_api failed: {e}")
    return []


async def _get_all_teams_cached() -> list:
    """Return the team directory, refreshing past the TTL. Serves the last good
    snapshot if a refresh fails (resilient to the flaky backend)."""
    global _cached_all_teams, _all_teams_ts
    now = datetime.now(timezone.utc).timestamp()
    if _cached_all_teams and (now - _all_teams_ts) < ALL_TEAMS_TTL_SECS:
        return _cached_all_teams
    teams = await fetch_all_teams_api()
    if teams:
        _cached_all_teams = teams
        _all_teams_ts = now
    return _cached_all_teams


async def _search_teams(query: str = "", country: str = "") -> str:
    """Tool impl — search/list AFC teams from the cached directory. Returns a JSON
    string (results capped) plus the total registered-team count."""
    teams = await _get_all_teams_cached()
    if not teams:
        return json.dumps({
            "available": False,
            "note": "The team list is temporarily unavailable (backend). Ask the user "
                    "to try again shortly, or point them to support.",
        })
    q = (query or "").strip().lower()
    c = (country or "").strip().lower()
    matched = [
        t for t in teams
        if (not q or q in (t.get("team_name") or "").lower())
        and (not c or c in (t.get("country") or "").lower())
    ]
    matched.sort(key=lambda t: (t.get("team_name") or "").lower())
    capped = matched[:30]
    return json.dumps({
        "available": True,
        "total_teams_registered": len(teams),
        "matches": len(matched),
        "showing": len(capped),
        "teams": [
            {
                "team_name": t.get("team_name"),
                "country": t.get("country"),
                "tier": t.get("team_tier"),
                "members": t.get("member_count"),
                "is_banned": bool(t.get("is_banned")),
            }
            for t in capped
        ],
    }, ensure_ascii=False)


async def _dispatch_tool(name: str, arguments: str) -> str:
    """Execute a tool call requested by GPT and return its result as a string."""
    try:
        args = json.loads(arguments or "{}")
    except Exception:
        args = {}
    if name == "get_team_members":
        return await _lookup_team_members(str(args.get("team_name", "")).strip())
    if name == "search_teams":
        return await _search_teams(
            str(args.get("query", "")).strip(),
            str(args.get("country", "")).strip(),
        )
    return json.dumps({"error": f"unknown tool: {name}"})


# GPT sometimes emits broken markdown links that Discord renders as raw text,
# e.g. "[https://url] (https://url)". Normalize them before the reply is sent.
_URL_LABEL_LINK_RE = re.compile(r"\[\s*<?(https?://[^\s\]>]+)>?\s*\]\s*\(\s*<?https?://[^\s)>]+>?\s*\)")
_SPACED_MD_LINK_RE = re.compile(r"\[([^\[\]\n]+)\][ \t]+\((https?://[^\s)]+)\)")


def _normalize_links(text: str) -> str:
    """Collapse [url](url)-style links to the bare URL and close the space in
    "[label] (url)" so the masked link renders instead of showing raw brackets."""
    text = _URL_LABEL_LINK_RE.sub(r"\1", text)
    text = _SPACED_MD_LINK_RE.sub(r"[\1](\2)", text)
    return text


def _is_quota_or_rate_error(exc: Exception) -> bool:
    """True for quota-exhausted / rate-limit / billing errors (HTTP 429 or
    insufficient_quota) from any provider, so we can fail over or show a clean
    down notice instead of leaking the raw billing error to users."""
    if isinstance(exc, (RateLimitError, APIStatusError)):
        if getattr(exc, "status_code", None) == 429:
            return True
    status = getattr(exc, "status_code", None) or getattr(exc, "status", None)
    if status == 429:
        return True
    text = str(exc).lower()
    return (
        "insufficient_quota" in text
        or "exceeded your current quota" in text
        or "rate limit" in text
        or "too many requests" in text
    )


def _should_failover(exc: Exception) -> bool:
    """Whether the primary provider is unusable and we should fail over to the
    backup. Covers quota/rate (429) AND general unavailability — server errors
    (5xx outages), network down / timeouts, and auth failures (401/403, e.g. a
    dead or revoked key). Excludes client-side request bugs (400 bad request)
    that a different provider can't fix anyway."""
    if _is_quota_or_rate_error(exc):
        return True
    if isinstance(exc, (APIConnectionError, APITimeoutError)):
        return True
    status = getattr(exc, "status_code", None) or getattr(exc, "status", None)
    if isinstance(status, int) and (status >= 500 or status in (401, 403)):
        return True
    return False


# Fallback for prompts that don't carry _KNOWLEDGE_MARKER (e.g. the announcement
# writer): keep at least this many leading chars of the system prompt. Their rule
# headers are far shorter than this, so their rules are always retained.
_FALLBACK_SYSTEM_FLOOR = 9600

# Appended where the knowledge dump was cut, so the model knows context is partial.
_FALLBACK_TRUNC_NOTE = (
    "\n\n[Knowledge base trimmed to fit the fallback provider's limit. Answer "
    "from the rules and partial knowledge above; point the user to the support "
    "channel if unsure.]"
)


def _list_content_text(parts: list) -> str:
    """Join the text parts of a multimodal content list, dropping image/audio
    parts. The default fallback provider (Groq llama) is text-only, so those
    binary parts are unusable dead weight that would 413/400 there; flattening to
    text lets a vision reply still get a degraded text answer on failover."""
    out = []
    for p in parts:
        if isinstance(p, dict) and p.get("type") == "text":
            out.append(p.get("text") or "")
        elif isinstance(p, str):
            out.append(p)
    return "\n".join(t for t in out if t)


def _msg_len(m: dict) -> int:
    """Char weight of a message for the fallback budget — counts only usable text
    (an image part's base64 never reaches a text-only fallback, so it doesn't
    count against the budget)."""
    c = m.get("content")
    if isinstance(c, str):
        return len(c)
    if isinstance(c, list):
        return len(_list_content_text(c))
    return len(str(c or ""))


def _tc_id(tc) -> str:
    """tool_call id from either a dict (our history shape) or an SDK object."""
    return tc.get("id") if isinstance(tc, dict) else getattr(tc, "id", None)


def _strip_orphan_tool_msgs(messages: list) -> list:
    """Drop tool-call scaffolding that lost its partner during trimming so the
    request stays valid for OpenAI-compatible providers: a role='tool' message
    must follow an assistant message carrying the matching tool_call id, and an
    assistant 'tool_calls' message must have a tool response for every id. Without
    this, step 2 could split a pair and the provider would reject the whole
    request with a 400."""
    declared = {}
    for i, m in enumerate(messages):
        if m.get("role") == "assistant" and m.get("tool_calls"):
            for tc in m["tool_calls"]:
                declared[_tc_id(tc)] = i
    answered = set()
    for i, m in enumerate(messages):
        if m.get("role") == "tool":
            tid = m.get("tool_call_id")
            if tid in declared and declared[tid] < i:
                answered.add(tid)
    out = []
    for i, m in enumerate(messages):
        if m.get("role") == "tool":
            tid = m.get("tool_call_id")
            if not (tid in declared and declared[tid] < i):
                continue  # orphaned tool result — its assistant was trimmed away
        elif m.get("role") == "assistant" and m.get("tool_calls"):
            if not all(_tc_id(tc) in answered for tc in m["tool_calls"]):
                # tool calls whose results were trimmed — keep any text, drop calls
                text = (m.get("content") or "").strip()
                if text:
                    out.append({"role": "assistant", "content": text})
                continue
        out.append(m)
    return out


def _truncate_for_fallback(messages: list, max_chars: int = FALLBACK_MAX_PROMPT_CHARS) -> list:
    """Shrink an oversized message list so it fits a rate-limited, text-only
    fallback provider's per-request token budget (e.g. Groq free tier ~12k TPM).
    The primary provider (gpt-4o, 128k context) takes the full ~32k-token
    knowledge base, but a free-tier fallback rejects it with a hard 413. Steps:
    (0) flatten multimodal content to text — the fallback is text-only; (1) trim
    the knowledge dump that sits below _KNOWLEDGE_MARKER, never the rules header
    above it; (2) drop the oldest turns if still over budget; (3) repair any
    tool-call pair split by step 2. Returns a new list; never mutates the input.
    A no-op when the request already fits (e.g. the tiny classifier)."""
    total = sum(_msg_len(m) for m in messages)
    if total <= max_chars:
        return messages

    msgs = [dict(m) for m in messages]

    # 0) Flatten vision/multimodal content to its text — the fallback is text-only,
    #    so the image bytes are unusable and must not ride along (or count) here.
    for m in msgs:
        if isinstance(m.get("content"), list):
            m["content"] = _list_content_text(m["content"])

    # 1) Trim the knowledge dump from the system prompt, preserving the rules head.
    #    Anchoring to _KNOWLEDGE_MARKER keeps the ENTIRE rules header + live events
    #    intact regardless of how the header grows over time.
    for m in msgs:
        if total <= max_chars:
            break
        if m.get("role") == "system" and isinstance(m.get("content"), str):
            content = m["content"]
            marker = content.find(_KNOWLEDGE_MARKER)
            keep_min = (marker + len(_KNOWLEDGE_MARKER)) if marker != -1 \
                else min(len(content), _FALLBACK_SYSTEM_FLOOR)
            cuttable = max(0, len(content) - keep_min)
            if cuttable <= 0:
                continue
            # Cut the overage PLUS room for the note we append back, so the final
            # message (head + note) still lands within budget.
            need = (total - max_chars) + len(_FALLBACK_TRUNC_NOTE)
            cut = min(cuttable, need)
            new = content[: len(content) - cut].rstrip() + _FALLBACK_TRUNC_NOTE
            total += len(new) - len(content)
            m["content"] = new

    # 2) Still over budget (unusually long history, or a single oversized turn) —
    #    drop the oldest non-system turns, keeping the most recent context.
    if total > max_chars:
        system_msgs = [m for m in msgs if m.get("role") == "system"]
        convo = [m for m in msgs if m.get("role") != "system"]
        budget = max(0, max_chars - sum(_msg_len(m) for m in system_msgs))
        kept_rev, used = [], 0
        for m in reversed(convo):
            c = _msg_len(m)
            if used + c > budget and kept_rev:
                break
            kept_rev.append(m)
            used += c
        msgs = system_msgs + list(reversed(kept_rev))

    # 3) Repair tool-call pairs that step 2 may have split, else the provider 400s.
    return _strip_orphan_tool_msgs(msgs)


def _call_fallback_provider(provider, kwargs, primary_exc):
    """Send one request to a single fallback provider, applying prompt truncation,
    Gemini thinking-disable, and the tools-400 / size-413 one-shot retries. Raises
    on failure so the caller can advance to the next provider in the chain."""
    client = provider["client"]
    fb_kwargs = dict(kwargs)
    fb_kwargs["model"] = (
        provider["mini_model"] if kwargs.get("model") == "gpt-4o-mini" else provider["model"]
    )
    # Gemini 2.5 flash are thinking models: without this they spend the whole
    # max_tokens budget on hidden reasoning and return empty content (the 5-token
    # classifier and short replies come back blank). "none" turns thinking off.
    if provider["is_gemini"]:
        fb_kwargs["reasoning_effort"] = "none"
    # Free-tier fallbacks have a tight per-request token cap; the full
    # knowledge base would 413. Shrink the prompt to fit before sending.
    if "messages" in fb_kwargs:
        fb_kwargs["messages"] = _truncate_for_fallback(fb_kwargs["messages"])
    print(
        f"⚠️  Primary AI unavailable ({primary_exc}) — failing over to fallback "
        f"provider ({provider['base_url']}, model={fb_kwargs['model']})"
    )
    try:
        return client.chat.completions.create(**fb_kwargs)
    except Exception as fb_exc:
        status = getattr(fb_exc, "status_code", None) or getattr(fb_exc, "status", None)
        # Some OpenAI-compatible providers reject tools/tool_choice with a 400.
        # Retry once without them so failover still produces an answer.
        if status == 400 and ("tools" in fb_kwargs or "tool_choice" in fb_kwargs):
            fb_kwargs.pop("tools", None)
            fb_kwargs.pop("tool_choice", None)
            print("⚠️  Fallback provider rejected tools — retrying without tools")
            return client.chat.completions.create(**fb_kwargs)
        # Still too large for the provider's token-per-minute cap — trim harder
        # (half the budget) and retry once so a reply still goes out.
        if status == 413 and "messages" in fb_kwargs:
            fb_kwargs["messages"] = _truncate_for_fallback(
                fb_kwargs["messages"], max_chars=max(4000, FALLBACK_MAX_PROMPT_CHARS // 2)
            )
            print("⚠️  Fallback provider 413 (request too large) — retrying with harder truncation")
            return client.chat.completions.create(**fb_kwargs)
        raise


def _chat_completion(**kwargs):
    """Create a chat completion on the primary provider, failing over through the
    configured fallback chain (FALLBACK → FALLBACK2) when the primary is unusable
    (quota/rate, 5xx outage, network/timeout, dead key). Re-raises if there is no
    usable fallback, the error isn't failover-eligible (e.g. a 400 bad request),
    or every provider in the chain also fails."""
    try:
        return client_ai.chat.completions.create(**kwargs)
    except Exception as exc:
        if not FALLBACK_PROVIDERS or not _should_failover(exc):
            raise
        last_exc = exc
        for provider in FALLBACK_PROVIDERS:
            try:
                return _call_fallback_provider(provider, kwargs, exc)
            except Exception as fb_exc:
                last_exc = fb_exc
                # Advance to the next provider on ANY failure — a 400/404 from a
                # fallback is usually provider-specific (decommissioned default
                # model, bad FALLBACK_MODEL name), which the next provider in the
                # chain (with its own model) may well fix. Re-raising here would
                # make FALLBACK2 unreachable behind a misconfigured FALLBACK.
                print(f"⚠️  Fallback provider ({provider['base_url']}) failed ({fb_exc}) — trying next in chain")
                continue
        # Every provider in the chain failed.
        raise last_exc


async def _achat(**kwargs):
    """Async wrapper for _chat_completion — runs the blocking OpenAI SDK call in
    a worker thread so the event loop (gateway heartbeat, poll loops, button
    interactions, other users' replies) never freezes while a request is in
    flight. Use this from ALL async code paths."""
    return await asyncio.to_thread(_chat_completion, **kwargs)


def _should_send_down_notice(channel_id: int) -> bool:
    """Throttle the 'AI is down' notice to once per channel per cooldown window so
    a dead quota doesn't spam an identical message on every incoming message."""
    now = datetime.now(timezone.utc).timestamp()
    if now - _ai_down_notice_at.get(channel_id, 0) >= AI_DOWN_NOTICE_COOLDOWN_SECS:
        _ai_down_notice_at[channel_id] = now
        return True
    return False


def resolve_ai_error_reply(channel_id: int, exc: Exception, force: bool = False) -> str | None:
    """Map an AI exception to a clean user-facing notice (never the raw error or a
    billing link). Logs the real error to stdout for ops. Returns None when a down
    notice was already sent to this channel recently — caller should stay quiet.
    Pass force=True to bypass the throttle (explicit @mentions and paths that
    already acknowledged the user must always get an answer, never silence)."""
    # AI unavailable (quota/rate, outage, timeout, dead key) — and either no
    # fallback or the fallback also failed. Show the clean "down" notice.
    if _should_failover(exc):
        print(f"⚠️  AI unavailable: {exc}")
        # Throttle check FIRST so the timestamp is always armed when due — a
        # short-circuiting `force or ...` skipped it on forced notices, letting
        # the very next unforced failure post a duplicate down notice seconds
        # later. `force` still bypasses suppression.
        if _should_send_down_notice(channel_id) or force:
            return AI_DOWN_NOTICE
        return None
    print(f"⚠️  AI call failed: {exc}")
    return GENERIC_ERROR_NOTICE


async def _run_chat(messages: list, allow_tools: bool = True) -> str:
    """Run a GPT-4o completion, resolving any tool calls, and return the final text.
    Tool round-trip messages stay local to this call so they never pollute the
    persisted channel history."""
    convo = list(messages)
    for _round in range(4):
        kwargs = {
            "model": "gpt-4o",
            "messages": convo,
            "max_tokens": 1024,
            "temperature": 0.7,
        }
        if allow_tools:
            kwargs["tools"] = TEAM_TOOLS
            kwargs["tool_choice"] = "auto"
        msg = (await _achat(**kwargs)).choices[0].message
        tool_calls = getattr(msg, "tool_calls", None)
        if not tool_calls:
            return _normalize_links((msg.content or "").strip())
        convo.append({
            "role": "assistant",
            "content": msg.content or "",
            "tool_calls": [
                {
                    "id": tc.id,
                    "type": "function",
                    "function": {"name": tc.function.name, "arguments": tc.function.arguments},
                }
                for tc in tool_calls
            ],
        })
        for tc in tool_calls:
            result = await _dispatch_tool(tc.function.name, tc.function.arguments)
            convo.append({"role": "tool", "tool_call_id": tc.id, "content": result})
    # Tool rounds exhausted — force a final answer with no further tools.
    msg = (await _achat(
        model="gpt-4o", messages=convo, max_tokens=1024, temperature=0.7
    )).choices[0].message
    return _normalize_links((msg.content or "").strip())


async def ask_openai_text(channel_id: int, user_text: str, username: str, is_staff: bool = False) -> tuple[str, bool]:
    """Standard text reply via GPT-4o. Returns (reply_text, needs_support_redirect)."""
    msgs = get_channel_messages(channel_id)
    system_prompt = build_system_prompt(is_staff=is_staff)

    msgs.append({"role": "user", "content": f"{username}: {user_text}"})
    trim_history(channel_id)
    touch_channel(channel_id)

    raw = await _run_chat([{"role": "system", "content": system_prompt}, *msgs])

    # Check if bot flagged this as needing support
    needs_support = "---SUPPORT_REDIRECT---" in raw
    reply = raw.replace("---SUPPORT_REDIRECT---", "").strip()

    msgs.append({"role": "assistant", "content": reply})
    trim_history(channel_id)
    touch_channel(channel_id)
    save_history_to_disk()
    return reply, needs_support


async def ask_openai_with_image(channel_id: int, user_text: str, username: str, image_bytes: bytes, media_type: str, is_staff: bool = False) -> tuple[str, bool]:
    """Send image + text to GPT-4o vision. Returns (reply_text, needs_support_redirect)
    — same contract as ask_openai_text, so escalations from screenshots (banned
    account, payment issue) trigger the support redirect instead of leaking the
    literal marker to the user."""
    msgs = get_channel_messages(channel_id)
    system_prompt = build_system_prompt(is_staff=is_staff)

    image_b64 = base64.b64encode(image_bytes).decode("utf-8")

    vision_message = {
        "role": "user",
        "content": [
            {
                "type": "image_url",
                "image_url": {
                    "url": f"data:{media_type};base64,{image_b64}"
                }
            },
            {
                "type": "text",
                "text": f"{username}: {user_text}" if user_text else f"{username} sent this image."
            }
        ]
    }

    raw = await _run_chat([
        {"role": "system", "content": system_prompt},
        *msgs,
        vision_message,
    ])
    needs_support = "---SUPPORT_REDIRECT---" in raw
    reply = raw.replace("---SUPPORT_REDIRECT---", "").strip()

    msgs = get_channel_messages(channel_id)
    msgs.append({"role": "user", "content": f"{username}: [sent an image] {user_text}"})
    msgs.append({"role": "assistant", "content": reply})
    trim_history(channel_id)
    touch_channel(channel_id)
    save_history_to_disk()
    return reply, needs_support


async def send_support_redirect(message: discord.Message):
    """Send a smart support redirect that tags roles and points to the support channel."""
    role_tags = " ".join([f"<@&{rid}>" for rid in SUPPORT_ROLES])
    embed = discord.Embed(
        description=(
            f"Hey {message.author.mention}, this one needs a human to sort out properly. 🙏\n\n"
            f"**Here's what you can do:**\n"
            f"1. Head over to <#{SUPPORT_CHANNEL_ID}> and create a support ticket\n"
            f"2. Or reach out directly via email: **info@africanfreefirecommunity.com**\n"
            f"3. Or join the AFC Discord: **{AFC_DISCORD_INVITE}**\n\n"
            f"Our support team has been notified 👇"
        ),
        color=0x00A550
    )
    embed.set_footer(text="African Freefire Community  •  africanfreefirecommunity.com")
    await message.channel.send(embed=embed)
    await message.channel.send(f"🔔 {role_tags} — support needed here.")


async def transcribe_audio(audio_bytes: bytes, filename: str) -> str:
    """Transcribe audio using OpenAI Whisper. Runs in a worker thread — the
    multipart upload + transcription can take many seconds and must not freeze
    the event loop."""
    ext = os.path.splitext(filename)[1].lower()

    def _do() -> str:
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(audio_bytes)
            tmp_path = tmp.name
        try:
            with open(tmp_path, "rb") as audio_file:
                transcript = client_ai.audio.transcriptions.create(
                    model="whisper-1",
                    file=audio_file,
                )
            return transcript.text
        finally:
            os.unlink(tmp_path)

    return await asyncio.to_thread(_do)


# ── Server management helpers ────────────────────────────────────────────────
# Tracks pending delete confirmations: {original_message_id: channel_to_delete_id}
pending_deletions: dict[int, int] = {}

# Active transcription sessions: guild_id → session info dict
active_transcriptions: dict[int, dict] = {}
# Pending stage transcription prompts: stage_channel_id → info dict
pending_stage_prompts: dict[int, dict] = {}

# Tracks who is currently waiting for a confirmation reply per channel.
# {channel_id: user_id} — only that specific user's non-command messages are suppressed.
_awaiting_confirmation: dict[int, int] = {}

def has_admin_role(member: discord.Member) -> bool:
    """Check if member has one of the announcement/admin roles."""
    return any(role.id in ANNOUNCE_ROLES for role in member.roles)


def parse_role_command(text: str):
    """
    Detect give/remove role commands for a single user.
    Returns (action, user_id, role_id) or None.
    """
    action = None
    if re.search(r"\bgive\b|\badd\b|\bassign\b|\bgrant\b", text, re.IGNORECASE):
        action = "give"
    elif re.search(r"\bremove\b|\btake\b|\brevoke\b|\bstrip\b", text, re.IGNORECASE):
        action = "remove"

    if not action:
        return None

    user_match = re.search(r"<@!?(\d+)>", text)
    role_match = re.search(r"<@&(\d+)>", text)

    if not user_match or not role_match:
        return None

    return action, int(user_match.group(1)), int(role_match.group(1))


def parse_mass_role_command(text: str):
    """
    Detect mass role operations.
    Returns dict: { action, target_role_id, condition_role_id } or None.

    Actions:
      remove_all   — remove a role from everyone who has it
      remove_if    — remove role_A from everyone who has role_B
      give_all     — give a role to everyone in the server
      give_if      — give role_A to everyone who has role_B

    Examples:
      remove @Role from everyone
      strip @Role from all members
      remove @Role from everyone who has @OtherRole
      take @RoleA from all members with @RoleB
      give @Role to everyone
      give @Role to everyone with @OtherRole
      assign @Role to all members who have @OtherRole
    """
    # Must have a role mention
    role_matches = re.findall(r"<@&(\d+)>", text)
    if not role_matches:
        return None

    # Must be a mass operation
    if not re.search(
        r"\beveryone\b|\ball\s+(members|players|users)\b|\ball\b",
        text, re.IGNORECASE
    ):
        return None

    result = {
        "action": None,
        "target_role_id": int(role_matches[0]),
        "condition_role_id": int(role_matches[1]) if len(role_matches) > 1 else None,
    }

    # Conditional (has second role)
    has_condition = bool(re.search(
        r"\bwho\s+has\b|\bwith\b|\bwho\s+have\b|\bthat\s+have\b|\bthat\s+has\b",
        text, re.IGNORECASE
    ))

    if re.search(r"\bgive\b|\bassign\b|\bgrant\b|\badd\b", text, re.IGNORECASE):
        result["action"] = "give_if" if has_condition else "give_all"
    elif re.search(r"\bremove\b|\bstrip\b|\btake\b|\brevoke\b", text, re.IGNORECASE):
        result["action"] = "remove_if" if has_condition else "remove_all"
    else:
        return None

    return result


def parse_role_manage_command(text: str):
    """
    Detect role creation, deletion, renaming, recoloring, permission edits.
    Returns dict: { action, role_id, name, color, permissions } or None.

    Actions: create, delete, rename, recolor, edit_perms

    Examples:
      create role Veteran
      create role Moderator with color #FF0000
      delete role @OldRole
      rename @Role to Senior Staff
      change @Role color to green
      recolor @Role to #FFD700
      make @Role mentionable
      make @Role not mentionable
      make @Role hoisted
    """
    if not re.search(
        r"\bcreate\s+role\b|\bmake\s+a?\s*role\b"
        r"|\bdelete\s+role\b|\bremove\s+role\b"
        r"|\brename\s+.*(role|<@&)\b"
        r"|\brecolor\b|\bchange\s+.*color\b"
        r"|\bmake\s+.*\b(mentionable|hoisted|pingable)\b"
        r"|\bmake\s+.*\bnot\s+(mentionable|hoisted|pingable)\b",
        text, re.IGNORECASE
    ):
        return None

    result = {
        "action": None,
        "role_id": None,
        "name": None,
        "color": None,
        "mentionable": None,
        "hoisted": None,
    }

    role_match = re.search(r"<@&(\d+)>", text)
    if role_match:
        result["role_id"] = int(role_match.group(1))

    # Color extraction — hex or color name
    hex_match = re.search(r"#([0-9A-Fa-f]{6})", text)
    if hex_match:
        result["color"] = int(hex_match.group(1), 16)
    else:
        color_names = {
            "red": 0xFF0000, "green": 0x00A550, "blue": 0x0000FF,
            "gold": 0xFFD700, "yellow": 0xFFFF00, "orange": 0xFF6600,
            "purple": 0x800080, "pink": 0xFF69B4, "white": 0xFFFFFF,
            "black": 0x000001, "cyan": 0x00FFFF, "teal": 0x008080,
            "silver": 0xC0C0C0, "grey": 0x808080, "gray": 0x808080,
        }
        for name, val in color_names.items():
            if re.search(rf"\b{name}\b", text, re.IGNORECASE):
                result["color"] = val
                break

    # Mentionable / hoisted
    if re.search(r"\bnot\s+(mentionable|pingable)\b", text, re.IGNORECASE):
        result["mentionable"] = False
    elif re.search(r"\b(mentionable|pingable)\b", text, re.IGNORECASE):
        result["mentionable"] = True

    if re.search(r"\bnot\s+hoisted\b", text, re.IGNORECASE):
        result["hoisted"] = False
    elif re.search(r"\bhoisted\b", text, re.IGNORECASE):
        result["hoisted"] = True

    # Action: DELETE
    if re.search(r"\bdelete\s+role\b|\bremove\s+role\b", text, re.IGNORECASE):
        result["action"] = "delete"
        return result

    # Action: RENAME
    rename_match = re.search(r"rename\s+(?:<@&\d+>|\S+)\s+to\s+(.+)$", text, re.IGNORECASE)
    if rename_match:
        result["action"] = "rename"
        result["name"] = rename_match.group(1).strip()
        return result

    # Action: RECOLOR
    if re.search(r"\brecolor\b|\bchange\s+.*color\b|\bcolor\s+.*to\b", text, re.IGNORECASE):
        result["action"] = "recolor"
        return result

    # Action: EDIT PROPS (mentionable/hoisted)
    if result["mentionable"] is not None or result["hoisted"] is not None:
        result["action"] = "edit_props"
        return result

    # Action: CREATE
    if re.search(r"\bcreate\s+role\b|\bmake\s+a?\s*role\b", text, re.IGNORECASE):
        result["action"] = "create"
        # Extract role name — everything after "role"
        name_match = re.search(r"(?:role\s+)([\w\s\-]+?)(?:\s+with|\s+color|\s+#|$)", text, re.IGNORECASE)
        if name_match:
            result["name"] = name_match.group(1).strip()
        return result

    return None


def parse_permission_command(text: str):
    """
    Detect permission edit commands.
    Returns (channel_or_cat_id, is_category, target_type, target_id, perm_name, allow) or None.
    Examples:
      lock <#channel>
      unlock <#channel>
      hide <#channel> from @role
      show <#channel> to @role
      deny send_messages in <#channel> for @role
      allow view_channel in <#channel> for @role
    """
    # Lock / unlock shortcuts
    if re.search(r"\block\b", text, re.IGNORECASE):
        ch = re.search(r"<#(\d+)>", text)
        if ch:
            return int(ch.group(1)), False, "everyone", None, "send_messages", False
    if re.search(r"\bunlock\b", text, re.IGNORECASE):
        ch = re.search(r"<#(\d+)>", text)
        if ch:
            return int(ch.group(1)), False, "everyone", None, "send_messages", True

    # Hide / show shortcuts
    if re.search(r"\bhide\b", text, re.IGNORECASE):
        ch = re.search(r"<#(\d+)>", text)
        role = re.search(r"<@&(\d+)>", text)
        if ch:
            return int(ch.group(1)), False, "role" if role else "everyone", int(role.group(1)) if role else None, "view_channel", False
    if re.search(r"\bshow\b|\bunhide\b", text, re.IGNORECASE):
        ch = re.search(r"<#(\d+)>", text)
        role = re.search(r"<@&(\d+)>", text)
        if ch:
            return int(ch.group(1)), False, "role" if role else "everyone", int(role.group(1)) if role else None, "view_channel", True

    return None


def parse_delete_command(text: str):
    """
    Detect delete CHANNEL commands only.
    Returns channel_id or None.
    """
    # Must explicitly say "channel" or "remove channel" — not just "delete messages"
    if not re.search(r"\bdelete\s+(the\s+)?channel\b|\bremove\s+channel\b|\bdelete\s+<#\d+>\s*$", text, re.IGNORECASE):
        return None

    # Never fire if talking about messages
    if re.search(r"\bmessages?\b|\bpurge\b|\bclear\b", text, re.IGNORECASE):
        return None

    ch = re.search(r"<#(\d+)>", text)
    if ch:
        return int(ch.group(1))
    return None


def parse_purge_command(text: str):
    """
    Detect message purge/clear commands.
    Returns dict with keys: mode, channel_id, amount, keyword, user_ids, role_id
    or None if not a purge command.

    Modes:
      count   — delete X messages
      keyword — delete messages containing a word/phrase
      user    — delete all messages from one or more specific users
      role    — delete all messages from users who have a specific role
      all     — delete ALL messages in the channel

    Examples:
      purge 50 messages in <#channel>
      clear 10 in <#channel>
      delete 20 messages in <#channel>
      delete the last 3 messages in <#channel>
      purge messages containing "spam" in <#channel>
      clear messages from @user in <#channel>
      delete messages from @user1 and @user2 in <#channel>
      delete messages from users with @role in <#channel>
      purge messages from @role members in <#channel>
      purge all messages in <#channel>
    """
    if not re.search(
        r"\bpurge\b|\bclear\b|\bclean\b|\bwipe\b"
        r"|\bdelete\s+(the\s+)?(last\s+)?\d+\b"
        r"|\bdelete\s+all\b|\bdelete\s+messages\b"
        r"|\bremove\s+(the\s+)?(last\s+)?\d+\s+messages\b",
        text, re.IGNORECASE
    ):
        return None

    result = {
        "mode":       None,
        "channel_id": None,
        "amount":     None,
        "keyword":    None,
        "user_ids":   [],     # list — supports multiple users
        "role_id":    None,
    }

    # Channel (optional — defaults to current channel if not specified)
    ch_match = re.search(r"<#(\d+)>", text)
    result["channel_id"] = int(ch_match.group(1)) if ch_match else None

    # Mode: ALL
    if re.search(r"\ball\b", text, re.IGNORECASE):
        result["mode"] = "all"
        return result

    # Mode: ROLE — "from users with @role" / "from @role members" / "from @role"
    role_match = re.search(r"<@&(\d+)>", text)
    if role_match and re.search(
        r"\bfrom\b|\bby\b|\bwith\b|\bmembers\b|\bwho\s+have\b|\bwho\s+has\b",
        text, re.IGNORECASE
    ):
        result["mode"]    = "role"
        result["role_id"] = int(role_match.group(1))
        return result

    # Mode: USER — one or more @mentions or plain user IDs with from/by context
    # Strip role mentions and channel mentions before collecting user mentions
    text_no_roles = re.sub(r"<@&\d+>|<#\d+>", "", text)
    user_ids = [int(uid) for uid in re.findall(r"<@!?(\d+)>", text_no_roles)]

    # Also handle plain user IDs like "from user 563399749231706123"
    if not user_ids:
        plain_ids = re.findall(
            r"(?:from\s+(?:this\s+)?(?:user\s+)?|by\s+(?:user\s+)?)(\d{15,19})\b",
            text, re.IGNORECASE
        )
        user_ids = [int(i) for i in plain_ids]

    if user_ids and re.search(r"\bfrom\b|\bby\b|\bthis\s+user\b|\buser\b", text, re.IGNORECASE):
        result["mode"]     = "user"
        result["user_ids"] = user_ids
        return result

    # Mode: KEYWORD
    kw_match = re.search(r'containing\s+"([^"]+)"', text, re.IGNORECASE)
    if not kw_match:
        kw_match = re.search(r"containing\s+'([^']+)'", text, re.IGNORECASE)
    if not kw_match:
        kw_match = re.search(r"containing\s+(\S+)", text, re.IGNORECASE)
    if not kw_match:
        kw_match = re.search(r"with\s+(?:the\s+word\s+)?[\"']?(\w+)[\"']?", text, re.IGNORECASE)
    if kw_match:
        result["mode"]    = "keyword"
        result["keyword"] = kw_match.group(1)
        return result

    # Mode: COUNT — handles "delete 20", "delete the last 3", "clear last 50"
    num_match = re.search(r"(?:last\s+)?(\d+)", text, re.IGNORECASE)
    if num_match:
        result["mode"]   = "count"
        result["amount"] = int(num_match.group(1))
        return result

    # Nothing matched clearly — don't guess, return None
    return None


def parse_create_command(text: str):
    """
    Detect create channel/category commands.
    Must have explicit "create" or "make" AND "channel", "category", or "voice channel"
    together — avoids false matches on normal messages.
    """
    # Must have a clear creation intent word
    if not re.search(r"\bcreate\b", text, re.IGNORECASE):
        # "make" only counts if directly followed by channel/category/voice
        if not re.search(r"\bmake\s+(a\s+)?(text\s+|voice\s+|private\s+)?channel\b|\bmake\s+(a\s+)?category\b", text, re.IGNORECASE):
            return None

    # Must explicitly mention channel or category
    if not re.search(r"\bchannel\b|\bcategory\b", text, re.IGNORECASE):
        return None

    result = {
        "type": "text",
        "name": None,
        "category_id": None,
        "private": False,
        "role_id": None,
    }

    if re.search(r"\bcategory\b", text, re.IGNORECASE):
        result["type"] = "category"
    elif re.search(r"\bvoice\b", text, re.IGNORECASE):
        result["type"] = "voice"
    else:
        result["type"] = "text"

    if re.search(r"\bprivate\b|\bsecret\b|\bhidden\b", text, re.IGNORECASE):
        result["private"] = True

    role_match = re.search(r"<@&(\d+)>", text)
    if role_match:
        result["role_id"] = int(role_match.group(1))

    cat_match = re.search(r"in\s+<#(\d+)>", text, re.IGNORECASE)
    if cat_match:
        result["category_id"] = int(cat_match.group(1))

    # Extract the name — strip command keywords
    name = text
    for pattern in [
        r"\bcreate\b", r"\bmake\b", r"\badd\b",
        r"\btext\b", r"\bvoice\b", r"\bcategory\b", r"\bchannel\b",
        r"\bprivate\b", r"\bsecret\b", r"\bhidden\b",
        r"\bfor\b", r"\bin\b", r"\bthe\b", r"\ba\b", r"\ban\b",
        r"<@&\d+>", r"<#\d+>",
    ]:
        name = re.sub(pattern, "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s+", "-", name.strip()).lower().strip("-")

    if not name:
        return None

    result["name"] = name
    return result


def parse_edit_command(text: str):
    """
    Detect edit message commands using fully natural language.
    Works with explicit commands AND pure feedback with no command words.

    Returns dict: { mode, channel_id, message_id, instruction } or None.

    Modes:
      last      — edit the bot's last reply in this channel
      id        — edit a specific message by its ID
      announce  — edit the last announcement the bot sent in a channel

    Examples — ALL of these work:
      too many emojis
      make it more formal
      change the tone, more hype
      that sounds off, fix it
      remove the last sentence
      way too long, shorten it
      less emojis
      more professional
      edit last reply, less emojis
      fix your last message
      edit last announcement in <#channel>, make it shorter
      edit message 123456789 — remove the last sentence
    """

    # ── ANNOUNCEMENT GUARD — never fire on clear announcement commands ────────
    if re.search(
        r"\bformulate\s+(an?\s+)?announcement\b"
        r"|\bnot\s+editing\b|\bnew\s+announcement\b"
        r"|\bsend\s+(an?\s+)?announcement\b"
        r"|\bformulate\s+and\s+send\b",
        text, re.IGNORECASE
    ):
        return None

    # Explicit edit/fix keywords
    explicit = bool(re.search(
        r"\bedit\b|\bupdate\b|\bchange\b|\bcorrect\b|\bfix\b|\brewrite\b|\brevise\b|\bshorten\b|\blonger\b",
        text, re.IGNORECASE
    ))

    # Pure feedback patterns — no command word needed
    pure_feedback = bool(re.search(
        r"\btoo\s+(many|much|long|short|formal|casual|hype|stiff|wordy)\b"
        r"|\bless\s+\w+\b|\bmore\s+\w+\b"
        r"|\bsounds?\s+(off|wrong|bad|weird|stiff|too)\b"
        r"|\bmake\s+it\b|\btone\s+(it|down|up)\b"
        r"|\b(remove|add|drop)\s+the\b"
        r"|\bway\s+too\b|\bnot\s+enough\b"
        r"|\bfeel\s+(more|less)\b"
        r"|\bthat\s+(phrase|word|line|sentence|part)\b",
        text, re.IGNORECASE
    ))

    if not explicit and not pure_feedback:
        return None

    # Also require some message context unless it's pure feedback
    if explicit and not re.search(r"\bmessage\b|\breply\b|\bannouncement\b|\blast\b|\bit\b", text, re.IGNORECASE):
        return None

    result = {
        "mode": None,
        "channel_id": None,
        "message_id": None,
        "instruction": text,  # pass full natural instruction to GPT
    }

    # Extract TARGET channel (where the message to edit lives)
    ch_match = re.search(r"(?:in\s+)?<#(\d+)>", text)
    result["channel_id"] = int(ch_match.group(1)) if ch_match else None

    # Mode: specific message ID
    # First strip all channel/user mentions so their IDs don't get confused as message IDs
    text_no_mentions = re.sub(r"<[#@!&]\d+>", "", text)

    id_match = re.search(
        r"(?:message\s+(?:with\s+(?:the\s+)?id\s+)?|message\s+id\s+|id\s+)"
        r"<?(\d{10,})>?",
        text_no_mentions, re.IGNORECASE
    )
    if not id_match:
        id_match = re.search(r"<(\d{15,})>", text_no_mentions)
    if not id_match:
        id_match = re.search(r"\b(\d{17,19})\b", text_no_mentions)

    if id_match:
        result["mode"] = "id"
        result["message_id"] = int(id_match.group(1))
        return result

    # Mode: last announcement — supports cross-channel via <#channel>
    if re.search(r"\bannouncement\b", text, re.IGNORECASE):
        result["mode"] = "announce"
        return result

    # Default: last reply
    # If a channel is specified, look for last bot message in THAT channel
    result["mode"] = "last"
    return result


async def ai_rewrite(original_text: str, instruction: str) -> str:
    """Use GPT-4o to rewrite a message based on feedback/instruction."""
    response = await _achat(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": (
                    "You are an editor for the African Freefire Community (AFC) Discord bot. "
                    "You will be given an original message and an instruction on how to improve it. "
                    "Rewrite the message based on the instruction. "
                    "Output ONLY the rewritten message — no explanations, no preamble, no quotes around it. "
                    "Preserve all Discord formatting like **bold**, links, and mentions unless told otherwise. "
                    "Keep the same general meaning and information unless told to change it."
                )
            },
            {
                "role": "user",
                "content": f"Original message:\n{original_text}\n\nInstruction: {instruction}\n\nRewrite:"
            }
        ],
        max_tokens=1024,
        temperature=0.7,
    )
    return response.choices[0].message.content.strip()


# Tracks last bot message per channel: {channel_id: message_id}
last_bot_messages: dict[int, int] = {}
# Tracks last bot announcement per channel: {channel_id: message_id}
last_bot_announcements: dict[int, int] = {}


# ── Voice transcription ───────────────────────────────────────────────────────

WHISPER_PROMPT = (
    "Transcribe exactly as spoken. This is a Nigerian/African gaming community voice session. "
    "Include Nigerian Pidgin English, African slang, informal speech, profanity, and all words "
    "exactly as said — do not censor, clean up, or modify any language. "
    "Common words you may hear: wahala, abeg, oya, sabi, dey, na, wey, comot, chop, sharp sharp, "
    "e don do, no wahala, guy, bro, fam, gbege, ginger, level up, carry last, shine your eye."
)


# discord.py has no built-in voice receive — it comes from the
# discord-ext-voice-recv extension (requirements.txt). Feature-detect so the
# transcription flows degrade gracefully (clean "not supported" message, no
# connect) if the extension is missing, instead of crashing AFTER connect()
# and stranding a zombie voice connection until restart.
try:
    from discord.ext import voice_recv  # discord-ext-voice-recv
    import audioop  # stdlib through 3.12 (prod pins 3.11) — PCM downmix/resample
    VOICE_RECEIVE_SUPPORTED = True
except Exception as _vr_exc:
    voice_recv = None
    VOICE_RECEIVE_SUPPORTED = False
    print(f"⚠️  Voice receive unavailable ({_vr_exc}) — transcription disabled")


if VOICE_RECEIVE_SUPPORTED:
    class PerUserTranscriptionSink(voice_recv.AudioSink):
        """Accumulates each speaker's audio as 16kHz mono s16 PCM.

        Downsampling from the decoder's 48kHz stereo (~6x smaller) matters for
        long stage sessions: RAM stays bounded and each Whisper upload fits the
        API's 25MB file cap. write() runs on the voice receive thread, so it
        only does cheap C-level audioop calls and bytearray appends — and must
        never raise, or the whole receive loop dies."""

        SRC_RATE  = 48000   # discord.opus.Decoder output rate
        SRC_WIDTH = 2       # 16-bit samples
        DST_RATE  = 16000   # Whisper's native rate
        # Hard per-speaker cap: 1 hour of actual speech (~115MB). Without it a
        # marathon stage session can OOM the small EC2 box.
        MAX_PCM_BYTES_PER_USER = 60 * 60 * 16000 * 2

        def __init__(self):
            super().__init__()
            self._pcm: dict[int, bytearray] = {}
            self._ratecv_state: dict[int, object] = {}
            self._capped: set[int] = set()
            # Monotonic timestamp of the last received voice packet — the
            # silence watchdog reads this to auto-end quiet sessions.
            self.last_audio_at: float = time.monotonic()

        def wants_opus(self) -> bool:
            return False

        def write(self, user, data) -> None:
            if user is None:
                return
            self.last_audio_at = time.monotonic()
            try:
                mono = audioop.tomono(data.pcm, self.SRC_WIDTH, 0.5, 0.5)
                converted, self._ratecv_state[user.id] = audioop.ratecv(
                    mono, self.SRC_WIDTH, 1, self.SRC_RATE, self.DST_RATE,
                    self._ratecv_state.get(user.id),
                )
                buf = self._pcm.setdefault(user.id, bytearray())
                if len(buf) < self.MAX_PCM_BYTES_PER_USER:
                    buf.extend(converted)
                elif user.id not in self._capped:
                    self._capped.add(user.id)
                    print(f"⚠️  Transcription buffer cap (1h speech) reached for user {user.id} — further audio dropped")
            except Exception:
                pass

        def cleanup(self) -> None:
            pass

        def take_audio(self) -> dict[int, bytes]:
            """Call after stop_listening(): {user_id: 16kHz mono s16 PCM}.

            stop_listening() does NOT join the library's router thread — one
            final write() can still land while this runs, so iterate a snapshot
            (list(...)) to avoid 'dict changed size during iteration'. Callers
            also sleep briefly first; any write after the snapshot is dropped
            (sub-second tail, acceptable)."""
            out = {uid: bytes(buf) for uid, buf in list(self._pcm.items()) if buf}
            self._pcm.clear()
            self._ratecv_state.clear()
            return out
else:
    PerUserTranscriptionSink = None


async def _finalize_transcription(guild_id: int, notice: str | None):
    """Single teardown path for a transcription session — used by the stop
    command, the silence watchdog, and the kicked-from-voice handler. Pops the
    session first (so concurrent/double finalize is a no-op), stops listening,
    disconnects, then posts the transcript from whatever was captured."""
    session = active_transcriptions.pop(guild_id, None)
    if not session:
        return
    vc, sink = session["vc"], session.get("sink")
    try:
        if hasattr(vc, "stop_listening"):
            vc.stop_listening()
    except Exception as e:
        print(f"⚠️  stop_listening failed: {e}")
    try:
        await vc.disconnect(force=True)
    except Exception as e:
        print(f"⚠️  voice disconnect failed: {e}")
    if notice:
        try:
            await session["text_channel"].send(notice)
        except Exception:
            pass
    # Give the library's detached router thread a beat to flush its final
    # write() before we read the buffers (see take_audio()).
    await asyncio.sleep(0.5)
    audio = sink.take_audio() if sink else {}

    async def _finish():
        try:
            await _process_and_send_transcript(
                audio, session["text_channel"], session["requester"], session["start_time"]
            )
        except Exception as e:
            print(f"⚠️  Transcript generation failed: {e}")

    asyncio.create_task(_finish())


def _start_silence_watchdog(guild_id: int, sink):
    """Auto-end the session after TRANSCRIPTION_SILENCE_TIMEOUT_SECS with no
    incoming voice packets, posting whatever was captured up to that point."""
    async def _watch():
        try:
            while True:
                await asyncio.sleep(30)
                session = active_transcriptions.get(guild_id)
                if not session or session.get("sink") is not sink:
                    return  # session ended (stop/kick) or replaced by a new one
                if time.monotonic() - sink.last_audio_at >= TRANSCRIPTION_SILENCE_TIMEOUT_SECS:
                    mins = TRANSCRIPTION_SILENCE_TIMEOUT_SECS // 60
                    await _finalize_transcription(
                        guild_id,
                        f"🔇 No one has spoken for {mins} minutes — stopped transcribing.",
                    )
                    return
        except Exception as e:
            print(f"⚠️  Silence watchdog error: {e}")

    asyncio.create_task(_watch())


def has_transcription_role(member: discord.Member) -> bool:
    return any(role.id in TRANSCRIPTION_ROLES for role in member.roles)


def parse_transcription_command(text: str):
    """
    Detect transcription start/stop commands.
    Returns {"action": "start"|"stop", "channel_id": int|None} or None.
    """
    t = text.lower()
    if re.search(r"\bstop\b.{0,20}(transcri|record)", t) or re.search(r"(transcri|record).{0,20}\bstop\b", t):
        return {"action": "stop", "channel_id": None}
    if re.search(r"\b(transcribe|transcription|record\s+(this\s+)?(call|meeting|stage|voice|session))\b", t):
        ch_match = re.search(r"<#(\d+)>", text)
        return {"action": "start", "channel_id": int(ch_match.group(1)) if ch_match else None}
    return None


# Whisper rejects uploads over 25MB; 12 minutes of 16kHz mono s16 PCM wrapped in
# a wav is ~23MB, so long sessions are transcribed in 12-minute chunks per user
# and the segment timestamps re-offset by the chunk position.
TRANSCRIBE_CHUNK_SECS = 12 * 60
_PCM_BYTES_PER_SEC = 16000 * 2  # 16kHz mono, 2 bytes per sample


async def _process_and_send_transcript(audio_data: dict, text_channel, requester, start_time):
    """Transcribe each speaker's 16kHz mono PCM (from PerUserTranscriptionSink
    .take_audio()) and send a merged document."""
    import io
    import wave as wave_mod

    if not audio_data:
        await text_channel.send("⚠️ No audio was captured.")
        return

    await text_channel.send("⏳ Transcribing... this may take a moment.")

    all_segments = []
    chunk_bytes = TRANSCRIBE_CHUNK_SECS * _PCM_BYTES_PER_SEC

    for user_id, pcm in audio_data.items():
        user = bot.get_user(user_id)
        username = user.display_name if user else f"User {user_id}"
        try:
            for ci in range(0, len(pcm), chunk_bytes):
                chunk = pcm[ci:ci + chunk_bytes]
                if len(chunk) < _PCM_BYTES_PER_SEC // 2:
                    continue  # under half a second — noise, not speech
                buf = io.BytesIO()
                with wave_mod.open(buf, "wb") as wf:
                    wf.setnchannels(1)
                    wf.setsampwidth(2)
                    wf.setframerate(16000)
                    wf.writeframes(chunk)
                buf.seek(0)

                def _transcribe(b=buf):
                    return client_ai.audio.transcriptions.create(
                        model="whisper-1",
                        file=("audio.wav", b, "audio/wav"),
                        response_format="verbose_json",
                        timestamp_granularities=["segment"],
                        prompt=WHISPER_PROMPT,
                    )

                result = await asyncio.to_thread(_transcribe)
                offset_secs = ci // _PCM_BYTES_PER_SEC
                for seg in (result.segments or []):
                    text = seg.text.strip()
                    if text:
                        all_segments.append({
                            "username": username,
                            "start":    seg.start + offset_secs,
                            "text":     text,
                        })
        except Exception as e:
            print(f"⚠️  Transcription failed for user {user_id}: {e}")

    if not all_segments:
        await text_channel.send("⚠️ No speech was detected or all transcriptions failed.")
        return

    all_segments.sort(key=lambda x: x["start"])

    # ── Format the document ───────────────────────────────────────────────────
    duration    = datetime.now(timezone.utc) - start_time
    total_secs  = int(duration.total_seconds())
    h, rem      = divmod(total_secs, 3600)
    m, s        = divmod(rem, 60)
    speakers    = sorted({seg["username"] for seg in all_segments})

    lines = [
        "=" * 60,
        "AFC VOICE SESSION TRANSCRIPT",
        f"Date     : {start_time.strftime('%Y-%m-%d')}",
        f"Started  : {start_time.strftime('%H:%M UTC')}",
        f"Duration : {h:02d}h {m:02d}m {s:02d}s",
        f"Recorded : {requester}",
        f"Speakers : {', '.join(speakers)}",
        "=" * 60,
        "",
    ]

    prev_speaker = None
    for seg in all_segments:
        ts_m, ts_s = divmod(int(seg["start"]), 60)
        ts         = f"[{ts_m:02d}:{ts_s:02d}]"
        if seg["username"] != prev_speaker:
            lines.append(f"\n{seg['username']}  {ts}")
            prev_speaker = seg["username"]
        lines.append(f"  {seg['text']}")

    transcript_text = "\n".join(lines)
    file_date       = start_time.strftime("%Y-%m-%d_%H-%M")

    transcript_file = discord.File(
        fp=io.BytesIO(transcript_text.encode("utf-8")),
        filename=f"transcript_{file_date}.txt",
    )

    embed = discord.Embed(
        title="🎙️ Session Transcript Ready",
        description=(
            f"**Duration:** {h:02d}h {m:02d}m {s:02d}s\n"
            f"**Speakers:** {len(speakers)}\n"
            f"**Recorded by:** {requester.mention}"
        ),
        color=0x00A550,
    )
    embed.set_footer(text="African Freefire Community  •  africanfreefirecommunity.com")
    embed.timestamp = datetime.now(timezone.utc)

    await text_channel.send(embed=embed, file=transcript_file)


@bot.event
async def on_stage_instance_create(stage: discord.StageInstance):
    """When a stage is created, ask mods if the bot should transcribe it."""
    # Don't offer transcription at all when this build can't record voice —
    # prompting mods and then failing after joining is worse than silence.
    if not VOICE_RECEIVE_SUPPORTED:
        return
    mods_channel = bot.get_channel(MODS_CHANNEL_ID)
    if not mods_channel:
        return

    # Find the creator via audit logs
    creator = None
    try:
        async for entry in stage.guild.audit_logs(
            action=discord.AuditLogAction.stage_instance_create, limit=1
        ):
            if (datetime.now(timezone.utc) - entry.created_at).total_seconds() < 15:
                creator = entry.user
            break
    except Exception:
        pass

    creator_mention = creator.mention if creator else "Someone"

    prompt_msg = await mods_channel.send(
        f"🎙️ {creator_mention} has started a stage in {stage.channel.mention}.\n"
        f"Should I join and transcribe the session? Reply **`yes`** or **`no`**."
    )

    pending_stage_prompts[stage.channel.id] = {"message_id": prompt_msg.id}

    def stage_confirm_check(m):
        return (
            m.channel.id == MODS_CHANNEL_ID
            and m.content.lower().strip() in ("yes", "no", "y", "n")
            and isinstance(m.author, discord.Member)
            and has_transcription_role(m.author)
        )

    try:
        reply = await bot.wait_for("message", check=stage_confirm_check, timeout=120.0)

        if reply.content.lower().strip() in ("yes", "y"):
            vc = None
            try:
                vc         = await stage.channel.connect(cls=voice_recv.VoiceRecvClient)
                sink       = PerUserTranscriptionSink()
                start_time = datetime.now(timezone.utc)
                requester  = reply.author

                active_transcriptions[stage.guild.id] = {
                    "vc":           vc,
                    "sink":         sink,
                    "text_channel": mods_channel,
                    "requester":    requester,
                    "start_time":   start_time,
                }

                vc.listen(sink)
                _start_silence_watchdog(stage.guild.id, sink)
                await mods_channel.send(
                    f"✅ Joined {stage.channel.mention} and started transcribing.\n"
                    f"Say `@bot stop transcribing` when the session is done "
                    f"(I'll also stop on my own after {TRANSCRIPTION_SILENCE_TIMEOUT_SECS // 60} minutes of silence)."
                )
            except Exception as e:
                # Never strand a live voice connection when a step after a
                # successful connect() fails — the "stop" command can't reach it
                # (no session entry) and the bot would sit in the stage forever.
                if vc is not None:
                    try:
                        await vc.disconnect(force=True)
                    except Exception:
                        pass
                    active_transcriptions.pop(stage.guild.id, None)
                await mods_channel.send(f"⚠️ Couldn't start transcribing: {e}")
        else:
            await mods_channel.send("👍 Got it — not transcribing this stage.")

    except asyncio.TimeoutError:
        await mods_channel.send("⏱️ No response in 2 minutes — not transcribing.")
    finally:
        pending_stage_prompts.pop(stage.channel.id, None)


@bot.event
async def on_voice_state_update(member: discord.Member, before: discord.VoiceState, after: discord.VoiceState):
    """If the BOT is disconnected from voice mid-session (stage ended, kicked,
    channel deleted) without anyone saying 'stop transcribing', finalize the
    transcription instead of leaking the session — a leaked entry would lock
    out every future start with 'Already transcribing' until restart, and the
    captured audio would be lost."""
    if bot.user is None or member.id != bot.user.id:
        return
    if before.channel is None or after.channel is not None:
        return  # not a disconnect
    if member.guild.id not in active_transcriptions:
        return
    await _finalize_transcription(
        member.guild.id,
        "⚠️ I was disconnected from the voice channel — generating the transcript from what was captured.",
    )


# ── Slash commands ────────────────────────────────────────────────────────────
# Discord-native application commands: typing "/" lists them with descriptions,
# and they show on the bot's profile. Each is a thin adapter that synthesizes
# the equivalent natural-language command and feeds it through the SAME
# _handle_message pipeline the @mention flow uses — zero duplicated logic, and
# follow-up interactions (announcement previews, purge confirmations) work
# exactly as they do today. The natural-language commands remain available.

class _SlashCommandShim:
    """Adapts a slash Interaction into the message shape _handle_message expects."""

    def __init__(self, interaction: discord.Interaction, content: str, attachments: list = None):
        self.id = interaction.id
        self.author = interaction.user
        self.channel = interaction.channel
        self.guild = interaction.guild
        self.content = content
        self.attachments = attachments or []
        self.mentions = [bot.user]   # force the mention path (skip the classifier)
        self.reference = None
        self._from_slash = True

    async def reply(self, content=None, **kwargs):
        kwargs.pop("mention_author", None)
        return await self.channel.send(content, **kwargs)


async def _dispatch_slash(interaction: discord.Interaction, content: str, attachments: list = None):
    """Ack the slash command ephemerally, then run the synthesized command
    through the normal message pipeline. Prompts (previews, confirmations)
    appear in the channel just like the @mention flow."""
    await interaction.response.send_message(f"➡️ Running: `{content[:180]}`", ephemeral=True)
    shim = _SlashCommandShim(interaction, content, attachments)

    async def _run():
        try:
            await _handle_message(shim)
        except Exception as e:
            print(f"⚠️  Slash command failed ({content[:60]}): {e}")
            try:
                await interaction.followup.send(GENERIC_ERROR_NOTICE, ephemeral=True)
            except Exception:
                pass

    asyncio.create_task(_run())


@tree.command(name="help", description="See everything AFC Bot can do")
@app_commands.guild_only()
async def slash_help(interaction: discord.Interaction):
    await _dispatch_slash(interaction, "help")


@tree.command(name="ask", description="Ask AFC Bot anything — tournaments, teams, registration, rules")
@app_commands.describe(question="Your question (Pidgin welcome)")
@app_commands.guild_only()
async def slash_ask(interaction: discord.Interaction, question: str):
    await interaction.response.defer(thinking=True)
    needs_support = False
    try:
        reply, needs_support = await ask_openai_text(
            interaction.channel_id, question, interaction.user.display_name,
            is_staff=has_staff_role(interaction.user),
        )
    except Exception as e:
        reply = resolve_ai_error_reply(interaction.channel_id, e, force=True)
    if not reply or not reply.strip():
        reply = GENERIC_ERROR_NOTICE
    for i in range(0, len(reply), 2000):
        await interaction.followup.send(reply[i:i + 2000])
    if needs_support:
        await send_support_redirect(_SlashCommandShim(interaction, question))


@tree.command(name="announce", description="Post an AI-formatted announcement (shows a preview first)")
@app_commands.describe(
    channel="Channel to announce in",
    message="What the announcement should say",
    user="Optionally tag a user in the announcement",
    image="Optional image for the announcement",
)
@app_commands.default_permissions(manage_messages=True)
@app_commands.guild_only()
async def slash_announce(
    interaction: discord.Interaction,
    channel: discord.TextChannel,
    message: str,
    user: Optional[discord.Member] = None,
    image: Optional[discord.Attachment] = None,
):
    text = (
        f"send {message} to {user.mention} in {channel.mention}"
        if user else f"send {message} to {channel.mention}"
    )
    await _dispatch_slash(interaction, text, attachments=[image] if image else None)


@tree.command(name="edit-last", description="Rewrite my last message or announcement with your instruction")
@app_commands.describe(
    instruction="How to change it (e.g. 'less emojis', 'make it shorter')",
    channel="Channel of the announcement to edit (default: last reply here)",
)
@app_commands.default_permissions(manage_messages=True)
@app_commands.guild_only()
async def slash_edit_last(
    interaction: discord.Interaction,
    instruction: str,
    channel: Optional[discord.TextChannel] = None,
):
    text = (
        f"edit last announcement in {channel.mention} — {instruction}"
        if channel else f"edit my last message — {instruction}"
    )
    await _dispatch_slash(interaction, text)


@tree.command(name="purge", description="Delete the last N messages in a channel (asks to confirm)")
@app_commands.describe(amount="How many messages to delete", channel="Channel (default: here)")
@app_commands.default_permissions(manage_messages=True)
@app_commands.guild_only()
async def slash_purge(
    interaction: discord.Interaction,
    amount: app_commands.Range[int, 1, 1000],
    channel: Optional[discord.TextChannel] = None,
):
    where = f" in {channel.mention}" if channel else ""
    await _dispatch_slash(interaction, f"delete the last {amount} messages{where}")


@tree.command(name="purge-user", description="Delete all messages from a user (asks to confirm)")
@app_commands.describe(user="Whose messages to delete", channel="Channel (default: here)")
@app_commands.default_permissions(manage_messages=True)
@app_commands.guild_only()
async def slash_purge_user(
    interaction: discord.Interaction,
    user: discord.Member,
    channel: Optional[discord.TextChannel] = None,
):
    where = f" in {channel.mention}" if channel else ""
    await _dispatch_slash(interaction, f"purge messages from {user.mention}{where}")


@tree.command(name="purge-all", description="Delete ALL messages in a channel (asks to confirm)")
@app_commands.describe(channel="Channel (default: here)")
@app_commands.default_permissions(manage_messages=True)
@app_commands.guild_only()
async def slash_purge_all(
    interaction: discord.Interaction,
    channel: Optional[discord.TextChannel] = None,
):
    where = f" in {channel.mention}" if channel else ""
    await _dispatch_slash(interaction, f"purge all messages{where}")


@tree.command(name="transcribe", description="Join a voice/stage channel and transcribe the session")
@app_commands.describe(channel="Voice or stage channel (default: the one you're in)")
@app_commands.default_permissions(manage_messages=True)
@app_commands.guild_only()
async def slash_transcribe(
    interaction: discord.Interaction,
    channel: Optional[Union[discord.VoiceChannel, discord.StageChannel]] = None,
):
    text = f"transcribe {channel.mention}" if channel else "transcribe this session"
    await _dispatch_slash(interaction, text)


@tree.command(name="stop-transcribing", description="Stop recording and generate the transcript")
@app_commands.default_permissions(manage_messages=True)
@app_commands.guild_only()
async def slash_stop_transcribing(interaction: discord.Interaction):
    await _dispatch_slash(interaction, "stop transcribing")


@tree.command(name="role-give", description="Give a role to a user")
@app_commands.describe(user="Who gets the role", role="Role to give")
@app_commands.default_permissions(manage_roles=True)
@app_commands.guild_only()
async def slash_role_give(interaction: discord.Interaction, user: discord.Member, role: discord.Role):
    await _dispatch_slash(interaction, f"give {role.mention} to {user.mention}")


@tree.command(name="role-remove", description="Remove a role from a user")
@app_commands.describe(user="Who loses the role", role="Role to remove")
@app_commands.default_permissions(manage_roles=True)
@app_commands.guild_only()
async def slash_role_remove(interaction: discord.Interaction, user: discord.Member, role: discord.Role):
    await _dispatch_slash(interaction, f"remove {role.mention} from {user.mention}")


@tree.command(name="lock", description="Lock a channel (everyone can read, nobody can send)")
@app_commands.describe(channel="Channel to lock")
@app_commands.default_permissions(manage_channels=True)
@app_commands.guild_only()
async def slash_lock(interaction: discord.Interaction, channel: discord.TextChannel):
    await _dispatch_slash(interaction, f"lock {channel.mention}")


@tree.command(name="unlock", description="Unlock a previously locked channel")
@app_commands.describe(channel="Channel to unlock")
@app_commands.default_permissions(manage_channels=True)
@app_commands.guild_only()
async def slash_unlock(interaction: discord.Interaction, channel: discord.TextChannel):
    await _dispatch_slash(interaction, f"unlock {channel.mention}")


@tree.command(name="create-channel", description="Create a text/voice channel or category")
@app_commands.describe(
    name="Name for the new channel",
    kind="What to create",
    private="Hide it from @everyone",
    category="Put the channel under this category",
    role="Role that can access it (for private channels)",
)
@app_commands.choices(kind=[
    app_commands.Choice(name="text", value="text"),
    app_commands.Choice(name="voice", value="voice"),
    app_commands.Choice(name="category", value="category"),
])
@app_commands.default_permissions(manage_channels=True)
@app_commands.guild_only()
async def slash_create_channel(
    interaction: discord.Interaction,
    name: str,
    kind: app_commands.Choice[str],
    private: Optional[bool] = False,
    category: Optional[discord.CategoryChannel] = None,
    role: Optional[discord.Role] = None,
):
    parts = ["create"]
    if private:
        parts.append("private")
    if kind.value == "category":
        parts.append(f"category {name}")
    else:
        parts.append(f"{kind.value} channel {name}")
    if role:
        parts.append(f"for {role.mention}")
    if category and kind.value != "category":
        parts.append(f"in <#{category.id}>")
    await _dispatch_slash(interaction, " ".join(parts))


# ── Events ───────────────────────────────────────────────────────────────────
# on_ready fires on EVERY gateway re-IDENTIFY (reconnect), not just first boot.
# One-time init (background loops, persistent view, approval state) must run
# exactly once per process or each reconnect would stack duplicate poll loops
# and double-post every announcement.
_bg_loops_started = False


@bot.event
async def on_ready():
    global _bg_loops_started
    # Ensure knowledge folder exists
    try:
        os.makedirs(KNOWLEDGE_DIR, exist_ok=True)
    except Exception as e:
        print(f"⚠️  Could not create knowledge folder: {e}")

    load_history_from_disk()
    purge_expired_history()

    # Pre-fetch events so the bot has live data from the very first message
    global _cached_events
    try:
        events = await fetch_all_events()
        if events is not None:
            _cached_events = events
        print(f"🎮  Pre-cached {len(_cached_events)} event(s) for live Q&A")
    except Exception as e:
        print(f"⚠️  Could not pre-cache events: {e}")

    if _bg_loops_started:
        print("🔁  Reconnected — background loops already running.")
        return
    _bg_loops_started = True

    # Restore organizer-event approval state and bind the persistent buttons so
    # previews posted before this restart keep working.
    global _pending_event_approvals, _rejected_event_ids
    _pending_event_approvals = load_pending_event_approvals()
    _rejected_event_ids = load_rejected_event_ids()
    bot.add_view(EventApprovalView())
    print(f"🕓  Approval gate: {len(_pending_event_approvals)} pending, {len(_rejected_event_ids)} rejected")

    # Sync slash commands per guild (instant, vs up to 1h for global sync).
    try:
        synced = []
        for g in bot.guilds:
            tree.copy_global_to(guild=g)
            synced = await tree.sync(guild=g)
        print(f"⌨️  Slash commands: synced {len(synced)} command(s) to {len(bot.guilds)} guild(s)")
    except Exception as e:
        print(f"⚠️  Slash command sync failed: {e}")

    bot.loop.create_task(auto_purge_loop())
    bot.loop.create_task(auto_scrape_loop())
    bot.loop.create_task(news_poll_loop())
    bot.loop.create_task(event_poll_loop())
    bot.loop.create_task(ban_poll_loop())
    print(f"✅  AFC Bot is online as {bot.user} (id: {bot.user.id})")
    print(f"📌  Listening in {len(ALLOWED_CHANNELS)} explicit channels + all channels in {len(ALLOWED_CATEGORIES)} category(ies)")
    print(f"📚  Knowledge base loaded: {len(load_knowledge())} characters")
    print(f"🕒  Conversation history: saved to disk, auto-clears after 24 hours")
    print(f"🔄  Auto-scrape: every {SCRAPE_INTERVAL_HOURS}h (first run in {SCRAPE_INTERVAL_HOURS}h)")
    print(f"📰  News poll: every {NEWS_POLL_INTERVAL_SECS}s → channel {NEWS_ANNOUNCEMENT_CHANNEL_ID}")
    print(f"🎮  Event poll: every {EVENT_POLL_INTERVAL_SECS}s → tournament ch {TOURNAMENT_ANNOUNCEMENT_CHANNEL_ID} / scrim ch {SCRIM_ANNOUNCEMENT_CHANNEL_ID} (+ status change tracking)")
    print(f"🔨  Ban poll: every {BAN_POLL_INTERVAL_SECS}s → ban ch {BAN_ANNOUNCEMENT_CHANNEL_ID} / unban ch {UNBAN_ANNOUNCEMENT_CHANNEL_ID}")


@bot.event
async def on_message(message: discord.Message):
    # Ignore bots entirely
    if message.author.bot:
        return
    try:
        await _handle_message(message)
    except Exception as e:
        print(f"⚠️  Unhandled error in on_message: {e}")


# Tracks which user-message IDs the bot has already replied to, so an edit of an
# answered message doesn't produce a duplicate reply. Insertion-ordered dict (not
# a set) so eviction genuinely drops the OLDEST entries.
_handled_message_ids: dict[int, None] = {}
_HANDLED_MAX = 2000


def _mark_handled(message_id: int):
    _handled_message_ids[message_id] = None
    if len(_handled_message_ids) > _HANDLED_MAX:
        # Drop the oldest ~25% to keep the dict bounded
        for mid in list(_handled_message_ids)[: _HANDLED_MAX // 4]:
            _handled_message_ids.pop(mid, None)


@bot.event
async def on_message_edit(before: discord.Message, after: discord.Message):
    """When a user edits a message, re-evaluate it.

    Real users sometimes type a vague message ('okay'), then edit it into a
    real question ('which time are we going to start the tournament').
    Without this handler the bot would never see the new content.
    """
    if after.author.bot:
        return
    if before.content == after.content and not after.attachments:
        return
    # Skip if we already replied directly to this message id (avoid duplicate
    # replies on every edit). The classifier path will still re-evaluate
    # because that flow does not call _mark_handled.
    if after.id in _handled_message_ids:
        return
    try:
        await _handle_message(after)
    except Exception as e:
        print(f"⚠️  Unhandled error in on_message_edit: {e}")


async def should_bot_respond(message_text: str, image_bytes: bytes = None, image_media_type: str = None) -> bool:
    """
    Use GPT to quickly decide if a message is worth the bot responding to.
    If an image is provided, it is included so the classifier can read it
    before deciding — the image content + the question together determine intent.
    Defaults to True on any error so the bot never silently ignores questions.
    """
    system_prompt = (
        "You are a classifier for a Discord bot on an African Free Fire gaming support channel (AFC).\n"
        "Decide if the BOT should respond to this message.\n\n"
        "The bot should respond when the message is about AFC, tournaments, registration, teams, or the platform — "
        "even if the person is not directly addressing the bot, even if it is a Discord reply to another user, "
        "and even if there is no question mark.\n\n"
        "Reply YES if the message:\n"
        "- Asks about how to use the AFC platform (register, join, create team, etc.)\n"
        "- Describes a problem with their account, team, registration, or tournament (even as a statement, e.g. 'my team is still pending', 'I registered but nothing happened')\n"
        "- Is asking about AFC rules, features, events, or how something works\n"
        "- Asks about tournament times, schedules, start/end dates, or event details (even without a question mark, even one or two words like 'when start?', 'time?', 'tournament time')\n"
        "- Is clearly seeking help or information, even if phrased as a statement rather than a question\n"
        "- Shares a screenshot of an AFC page, email, or app (with or without a question)\n"
        "- Reports a bug, error, or unexpected behavior on the platform\n"
        "- Expresses confusion or frustration about a platform feature or process\n"
        "- Is a new member introducing themselves or asking to be added/included (e.g. 'I'm new here', 'can someone add me', 'how do I join')\n"
        "- Is a vague or unclear request that COULD be about the platform (e.g. 'can someone help me', 'add me pls', 'how to register') — the bot can ask for clarification\n"
        "- Contains typos or informal language but the INTENT is asking for help (e.g. 'had me' meaning 'add me', 'tornument' meaning 'tournament')\n"
        "- Is a Discord reply to another user but the message itself is a platform question, problem report, or status update about a team/tournament/registration\n\n"
        "Reply NO ONLY if the message is:\n"
        "- Pure casual chat, greetings, reactions, hype with NO AFC content (e.g. 'lol', 'gg', 'nice', 'gm', 'fire bro', 'okay', 'cool')\n"
        "- A question one PERSON is asking ANOTHER SPECIFIC PERSON by name or @mention about purely personal details unrelated to the platform (e.g. '@John what is your in-game name?')\n"
        "- Pure banter, jokes, or off-topic conversation with no AFC/platform relevance\n"
        "- A screenshot shared casually (meme, flex, celebration) with no question or help-seeking intent\n\n"
        "IMPORTANT DISTINCTIONS:\n"
        "- '@John what is your team name?' → directed at A SPECIFIC USER by name → NO\n"
        "- 'what is your team name?' with no specific target → could be asking the community → YES\n"
        "- 'how do I register my team?' → platform question → YES\n"
        "- 'Guys how to register to tournament' → asking the community for help → YES\n"
        "- 'Can someone add me' / 'Can someone had me' → new member seeking help (typos included) → YES\n"
        "- 'I'm new here' / 'Hi guys new here' → new member who likely needs guidance → YES\n"
        "- 'which time are we going to start the tournament' → tournament question → YES (no question mark needed)\n"
        "- 'Which time are we going to start the tournament' (as a reply to anyone) → still a tournament question → YES\n"
        "- 'I have registered my team but we are still on pending' → describes a platform problem → YES\n"
        "- 'when is the next event' → event/schedule question → YES\n"
        "- 'has the tournament started yet' / 'is it live' / 'tournament time?' → tournament status question → YES\n"
        "- Messages do NOT need a question mark to be questions. Look at intent, not punctuation.\n"
        "- Statements describing problems ('still pending', 'not working', 'can't join') are implicit help requests → YES\n"
        "- Messages with typos or broken English should still be understood by INTENT, not spelling → YES if help-seeking\n"
        "- Reply context (the [This is a reply to ...] line) is FYI only — it does NOT make the message ineligible. Judge the message itself.\n\n"
        "If an image is provided, read it first — its content should inform your decision alongside the text.\n"
        "When in doubt, reply YES — it is FAR better to respond unnecessarily than to ignore someone who needs help.\n"
        "Reply with only YES or NO."
    )
    try:
        if image_bytes and image_media_type:
            image_b64 = base64.b64encode(image_bytes).decode("utf-8")
            user_content = [
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:{image_media_type};base64,{image_b64}"}
                },
                {
                    "type": "text",
                    "text": message_text if message_text else "(no text — image only)"
                }
            ]
        else:
            user_content = message_text

        response = await _achat(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_content}
            ],
            max_tokens=5,
            temperature=0,
        )
        answer = response.choices[0].message.content.strip().upper()
        return answer.startswith("YES")
    except Exception as e:
        print(f"⚠️  should_bot_respond error: {e} — defaulting to respond")
        return True  # always respond if classifier fails


async def _handle_message(message: discord.Message):
    # Ignore self
    if message.author == bot.user:
        return

    # Only allowed channels — except slash commands, which Discord already
    # scopes via the command permissions and can run anywhere staff invoke them.
    if not is_allowed_channel(message.channel.id, message.channel) and not getattr(message, "_from_slash", False):
        return

    # If this channel is waiting for a confirmation reply, suppress messages from
    # the triggering user — but only if they're not issuing a new bot command.
    pending_user = _awaiting_confirmation.get(message.channel.id)
    if pending_user and message.author.id == pending_user:
        # Allow through if the user is @mentioning the bot (new command)
        if bot.user not in message.mentions:
            return

    is_mentioned = bot.user in message.mentions

    # Auto-reply runs in EVERY allowed channel — the classifier decides whether
    # the message actually needs a bot response. This way no question gets ignored
    # just because it lives in a channel that wasn't explicitly listed.
    if not is_mentioned:
        content = message.content.strip()

        # Quick filter — skip very short messages and pure emoji reactions
        if len(content) < 4 and not message.attachments:
            return
        if content and re.match(r'^[\U00010000-\U0010ffff\U00002000-\U00002BFF\s]+$', content):
            return

        # If this message is a Discord reply, fetch the parent so the classifier
        # can use it as context. We DO NOT silently skip replies anymore — many
        # real questions ("we are still on pending", "which time are we going to
        # start the tournament") arrive as replies to other users.
        reply_context_text = ""
        if message.reference is not None:
            ref = message.reference.resolved
            if ref is None and message.reference.message_id:
                try:
                    ref = await message.channel.fetch_message(message.reference.message_id)
                except Exception:
                    ref = None
            if isinstance(ref, discord.Message):
                ref_author = "the bot" if ref.author == bot.user else ref.author.display_name
                ref_snippet = (ref.content or "")[:200]
                reply_context_text = f"[This is a reply to {ref_author}: \"{ref_snippet}\"]\n"

        # If there's an image, download it so the classifier can read it
        # alongside the text before deciding whether to respond.
        classifier_image_bytes = None
        classifier_media_type = None
        image_attachment = next(
            (a for a in message.attachments if get_attachment_type(a.filename) == "image"),
            None
        )
        if image_attachment:
            ext = os.path.splitext(image_attachment.filename)[1].lower().strip(".")
            media_type_map = {
                "jpg": "image/jpeg", "jpeg": "image/jpeg",
                "png": "image/png", "gif": "image/gif", "webp": "image/webp"
            }
            classifier_media_type = media_type_map.get(ext, "image/jpeg")
            try:
                classifier_image_bytes = await download_attachment(image_attachment)
            except Exception:
                pass  # if download fails, classify on text alone

        # Ask GPT (with image if present) whether this message needs a response
        classifier_text = (reply_context_text + content) if reply_context_text else content
        should_respond = await should_bot_respond(classifier_text, classifier_image_bytes, classifier_media_type)
        if not should_respond:
            return

    # Strip the @mention
    user_text = message.content.replace(f"<@{bot.user.id}>", "").strip()
    username  = message.author.display_name
    is_staff  = has_staff_role(message.author)

    # ── Edit command check FIRST — must run before announcement to avoid confusion ──
    if has_admin_role(message.author):
        edit_cmd = parse_edit_command(user_text)
        if edit_cmd:
            mode        = edit_cmd["mode"]
            instruction = edit_cmd["instruction"]
            edit_ch_id  = edit_cmd["channel_id"] or message.channel.id

            # Use fetch_channel to guarantee we get the channel even if not cached
            try:
                edit_ch = await bot.fetch_channel(edit_ch_id)
            except Exception:
                edit_ch = message.channel

            async def apply_edit(target_msg: discord.Message):
                if target_msg.author.id != bot.user.id:
                    await message.reply("❌ I can only edit my own messages.", mention_author=True)
                    return
                if target_msg.embeds and target_msg.embeds[0].description:
                    original = target_msg.embeds[0].description
                    is_embed = True
                    old_embed = target_msg.embeds[0]
                else:
                    original = target_msg.content or ""
                    is_embed = False
                if not original:
                    await message.reply("❌ That message has no text content I can edit.", mention_author=True)
                    return
                async with message.channel.typing():
                    new_text = await ai_rewrite(original, instruction)
                if is_embed:
                    new_embed = discord.Embed(
                        title=old_embed.title,
                        description=new_text,
                        color=old_embed.color
                    )
                    new_embed.set_footer(text="African Freefire Community  •  africanfreefirecommunity.com  •  (edited)")
                    new_embed.timestamp = __import__("datetime").datetime.now(__import__("datetime").timezone.utc)
                    await target_msg.edit(embed=new_embed)
                else:
                    await target_msg.edit(content=new_text)
                await message.reply("✅ Done! Message updated.", mention_author=True)

            try:
                if mode == "id":
                    msg_id = edit_cmd["message_id"]
                    try:
                        target_msg = await edit_ch.fetch_message(msg_id)
                    except discord.NotFound:
                        await message.reply("❌ Couldn't find that message. Check the ID is correct.", mention_author=True)
                        return
                    await apply_edit(target_msg)
                elif mode == "announce":
                    msg_id = last_bot_announcements.get(edit_ch_id)
                    if not msg_id:
                        await message.reply(
                            f"❌ No announcement record for {edit_ch.mention}. Use the message ID instead.",
                            mention_author=True
                        )
                        return
                    target_msg = await edit_ch.fetch_message(msg_id)
                    await apply_edit(target_msg)
                elif mode == "last":
                    # If a target channel was specified, look for last message there
                    # Otherwise fall back to current channel
                    lookup_channel_id = edit_ch_id if edit_ch_id != message.channel.id else message.channel.id
                    msg_id = (
                        last_bot_announcements.get(lookup_channel_id)
                        or last_bot_messages.get(lookup_channel_id)
                        or last_bot_messages.get(message.channel.id)
                    )
                    if not msg_id:
                        await message.reply(
                            f"❌ No record of my last message in {edit_ch.mention}. "
                            "Use the message ID directly — right-click the message → Copy Message ID.",
                            mention_author=True
                        )
                        return
                    target_msg = await edit_ch.fetch_message(msg_id)
                    await apply_edit(target_msg)
            except discord.Forbidden:
                await message.reply("❌ I don't have permission to edit that message.", mention_author=True)
            except Exception as e:
                # ai_rewrite() runs through _chat_completion, so an AI-unavailability
                # error can surface here — never dump the raw OpenAI billing error.
                if _should_failover(e):
                    print(f"⚠️  AI unavailable during edit: {e}")
                    await message.reply(AI_DOWN_NOTICE, mention_author=True)
                else:
                    # Never dump the raw exception (it can carry the provider's
                    # full API error body) into a Discord channel.
                    print(f"⚠️  Edit failed: {e}")
                    await message.reply(GENERIC_ERROR_NOTICE, mention_author=True)
            return

    # ── Announcement command ─────────────────────────────────────────────────
    announce = parse_announce_command(user_text)
    if announce:
        if not has_announce_role(message.author):
            await message.reply(
                "❌ You don't have permission to use the announcement command. Only AFC staff roles can do that.",
                mention_author=True
            )
            return

        target_channel_id, target_user_id, msg_content = announce
        target_channel = bot.get_channel(target_channel_id)
        if not target_channel:
            try:
                target_channel = await bot.fetch_channel(target_channel_id)
            except Exception:
                await message.reply("❌ I couldn't find that channel. Make sure I have access to it.", mention_author=True)
                return

        generate_keywords = [
            "formulate", "generate", "write", "create", "draft", "help me",
            "make", "compose", "craft", "prepare", "put together"
        ]
        should_generate = any(kw in user_text.lower() for kw in generate_keywords)

        # Collect attached files
        import io
        image_files = []
        other_files = []
        if message.attachments:
            for attachment in message.attachments:
                file_bytes = await download_attachment(attachment)
                att_type = get_attachment_type(attachment.filename)
                if att_type == "image":
                    image_files.append((attachment.filename, file_bytes))
                else:
                    other_files.append((attachment.filename, file_bytes))

        layout_keywords = ["above", "below", "middle", "between", "then", "followed by",
                           "first image", "second image", "image then text", "text then image",
                           "top", "bottom", "after the text", "before the text"]
        is_multi_layout = len(image_files) > 1 or any(kw in user_text.lower() for kw in layout_keywords)

        # ── Generate the announcement ────────────────────────────────────────
        current_hints = msg_content  # kept for re-generation on feedback

        async def generate_embed_and_content(hints, uid):
            if should_generate or hints:
                try:
                    ann_data = await generate_announcement(hints, uid)
                    emb, ping = build_embed(ann_data)
                    return emb, ping
                except Exception as e:
                    return None, None
            else:
                plain = f"<@{uid}> {hints}" if uid else hints
                emb = discord.Embed(description=plain, color=EMBED_COLORS["announcement"])
                emb.set_footer(text="African Freefire Community  •  africanfreefirecommunity.com")
                emb.timestamp = __import__("datetime").datetime.now(__import__("datetime").timezone.utc)
                return emb, None

        async with message.channel.typing():
            embed, ping_content = await generate_embed_and_content(current_hints, target_user_id)

        if embed is None:
            await message.reply("⚠️ Couldn't generate the announcement. Please try again.", mention_author=True)
            return

        # ── PREVIEW LOOP — show preview, wait for approval ───────────────────
        max_attempts = 5
        attempt = 0

        while attempt < max_attempts:
            attempt += 1

            # Build preview embed with header
            preview_embed = discord.Embed(
                title=embed.title,
                description=embed.description,
                color=embed.color
            )
            preview_embed.set_footer(text=f"📋 PREVIEW — will be sent to #{target_channel.name}  •  africanfreefirecommunity.com")
            if embed.timestamp:
                preview_embed.timestamp = embed.timestamp

            # Show preview with images if any
            preview_files = []
            if image_files:
                preview_files = [discord.File(fp=io.BytesIO(fb), filename=fn) for fn, fb in image_files[:1]]

            await message.channel.send(
                content=f"{message.author.mention} **Preview** → will send to <#{target_channel_id}>",
                embed=preview_embed,
                files=preview_files if preview_files else discord.utils.MISSING
            )

            # Show approval prompt
            approval_msg = await message.channel.send(
                "✅ Type `send` to send  |  ✏️ Type your correction to fix it  |  ❌ Type `cancel` to cancel"
            )

            def approval_check(m):
                return (
                    m.author.id == message.author.id
                    and m.channel.id == message.channel.id
                )

            try:
                _awaiting_confirmation[message.channel.id] = message.author.id
                response = await bot.wait_for("message", check=approval_check, timeout=120.0)
                resp_text = response.content.strip().lower()

                # Clean up the approval prompt and response after reading
                try:
                    await approval_msg.delete()
                    await response.delete()
                except Exception:
                    pass

                if resp_text == "cancel":
                    await message.channel.send("❌ Announcement cancelled.", delete_after=5)
                    return

                elif resp_text == "send":
                    # ── Send to target channel ───────────────────────────────
                    break  # exit loop and send below

                else:
                    # User gave feedback — regenerate with correction
                    async with message.channel.typing():
                        # Combine original hints with the correction
                        new_hints = f"{current_hints}\n\nCorrection: {response.content.strip()}"
                        embed, ping_content = await generate_embed_and_content(new_hints, target_user_id)
                    if embed is None:
                        await message.channel.send("⚠️ Couldn't regenerate. Try again.", delete_after=5)
                        return
                    continue  # show updated preview

            except asyncio.TimeoutError:
                await message.channel.send(
                    "⏱️ No response for 2 minutes. Announcement cancelled.",
                    delete_after=10
                )
                return
            finally:
                _awaiting_confirmation.pop(message.channel.id, None)
        else:
            # Loop exhausted without an explicit `send` (the 5th message was yet
            # another correction) — NEVER auto-post an announcement the admin
            # hasn't previewed and approved.
            await message.channel.send(
                "❌ Too many revisions without `send` — announcement cancelled.",
                delete_after=10
            )
            return

        # ── Actually send to target channel ──────────────────────────────────
        try:
            allowed = discord.AllowedMentions(everyone=True, roles=True, users=True)
            sent = None

            if is_multi_layout and image_files:
                text_lower = user_text.lower()
                if re.search(r"\bimage\s+(above|first|on\s+top|before\s+text)\b"
                             r"|\babove\s+the\s+text\b|\bimage\s+then\s+text\b"
                             r"|\btop\b.*\bimage\b", text_lower):
                    order = "images_first"
                elif re.search(r"\bimage\s+(below|after|at\s+the\s+bottom|after\s+text)\b"
                               r"|\bbelow\s+the\s+text\b|\btext\s+then\s+image\b"
                               r"|\bbottom\b.*\bimage\b", text_lower):
                    order = "text_first"
                elif re.search(r"\bmiddle\b|\bbetween\b|\bsandwich\b", text_lower):
                    order = "sandwich"
                else:
                    order = "images_first"

                if order == "images_first":
                    for fname, fbytes in image_files:
                        await target_channel.send(
                            file=discord.File(fp=io.BytesIO(fbytes), filename=fname),
                            allowed_mentions=allowed
                        )
                        await asyncio.sleep(0.5)
                    sent = await target_channel.send(content=ping_content, embed=embed, allowed_mentions=allowed)
                    for fname, fbytes in other_files:
                        await target_channel.send(file=discord.File(fp=io.BytesIO(fbytes), filename=fname))

                elif order == "text_first":
                    sent = await target_channel.send(content=ping_content, embed=embed, allowed_mentions=allowed)
                    for fname, fbytes in image_files:
                        await target_channel.send(
                            file=discord.File(fp=io.BytesIO(fbytes), filename=fname),
                            allowed_mentions=allowed
                        )
                        await asyncio.sleep(0.5)

                elif order == "sandwich":
                    first_fname, first_fbytes = image_files[0]
                    await target_channel.send(
                        file=discord.File(fp=io.BytesIO(first_fbytes), filename=first_fname),
                        allowed_mentions=allowed
                    )
                    await asyncio.sleep(0.5)
                    sent = await target_channel.send(content=ping_content, embed=embed, allowed_mentions=allowed)
                    for fname, fbytes in image_files[1:]:
                        await asyncio.sleep(0.5)
                        await target_channel.send(
                            file=discord.File(fp=io.BytesIO(fbytes), filename=fname),
                            allowed_mentions=allowed
                        )

            elif image_files:
                files_to_send = [discord.File(fp=io.BytesIO(fb), filename=fn) for fn, fb in image_files]
                for fname, fbytes in other_files:
                    files_to_send.append(discord.File(fp=io.BytesIO(fbytes), filename=fname))
                sent = await target_channel.send(
                    content=ping_content, embed=embed,
                    files=files_to_send, allowed_mentions=allowed
                )
            else:
                other_fs = [discord.File(fp=io.BytesIO(fb), filename=fn) for fn, fb in other_files]
                sent = await target_channel.send(
                    content=ping_content, embed=embed,
                    files=other_fs if other_fs else discord.utils.MISSING,
                    allowed_mentions=allowed
                )

            if sent:
                last_bot_announcements[target_channel.id] = sent.id

            await message.channel.send(
                f"✅ Announcement sent to <#{target_channel_id}>!",
                delete_after=10
            )

        except discord.Forbidden:
            await message.reply(
                f"❌ I don't have permission to send messages in <#{target_channel_id}>.",
                mention_author=True
            )
        return
    # ── End announcement ─────────────────────────────────────────────────────

    # ── Transcription commands (mod/support only) ────────────────────────────
    if has_transcription_role(message.author):
        trans_cmd = parse_transcription_command(user_text)
        if trans_cmd:
            action = trans_cmd["action"]

            if action == "stop":
                if message.guild.id not in active_transcriptions:
                    await message.reply("❌ No active transcription session running.", mention_author=True)
                    return
                await message.reply("⏹️ Stopped recording. Generating transcript...", mention_author=True)
                await _finalize_transcription(message.guild.id, None)
                return

            elif action == "start":
                if not VOICE_RECEIVE_SUPPORTED:
                    await message.reply(
                        "❌ Voice transcription isn't supported in this bot build "
                        "(the library has no voice-receive support).",
                        mention_author=True
                    )
                    return
                if message.guild.id in active_transcriptions:
                    await message.reply(
                        "❌ Already transcribing a session. Say `@bot stop transcribing` first.",
                        mention_author=True
                    )
                    return

                ch_id    = trans_cmd.get("channel_id")
                voice_ch = None

                if ch_id:
                    voice_ch = bot.get_channel(ch_id)
                    if not isinstance(voice_ch, (discord.VoiceChannel, discord.StageChannel)):
                        await message.reply("❌ That's not a valid voice or stage channel.", mention_author=True)
                        return
                elif message.author.voice:
                    voice_ch = message.author.voice.channel
                else:
                    await message.reply(
                        "❌ Please mention the channel: `@bot transcribe <#channel>` "
                        "or join the voice channel first.",
                        mention_author=True
                    )
                    return

                vc = None
                try:
                    vc         = await voice_ch.connect(cls=voice_recv.VoiceRecvClient)
                    sink       = PerUserTranscriptionSink()
                    start_time = datetime.now(timezone.utc)
                    requester  = message.author
                    text_ch    = message.channel

                    active_transcriptions[message.guild.id] = {
                        "vc":           vc,
                        "sink":         sink,
                        "text_channel": text_ch,
                        "requester":    requester,
                        "start_time":   start_time,
                    }

                    vc.listen(sink)
                    _start_silence_watchdog(message.guild.id, sink)
                    await message.reply(
                        f"✅ Joined {voice_ch.mention} and started transcribing.\n"
                        f"Say `@bot stop transcribing` when the session is done "
                        f"(I'll also stop on my own after {TRANSCRIPTION_SILENCE_TIMEOUT_SECS // 60} minutes of silence).",
                        mention_author=True
                    )
                except discord.ClientException:
                    await message.reply("❌ I'm already connected to a voice channel in this server.", mention_author=True)
                except Exception as e:
                    # Never strand a live voice connection when a step after a
                    # successful connect() fails.
                    if vc is not None:
                        try:
                            await vc.disconnect(force=True)
                        except Exception:
                            pass
                        active_transcriptions.pop(message.guild.id, None)
                    await message.reply(f"⚠️ Couldn't start transcribing: {e}", mention_author=True)
                return

    # ── Help command ─────────────────────────────────────────────────────────
    if re.search(r"^\s*(help|commands?|what\s+can\s+you\s+do|list\s+commands?|show\s+commands?)\s*$", user_text, re.IGNORECASE):
        is_admin = has_admin_role(message.author)
        is_trans = has_transcription_role(message.author)

        # ── Page 1: General / Q&A ─────────────────────────────────────────
        general = discord.Embed(
            title="🤖 AFC Bot — What I Can Do",
            description=(
                "I'm the official AFC assistant. Here's everything I can help with.\n"
                "All commands are natural language — just @mention me and describe what you need."
            ),
            color=EMBED_COLORS["general"],
        )
        general.add_field(
            name="💬 Questions & Support",
            value=(
                "Ask me anything about the AFC platform — tournaments, teams, registration, rules, rankings, and more.\n"
                "I understand Nigerian Pidgin too. Just ask naturally.\n"
                "*(Works in support & general channels without @mention)*"
            ),
            inline=False,
        )
        general.set_footer(text="African Freefire Community  •  africanfreefirecommunity.com")
        general.timestamp = datetime.now(timezone.utc)

        await message.reply(embed=general, mention_author=False)

        # ── Page 2: Staff commands ────────────────────────────────────────
        if is_admin or is_trans:
            staff = discord.Embed(
                title="🛠️ Staff Commands",
                description="These commands are available to mods, admins, and support roles.",
                color=EMBED_COLORS["announcement"],
            )
            staff.add_field(
                name="📢 Announcements",
                value=(
                    "`@bot send [message] to #channel` — post an AI-formatted announcement\n"
                    "`@bot send [message] to @user in #channel` — tag a user in the announcement\n"
                    "You can also attach an image and it'll be included in the embed."
                ),
                inline=False,
            )
            staff.add_field(
                name="✏️ Edit Messages",
                value=(
                    "`@bot edit last message — [instruction]` — rewrite your last bot message\n"
                    "`@bot edit last announcement in #channel — [instruction]`\n"
                    "`@bot edit message [ID] — [instruction]`\n"
                    "Or just give feedback naturally: *\"too many emojis\"*, *\"make it shorter\"*"
                ),
                inline=False,
            )
            staff.add_field(
                name="🗑️ Purge Messages",
                value=(
                    "`@bot delete last [N] messages in #channel`\n"
                    "`@bot purge messages from @user in #channel`\n"
                    "`@bot delete messages from @user1 and @user2 in #channel`\n"
                    "`@bot delete messages from users with @role in #channel`\n"
                    "`@bot purge messages containing \"word\" in #channel`\n"
                    "`@bot purge all messages in #channel`"
                ),
                inline=False,
            )
            if is_trans and VOICE_RECEIVE_SUPPORTED:
                staff.add_field(
                    name="🎙️ Voice Transcription",
                    value=(
                        "`@bot transcribe #voice-channel` — join and start recording\n"
                        "`@bot transcribe` — join the voice channel you're in\n"
                        "`@bot stop transcribing` — stop and generate the transcript\n"
                        "When a stage is created, I'll ask in the mods channel if I should transcribe."
                    ),
                    inline=False,
                )
            if is_admin:
                staff.add_field(
                    name="⚙️ Server Management",
                    value=(
                        "**Channels:** `create channel/voice/category [name]` · `delete #channel` · `lock/unlock #channel` · `hide/show #channel`\n"
                        "**Roles (single user):** `give @role to @user` · `remove @role from @user`\n"
                        "**Roles (mass):** `give @role to everyone` · `remove @role from everyone with @otherrole`\n"
                        "**Role management:** `create role [name]` · `delete role @role` · `rename @role to [name]` · `recolor @role to #hex`"
                    ),
                    inline=False,
                )
            staff.set_footer(text="African Freefire Community  •  Staff commands")
            staff.timestamp = datetime.now(timezone.utc)
            await message.channel.send(embed=staff)

        return

    # ── Server management commands (admin only) ──────────────────────────────
    if has_admin_role(message.author):

        # ── Create channel / category ─────────────────────────────────────────
        create_cmd = parse_create_command(user_text)
        if create_cmd:
            guild     = message.guild
            ch_type   = create_cmd["type"]
            ch_name   = create_cmd["name"]
            is_private = create_cmd["private"]
            role_id   = create_cmd["role_id"]
            cat_id    = create_cmd["category_id"]

            # Resolve category if given
            category = None
            if cat_id:
                category = guild.get_channel(cat_id)
                if category and not isinstance(category, discord.CategoryChannel):
                    category = None  # not actually a category

            try:
                overwrites = {}

                if is_private:
                    # Hide from everyone by default
                    overwrites[guild.default_role] = discord.PermissionOverwrite(view_channel=False)
                    # Give access to the specified role if provided
                    if role_id:
                        role = guild.get_role(role_id)
                        if role:
                            overwrites[role] = discord.PermissionOverwrite(view_channel=True, send_messages=True)
                    # Always give access to admin roles
                    for admin_role_id in ANNOUNCE_ROLES:
                        admin_role = guild.get_role(admin_role_id)
                        if admin_role:
                            overwrites[admin_role] = discord.PermissionOverwrite(view_channel=True, send_messages=True)

                if ch_type == "category":
                    new_ch = await guild.create_category(
                        name=ch_name,
                        overwrites=overwrites,
                        reason=f"AFC Bot — created by {message.author}"
                    )
                    ch_label = f"📁 Category **{new_ch.name}**"

                elif ch_type == "voice":
                    new_ch = await guild.create_voice_channel(
                        name=ch_name,
                        category=category,
                        overwrites=overwrites,
                        reason=f"AFC Bot — created by {message.author}"
                    )
                    ch_label = f"🔊 Voice channel **{new_ch.name}**"

                else:  # text
                    new_ch = await guild.create_text_channel(
                        name=ch_name,
                        category=category,
                        overwrites=overwrites,
                        reason=f"AFC Bot — created by {message.author}"
                    )
                    ch_label = f"💬 Channel {new_ch.mention}"

                # Build response embed
                desc = f"✅ {ch_label} has been created."
                if category and ch_type != "category":
                    desc += f"\n📁 Category: **{category.name}**"
                if is_private:
                    desc += f"\n🔒 **Private** — hidden from @everyone"
                    if role_id:
                        desc += f" | Accessible to <@&{role_id}>"

                embed = discord.Embed(description=desc, color=EMBED_COLORS["general"])
                embed.set_footer(text=f"Created by {message.author.display_name}")
                await message.reply(embed=embed, mention_author=True)

            except discord.Forbidden:
                await message.reply("❌ I don't have permission to create channels. Make sure I have the Administrator role.", mention_author=True)
            return

        # ── Mass role operations ──────────────────────────────────────────────
        mass_role_cmd = parse_mass_role_command(user_text)
        if mass_role_cmd:
            action           = mass_role_cmd["action"]
            target_role_id   = mass_role_cmd["target_role_id"]
            condition_role_id = mass_role_cmd["condition_role_id"]
            guild            = message.guild
            target_role      = guild.get_role(target_role_id)

            if not target_role:
                await message.reply("❌ Couldn't find that role.", mention_author=True)
                return

            condition_role = None
            if condition_role_id:
                condition_role = guild.get_role(condition_role_id)
                if not condition_role:
                    await message.reply("❌ Couldn't find the condition role.", mention_author=True)
                    return

            # Build confirmation message
            if action == "remove_all":
                summary = f"Remove **@{target_role.name}** from **everyone** who has it"
            elif action == "remove_if":
                summary = f"Remove **@{target_role.name}** from everyone who has **@{condition_role.name}**"
            elif action == "give_all":
                summary = f"Give **@{target_role.name}** to **every member** in the server"
            elif action == "give_if":
                summary = f"Give **@{target_role.name}** to everyone who has **@{condition_role.name}**"
            else:
                summary = "Unknown action"

            embed = discord.Embed(
                title="⚠️ Confirm Mass Role Action",
                description=f"{summary}\n\n**This will affect multiple members.**\n\nReply `yes` to confirm or `no` to cancel.",
                color=EMBED_COLORS["urgent"]
            )
            embed.set_footer(text=f"Requested by {message.author.display_name} • Expires in 30 seconds")
            await message.reply(embed=embed, mention_author=True)

            def mass_check(m):
                return (
                    m.author.id == message.author.id
                    and m.channel.id == message.channel.id
                    and m.content.lower().strip() in ("yes", "no", "y", "n")
                )

            try:
                _awaiting_confirmation[message.channel.id] = message.author.id
                confirm = await bot.wait_for("message", check=mass_check, timeout=30.0)
                if confirm.content.lower().strip() not in ("yes", "y"):
                    await message.channel.send("❌ Mass role action cancelled.")
                    return
            except asyncio.TimeoutError:
                await message.channel.send("⏱️ Confirmation timed out. Action cancelled.")
                return
            finally:
                _awaiting_confirmation.pop(message.channel.id, None)

            # Execute — fetch all members and apply
            await message.channel.send(f"⚙️ Working on it... this may take a moment for large servers.")
            affected = 0
            failed   = 0

            try:
                await guild.chunk()  # ensure member cache is full
            except Exception:
                pass

            for member in guild.members:
                if member.bot:
                    continue

                # Determine if this member should be affected
                if action in ("remove_all", "give_all"):
                    should_act = True
                elif action == "remove_if":
                    should_act = condition_role in member.roles
                elif action == "give_if":
                    should_act = condition_role in member.roles
                else:
                    should_act = False

                if not should_act:
                    continue

                # Skip if already has/doesn't have the role
                if action in ("remove_all", "remove_if") and target_role not in member.roles:
                    continue
                if action in ("give_all", "give_if") and target_role in member.roles:
                    continue

                try:
                    success = False
                    retries = 0
                    while not success and retries < 5:
                        try:
                            if action in ("remove_all", "remove_if"):
                                await member.remove_roles(target_role, reason=f"AFC Bot mass action — by {message.author}")
                            else:
                                await member.add_roles(target_role, reason=f"AFC Bot mass action — by {message.author}")
                            affected += 1
                            success = True
                            await asyncio.sleep(0.8)  # safe delay between each member
                        except discord.HTTPException as e:
                            if e.status == 429:  # rate limited
                                retry_after = float(e.response.headers.get("Retry-After", 2))
                                await asyncio.sleep(retry_after + 0.5)
                                retries += 1
                            else:
                                failed += 1
                                success = True  # skip this member
                except Exception:
                    failed += 1

            verb = "removed from" if action in ("remove_all", "remove_if") else "given to"
            result_embed = discord.Embed(
                description=(
                    f"✅ **@{target_role.name}** {verb} **{affected}** member(s)."
                    + (f"\n⚠️ Failed for {failed} member(s)." if failed else "")
                ),
                color=EMBED_COLORS["general"]
            )
            result_embed.set_footer(text=f"Action by {message.author.display_name}")
            await message.channel.send(embed=result_embed)
            return

        # ── Give / Remove role (single user) ─────────────────────────────────
        role_cmd = parse_role_command(user_text)
        if role_cmd:
            action, target_user_id, target_role_id = role_cmd
            guild = message.guild
            member = guild.get_member(target_user_id)
            role   = guild.get_role(target_role_id)

            if not member:
                await message.reply("❌ Couldn't find that user in this server.", mention_author=True)
                return
            if not role:
                await message.reply("❌ Couldn't find that role.", mention_author=True)
                return

            try:
                if action == "give":
                    await member.add_roles(role, reason=f"AFC Bot — assigned by {message.author}")
                    embed = discord.Embed(
                        description=f"✅ **{role.name}** has been given to {member.mention}.",
                        color=EMBED_COLORS["general"]
                    )
                else:
                    await member.remove_roles(role, reason=f"AFC Bot — removed by {message.author}")
                    embed = discord.Embed(
                        description=f"✅ **{role.name}** has been removed from {member.mention}.",
                        color=EMBED_COLORS["urgent"]
                    )
                embed.set_footer(text=f"Action by {message.author.display_name}")
                await message.reply(embed=embed, mention_author=True)
            except discord.Forbidden:
                await message.reply("❌ I don't have permission to manage that role. Make sure my role is above it in the role list.", mention_author=True)
            return

        # ── Role management (create, delete, rename, recolor, edit props) ────
        role_manage_cmd = parse_role_manage_command(user_text)
        if role_manage_cmd:
            action   = role_manage_cmd["action"]
            role_id  = role_manage_cmd["role_id"]
            guild    = message.guild

            try:
                if action == "create":
                    name  = role_manage_cmd["name"] or "New Role"
                    color = discord.Color(role_manage_cmd["color"]) if role_manage_cmd["color"] else discord.Color.default()
                    mentionable = role_manage_cmd["mentionable"] or False
                    hoisted     = role_manage_cmd["hoisted"] or False
                    new_role = await guild.create_role(
                        name=name, color=color,
                        mentionable=mentionable, hoist=hoisted,
                        reason=f"AFC Bot — created by {message.author}"
                    )
                    embed = discord.Embed(
                        description=f"✅ Role **{new_role.name}** created — {new_role.mention}",
                        color=new_role.color
                    )
                    embed.set_footer(text=f"Created by {message.author.display_name}")
                    await message.reply(embed=embed, mention_author=True)

                elif action == "delete":
                    role = guild.get_role(role_id)
                    if not role:
                        await message.reply("❌ Couldn't find that role.", mention_author=True)
                        return
                    role_name = role.name
                    await role.delete(reason=f"AFC Bot — deleted by {message.author}")
                    embed = discord.Embed(
                        description=f"🗑️ Role **{role_name}** has been deleted.",
                        color=EMBED_COLORS["urgent"]
                    )
                    embed.set_footer(text=f"Action by {message.author.display_name}")
                    await message.reply(embed=embed, mention_author=True)

                elif action == "rename":
                    role = guild.get_role(role_id)
                    if not role:
                        await message.reply("❌ Couldn't find that role.", mention_author=True)
                        return
                    old_name = role.name
                    await role.edit(name=role_manage_cmd["name"], reason=f"AFC Bot — renamed by {message.author}")
                    embed = discord.Embed(
                        description=f"✅ Role **{old_name}** renamed to **{role.name}**.",
                        color=role.color
                    )
                    embed.set_footer(text=f"Action by {message.author.display_name}")
                    await message.reply(embed=embed, mention_author=True)

                elif action == "recolor":
                    role = guild.get_role(role_id)
                    if not role:
                        await message.reply("❌ Couldn't find that role.", mention_author=True)
                        return
                    new_color = discord.Color(role_manage_cmd["color"]) if role_manage_cmd["color"] else discord.Color.default()
                    await role.edit(color=new_color, reason=f"AFC Bot — recolored by {message.author}")
                    embed = discord.Embed(
                        description=f"✅ Role **{role.name}** color updated.",
                        color=new_color
                    )
                    embed.set_footer(text=f"Action by {message.author.display_name}")
                    await message.reply(embed=embed, mention_author=True)

                elif action == "edit_props":
                    role = guild.get_role(role_id)
                    if not role:
                        await message.reply("❌ Couldn't find that role.", mention_author=True)
                        return
                    kwargs = {}
                    if role_manage_cmd["mentionable"] is not None:
                        kwargs["mentionable"] = role_manage_cmd["mentionable"]
                    if role_manage_cmd["hoisted"] is not None:
                        kwargs["hoist"] = role_manage_cmd["hoisted"]
                    await role.edit(**kwargs, reason=f"AFC Bot — edited by {message.author}")
                    changes = []
                    if "mentionable" in kwargs:
                        changes.append(f"mentionable: **{'yes' if kwargs['mentionable'] else 'no'}**")
                    if "hoist" in kwargs:
                        changes.append(f"hoisted: **{'yes' if kwargs['hoist'] else 'no'}**")
                    embed = discord.Embed(
                        description=f"✅ Role **{role.name}** updated — {', '.join(changes)}.",
                        color=role.color
                    )
                    embed.set_footer(text=f"Action by {message.author.display_name}")
                    await message.reply(embed=embed, mention_author=True)

            except discord.Forbidden:
                await message.reply("❌ I don't have permission to manage that role. Make sure my role is above it in the role list.", mention_author=True)
            except Exception as e:
                await message.reply(f"⚠️ Something went wrong with the role action. Error: {e}", mention_author=True)
            return
        perm_cmd = parse_permission_command(user_text)
        if perm_cmd:
            ch_id, is_category, target_type, target_role_id, perm_name, allow = perm_cmd
            guild   = message.guild
            channel = guild.get_channel(ch_id)

            if not channel:
                await message.reply("❌ Couldn't find that channel.", mention_author=True)
                return

            # Determine the target (a role or @everyone)
            if target_type == "role" and target_role_id:
                target = guild.get_role(target_role_id)
                target_label = f"@{target.name}" if target else "that role"
            else:
                target = guild.default_role   # @everyone
                target_label = "@everyone"

            if not target:
                await message.reply("❌ Couldn't find that role.", mention_author=True)
                return

            try:
                perm_overwrite = channel.overwrites_for(target)
                setattr(perm_overwrite, perm_name, allow if allow else None)
                await channel.set_permissions(target, overwrite=perm_overwrite, reason=f"AFC Bot — changed by {message.author}")

                action_word = "allowed" if allow else "denied/removed"
                perm_display = perm_name.replace("_", " ").title()
                embed = discord.Embed(
                    description=f"✅ **{perm_display}** {action_word} for **{target_label}** in {channel.mention}.",
                    color=EMBED_COLORS["general"]
                )
                embed.set_footer(text=f"Action by {message.author.display_name}")
                await message.reply(embed=embed, mention_author=True)
            except discord.Forbidden:
                await message.reply("❌ I don't have permission to edit that channel's permissions.", mention_author=True)
            return

        # ── Delete channel (with confirmation) ───────────────────────────────
        del_cmd = parse_delete_command(user_text)
        if del_cmd:
            target_ch = message.guild.get_channel(del_cmd)
            if not target_ch:
                await message.reply("❌ Couldn't find that channel.", mention_author=True)
                return

            # Store pending and ask for confirmation
            pending_deletions[message.id] = del_cmd
            embed = discord.Embed(
                title="⚠️ Confirm Channel Deletion",
                description=(
                    f"You are about to permanently delete **{target_ch.name}**.\n\n"
                    f"**This cannot be undone.**\n\n"
                    f"Reply with ✅ `yes` or ❌ `no` to confirm."
                ),
                color=EMBED_COLORS["urgent"]
            )
            embed.set_footer(text=f"Requested by {message.author.display_name} • Expires in 30 seconds")
            confirm_msg = await message.reply(embed=embed, mention_author=True)

            # Wait for confirmation
            def check(m):
                return (
                    m.author.id == message.author.id
                    and m.channel.id == message.channel.id
                    and m.content.lower().strip() in ("yes", "no", "y", "n")
                )

            try:
                reply_msg = await bot.wait_for("message", check=check, timeout=30.0)
                if reply_msg.content.lower().strip() in ("yes", "y"):
                    ch_name = target_ch.name
                    await target_ch.delete(reason=f"AFC Bot — deleted by {message.author}")
                    embed2 = discord.Embed(
                        description=f"🗑️ Channel **#{ch_name}** has been deleted.",
                        color=EMBED_COLORS["urgent"]
                    )
                    embed2.set_footer(text=f"Action by {message.author.display_name}")
                    await message.channel.send(embed=embed2)
                else:
                    await message.channel.send("❌ Channel deletion cancelled.")
            except asyncio.TimeoutError:
                await message.channel.send("⏱️ Confirmation timed out. Channel deletion cancelled.")
            finally:
                pending_deletions.pop(message.id, None)
            return

        # ── Purge messages (MUST come before delete channel check) ───────────
        purge_cmd = parse_purge_command(user_text)
        if purge_cmd:
            mode        = purge_cmd["mode"]
            target_cid  = purge_cmd["channel_id"] or message.channel.id
            amount      = purge_cmd["amount"]
            keyword     = purge_cmd["keyword"]
            pu_user_ids = purge_cmd["user_ids"]
            pu_role_id  = purge_cmd["role_id"]
            guild       = message.guild
            try:
                target_ch = await bot.fetch_channel(target_cid)
            except Exception:
                target_ch = message.channel

            # Build confirmation embed
            if mode == "count":
                summary = f"Delete the last **{amount}** messages in {target_ch.mention}"
            elif mode == "keyword":
                summary = f"Delete all messages containing **\"{keyword}\"** in {target_ch.mention}"
            elif mode == "user":
                user_tags = " ".join(f"<@{uid}>" for uid in pu_user_ids)
                summary = f"Delete all messages from {user_tags} in {target_ch.mention}"
            elif mode == "role":
                summary = f"Delete all messages from <@&{pu_role_id}> members in {target_ch.mention}"
            elif mode == "all":
                summary = f"Delete **ALL** messages in {target_ch.mention} — no matter how old"
            else:
                await message.reply("❌ Couldn't understand that purge command.", mention_author=True)
                return

            embed = discord.Embed(
                title="⚠️ Confirm Message Purge",
                description=f"{summary}\n\n**This cannot be undone.**\n\nReply `yes` to confirm or `no` to cancel.",
                color=EMBED_COLORS["urgent"]
            )
            embed.set_footer(text=f"Requested by {message.author.display_name} • Expires in 30 seconds")
            confirm_bot_msg = await message.reply(embed=embed, mention_author=True)

            # IDs of confirmation-related messages to clean up after
            # NOTE: message.id (admin's command) is intentionally excluded — it stays
            cleanup_ids = {
                confirm_bot_msg.id,   # bot's confirmation request
            }

            def purge_check(m):
                return (
                    m.author.id == message.author.id
                    and m.channel.id == message.channel.id
                    and m.content.lower().strip() in ("yes", "no", "y", "n")
                )

            try:
                _awaiting_confirmation[message.channel.id] = message.author.id
                confirm = await bot.wait_for("message", check=purge_check, timeout=30.0)
                cleanup_ids.add(confirm.id)  # admin's yes/no reply

                if confirm.content.lower().strip() not in ("yes", "y"):
                    cancelled_msg = await message.channel.send("❌ Purge cancelled.")
                    cleanup_ids.add(cancelled_msg.id)
                    await asyncio.sleep(2)
                    for mid in cleanup_ids:
                        try:
                            m = await message.channel.fetch_message(mid)
                            await m.delete()
                        except Exception:
                            pass
                    return
            except asyncio.TimeoutError:
                timeout_msg = await message.channel.send("⏱️ Confirmation timed out. Purge cancelled.")
                cleanup_ids.add(timeout_msg.id)
                await asyncio.sleep(2)
                for mid in cleanup_ids:
                    try:
                        m = await message.channel.fetch_message(mid)
                        await m.delete()
                    except Exception:
                        pass
                return
            finally:
                _awaiting_confirmation.pop(message.channel.id, None)

            # Execute the purge in a background task so the bot stays responsive
            async def do_purge():
                deleted_count = 0

                async def safe_delete(msg):
                    """Delete one message with rate-limit retry."""
                    for _ in range(5):
                        try:
                            await msg.delete()
                            return True
                        except discord.HTTPException as e:
                            if e.status == 429:
                                wait = getattr(e, "retry_after", 2.0) or 2.0
                                await asyncio.sleep(float(wait) + 0.5)
                            elif e.status == 404:
                                return True   # already deleted
                            else:
                                return False
                        except Exception:
                            return False
                    return False

                try:
                    if mode == "count":
                        # Exclude confirmation messages from the count by filtering
                        def not_cleanup(m): return m.id not in cleanup_ids
                        deleted = await target_ch.purge(limit=amount, check=not_cleanup)
                        deleted_count = len(deleted)

                    elif mode == "keyword":
                        def kw_check(m): return keyword.lower() in m.content.lower() and m.id not in cleanup_ids
                        deleted = await target_ch.purge(limit=None, check=kw_check)
                        deleted_count = len(deleted)
                        async for old_msg in target_ch.history(limit=None):
                            if old_msg.id in cleanup_ids:
                                continue
                            if keyword.lower() in old_msg.content.lower():
                                if await safe_delete(old_msg):
                                    deleted_count += 1
                                    await asyncio.sleep(1.2)

                    elif mode == "user":
                        uid_set = set(pu_user_ids)
                        def user_check(m): return m.author.id in uid_set and m.id not in cleanup_ids
                        deleted = await target_ch.purge(limit=None, check=user_check)
                        deleted_count = len(deleted)
                        async for old_msg in target_ch.history(limit=None):
                            if old_msg.id in cleanup_ids:
                                continue
                            if old_msg.author.id in uid_set:
                                if await safe_delete(old_msg):
                                    deleted_count += 1
                                    await asyncio.sleep(1.2)

                    elif mode == "role":
                        role_obj = guild.get_role(pu_role_id)
                        role_member_ids = {m.id for m in role_obj.members} if role_obj else set()
                        def role_check(m): return m.author.id in role_member_ids and m.id not in cleanup_ids
                        deleted = await target_ch.purge(limit=None, check=role_check)
                        deleted_count = len(deleted)
                        async for old_msg in target_ch.history(limit=None):
                            if old_msg.id in cleanup_ids:
                                continue
                            if old_msg.author.id in role_member_ids:
                                if await safe_delete(old_msg):
                                    deleted_count += 1
                                    await asyncio.sleep(1.2)

                    elif mode == "all":
                        def not_cleanup_check(m): return m.id not in cleanup_ids
                        deleted = await target_ch.purge(limit=None, check=not_cleanup_check)
                        deleted_count = len(deleted)
                        async for old_msg in target_ch.history(limit=None):
                            if old_msg.id in cleanup_ids:
                                continue
                            if await safe_delete(old_msg):
                                deleted_count += 1
                                await asyncio.sleep(1.2)

                    # Send result — stays permanently so admin has a record
                    result_embed = discord.Embed(
                        description=f"✅ **{deleted_count}** message(s) deleted from {target_ch.mention}.",
                        color=EMBED_COLORS["general"]
                    )
                    result_embed.set_footer(text=f"Action by {message.author.display_name}")
                    await message.channel.send(embed=result_embed)

                    # Clean up only the confirmation prompt and the user's yes/no reply
                    for mid in list(cleanup_ids):
                        try:
                            m = await message.channel.fetch_message(mid)
                            await m.delete()
                        except Exception:
                            pass

                except discord.Forbidden:
                    await message.channel.send("❌ I don't have permission to delete messages in that channel.")
                except Exception as e:
                    await message.channel.send(f"⚠️ Something went wrong during purge: {e}")

            asyncio.ensure_future(do_purge())
            return

    # ── End server management ─────────────────────────────────────────────────

    # ── Handle attachments ───────────────────────────────────────────────────
    if message.attachments:
        attachment = message.attachments[0]
        att_type   = get_attachment_type(attachment.filename)

        # 🖼️ IMAGE — GPT-4o Vision
        if att_type == "image":
            stop = asyncio.Event()
            asyncio.create_task(keep_typing(message.channel, stop))
            needs_support = False
            try:
                image_bytes = await download_attachment(attachment)
                ext = os.path.splitext(attachment.filename)[1].lower().strip(".")
                media_type_map = {
                    "jpg": "image/jpeg", "jpeg": "image/jpeg",
                    "png": "image/png", "gif": "image/gif", "webp": "image/webp"
                }
                media_type = media_type_map.get(ext, "image/jpeg")
                prompt = user_text if user_text else "What is in this image? Give context relevant to AFC or Free Fire if applicable."
                reply, needs_support = await ask_openai_with_image(message.channel.id, prompt, username, image_bytes, media_type, is_staff=is_staff)
            except Exception as e:
                reply = resolve_ai_error_reply(message.channel.id, e, force=is_mentioned)
            finally:
                stop.set()
            if reply is not None:
                if not reply.strip() and not needs_support:
                    reply = GENERIC_ERROR_NOTICE
                if reply.strip():
                    await _reply_chunked(message, reply)
                _mark_handled(message.id)
                if needs_support:
                    await send_support_redirect(message)
            return

        # 🎵 AUDIO — Whisper transcription → GPT-4o reply
        elif att_type == "audio":
            try:
                await message.reply("🎵 Got your audio! Give me a sec to listen...", mention_author=True)
                audio_bytes = await download_attachment(attachment)
                transcript  = await transcribe_audio(audio_bytes, attachment.filename)

                if not transcript.strip():
                    await message.reply("⚠️ I couldn't make out what was said. Try sending a clearer recording.", mention_author=True)
                    return

                combined = f"[Audio message transcribed]: {transcript}"
                if user_text:
                    combined += f"\n[User also typed]: {user_text}"

                stop = asyncio.Event()
                asyncio.create_task(keep_typing(message.channel, stop))
                try:
                    reply, needs_support = await ask_openai_text(message.channel.id, combined, username, is_staff=is_staff)
                finally:
                    stop.set()
                if not reply.strip():
                    reply = GENERIC_ERROR_NOTICE
                # Cap the quoted transcript — a long voice note plus the reply
                # would exceed Discord's 2000-char content limit (the full text
                # already went to the model via `combined` above).
                quoted = transcript if len(transcript) <= 700 else transcript[:700] + "…"
                await _reply_chunked(message, f"🎙️ **I heard:** _{quoted}_\n\n{reply}")
                _mark_handled(message.id)
                if needs_support:
                    await send_support_redirect(message)
            except Exception as e:
                # Audio path already posted a "give me a sec" ack — always deliver a
                # resolution (force) so we never leave the user hanging after that.
                notice = resolve_ai_error_reply(message.channel.id, e, force=True)
                if notice is not None:
                    await message.reply(notice, mention_author=True)
            return

        # 🎥 VIDEO — Acknowledge, can't analyze
        elif att_type == "video":
            context = f"{username} sent a video called '{attachment.filename}'. " \
                      f"Acknowledge you received it but explain warmly that you can't watch or analyze videos yet. " \
                      f"They also said: '{user_text}'. Reply naturally and in character as AFC Bot."
            stop = asyncio.Event()
            asyncio.create_task(keep_typing(message.channel, stop))
            needs_support = False
            try:
                reply, needs_support = await ask_openai_text(message.channel.id, context, username, is_staff=is_staff)
            except Exception as e:
                reply = resolve_ai_error_reply(message.channel.id, e, force=is_mentioned)
            finally:
                stop.set()
            if reply is not None:
                if not reply.strip() and not needs_support:
                    reply = GENERIC_ERROR_NOTICE
                if reply.strip():
                    await _reply_chunked(message, reply)
                _mark_handled(message.id)
                if needs_support:
                    await send_support_redirect(message)
            return

        # ❓ Unknown file type
        else:
            stop = asyncio.Event()
            asyncio.create_task(keep_typing(message.channel, stop))
            needs_support = False
            try:
                reply, needs_support = await ask_openai_text(
                    message.channel.id,
                    f"{username} sent a file called '{attachment.filename}'. {user_text}",
                    username,
                    is_staff=is_staff,
                )
            except Exception as e:
                reply = resolve_ai_error_reply(message.channel.id, e, force=is_mentioned)
            finally:
                stop.set()
            if reply is not None:
                if not reply.strip() and not needs_support:
                    reply = GENERIC_ERROR_NOTICE
                if reply.strip():
                    await _reply_chunked(message, reply)
                _mark_handled(message.id)
                if needs_support:
                    await send_support_redirect(message)
            return

    # ── Standard text reply ──────────────────────────────────────────────────
    if not user_text:
        user_text = "Hello!"

    reply_context = await get_reply_context(message)
    if reply_context:
        user_text = reply_context + user_text

    stop = asyncio.Event()
    asyncio.create_task(keep_typing(message.channel, stop))
    needs_support = False
    try:
        reply, needs_support = await ask_openai_text(message.channel.id, user_text, username, is_staff=is_staff)
    except Exception as exc:
        reply = resolve_ai_error_reply(message.channel.id, exc, force=is_mentioned)
    finally:
        stop.set()

    # reply is None only when the AI is down and we already notified this channel
    # recently — stay quiet instead of spamming an identical down notice.
    if reply is None:
        return

    if not reply.strip():
        # Marker-only output (the model sent just the redirect marker) or an
        # empty completion — message.reply("") would 400, and an escalation
        # must never be dropped.
        if needs_support:
            await send_support_redirect(message)
            _mark_handled(message.id)
            return
        reply = GENERIC_ERROR_NOTICE

    sent_reply = await _reply_chunked(message, reply)
    if sent_reply:
        last_bot_messages[message.channel.id] = sent_reply.id
    _mark_handled(message.id)
    if needs_support:
        await send_support_redirect(message)


if __name__ == "__main__":
    bot.run(DISCORD_TOKEN)
